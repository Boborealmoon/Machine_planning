"""Tests for First Article Tracker (Archive)."""
from __future__ import annotations

import io
import os
import unittest
from datetime import date
from unittest.mock import patch

from app import app
from planning.first_article_service import (
    diff_tracked_fields,
    flag_process_sheets,
    flatten_sales_order_jobs,
    history_text,
    job_from_sales_order_pp,
    json_error,
    list_flag_candidates,
    list_new_part_rows,
    _is_complete_status,
    lookup_sales_order_job,
    parse_npi_import_workbook,
    search_flag_candidates,
    search_jobs,
    _ensure_tables,
    _job_from_pp_cache_row,
    _live_job_map,
    _merge_live_job,
    _merge_new_part_row,
    _parse_check_cell,
    _parse_machine_codes,
    _parse_pic_names,
    _resolve_machine_codes,
    _serialize_tracker_row,
    build_import_template_bytes,
)


def _pp(**overrides):
    row = {
        "process_sheet_no": "APS-1001",
        "pp_voucher_no": "PP-1001",
        "inventory_code": "8816-01",
        "description": "Valve body",
        "pp_qty": 12,
        "due_date": "2026-09-01",
        "queued_machines": ["CNC-01"],
        "coway_proposed_edd": "",
        "partials": [],
    }
    row.update(overrides)
    return row


class FirstArticleServiceTests(unittest.TestCase):
    def test_job_uses_coway_edd_from_partial_when_header_blank(self):
        pp = _pp(
            partials=[
                {"pp_partial_no": 1, "coway_proposed_edd": "2026-08-20"},
                {"pp_partial_no": 2, "coway_proposed_edd": "2026-08-22"},
            ]
        )
        job = job_from_sales_order_pp({"sales_order_no": "SO-9", "customer_name": "Coway"}, pp)
        self.assertEqual(job["coway_proposed_edd"], "2026-08-20")
        self.assertEqual(job["part_no"], "8816-01")
        self.assertEqual(job["total_qty"], 12)
        self.assertEqual(job["po_due_date"], "2026-09-01")
        self.assertEqual(job["machine_cnc"], "CNC-01")
        self.assertEqual(job["posted_date"], "")
        self.assertFalse(job["is_new_part"])
        self.assertEqual(job["ps_type"], "APS")
        self.assertEqual(job["bom_code"], "")
        self.assertFalse(job["has_bom"])
        self.assertEqual(job["so_scope"], "active")
        self.assertEqual(job["erp_stage_mode"], "unassigned")

    def test_job_copies_posted_date_material_and_new_flag(self):
        pp = _pp(
            is_new_part=True,
            order_date="2026-08-12",
            material_subcon="ARRIVED",
            bom_code="BOM-1",
        )
        job = job_from_sales_order_pp(
            {
                "sales_order_no": "SO-9",
                "customer_name": "Coway",
                "first_posted_datetime": "2026-08-11 09:30:00",
            },
            pp,
        )
        self.assertTrue(job["is_new_part"])
        self.assertEqual(job["posted_date"], "2026-08-11")
        self.assertTrue(job["material_arrived"])
        self.assertEqual(job["material_display"], "Arrived")
        self.assertEqual(job["bom_code"], "BOM-1")
        self.assertFalse(job["has_bom"])

    def test_placeholder_bom_is_copied_without_implying_materials(self):
        job = job_from_sales_order_pp({"sales_order_no": "SO-9"}, _pp(bom_code="PLACEHOLDER"))
        self.assertEqual(job["bom_code"], "PLACEHOLDER")
        self.assertFalse(job["has_bom"])

    def test_job_prefers_header_coway_edd(self):
        pp = _pp(
            coway_proposed_edd="2026-07-15",
            partials=[{"coway_proposed_edd": "2026-08-01"}],
        )
        job = job_from_sales_order_pp({"sales_order_no": "SO-9"}, pp)
        self.assertEqual(job["coway_proposed_edd"], "2026-07-15")

    def test_job_copies_erp_stage_and_complete_scope(self):
        pp = _pp(
            current_stage_desc="Rough turning",
            current_stage_status="I",
            erp_stage_mode="open",
            shipped_completed=True,
        )
        job = job_from_sales_order_pp({"sales_order_no": "SO-9"}, pp, so_scope="complete")
        self.assertEqual(job["current_stage_desc"], "Rough turning")
        self.assertEqual(job["current_stage_status"], "I")
        self.assertEqual(job["erp_stage_mode"], "open")
        self.assertEqual(job["so_scope"], "complete")
        self.assertTrue(job["shipped_completed"])

    def test_parse_machine_codes_dedupes(self):
        self.assertEqual(_parse_machine_codes("CNC 10, cnc 10, CNC 20"), ["CNC 10", "CNC 20"])
        self.assertEqual(_parse_machine_codes(["CNC 38", "CNC 38", ""]), ["CNC 38"])

    def test_serialize_prefers_saved_machines_over_live_queue(self):
        row = {
            "first_article_id": 1,
            "process_sheet_no": "APS-1",
            "pp_voucher_no": "",
            "pic_ids": [],
            "machine_codes": ["CNC 38"],
            "tooling_mode": "tick",
            "tooling_tick": False,
            "tooling_text": "",
            "fixture_mode": "tick",
            "fixture_tick": False,
            "fixture_text": "",
            "gauges_mode": "tick",
            "gauges_tick": False,
            "gauges_text": "",
            "remarks": "",
        }
        live = {
            "queued_machines": ["CNC 01"],
            "machine_cnc": "CNC 01",
            "part_no": "8816-01",
            "so_scope": "active",
        }
        saved = _serialize_tracker_row(row, live=live, pics_by_id={})
        self.assertEqual(saved["machine_codes"], ["CNC 38"])
        self.assertEqual(saved["machine_cnc"], "CNC 38")
        unset = _serialize_tracker_row({**row, "machine_codes": None}, live=live, pics_by_id={})
        self.assertEqual(unset["machine_codes"], ["CNC 01"])
        self.assertEqual(unset["current_stage_desc"], "")

    def test_pp_cache_row_fills_part_fields(self):
        job = _job_from_pp_cache_row({
            "process_sheet_no": "NPS26-0324",
            "part_no": "BBD012702A",
            "part_description": "Housing",
            "total_qty": 20,
            "po_due_date": "2026-09-01",
            "status": "History",
        })
        self.assertEqual(job["process_sheet_no"], "NPS26-0324")
        self.assertEqual(job["part_no"], "BBD012702A")
        self.assertEqual(job["part_description"], "Housing")
        self.assertEqual(job["total_qty"], 20)
        self.assertEqual(job["so_scope"], "complete")
        self.assertTrue(job["from_erp_cache"])

    def test_merge_live_job_fills_blank_part_fields(self):
        primary = job_from_sales_order_pp({"sales_order_no": "SO-1"}, _pp(inventory_code="", description=""))
        extra = _job_from_pp_cache_row({
            "process_sheet_no": "APS-1001",
            "part_no": "8816-01",
            "part_description": "Valve body",
            "total_qty": 12,
        })
        merged = _merge_live_job(primary, extra)
        self.assertEqual(merged["part_no"], "8816-01")
        self.assertEqual(merged["part_description"], "Valve body")
        self.assertEqual(merged["sales_order_no"], "SO-1")

    def test_flatten_complete_jobs_keep_historical_scope(self):
        jobs = flatten_sales_order_jobs(
            [{"sales_order_no": "SO-H", "pp_vouchers": [_pp(process_sheet_no="AP526-0151", pp_voucher_no="PP-H")]}],
            so_scope="complete",
        )
        self.assertEqual(len(jobs), 1)
        self.assertEqual(jobs[0]["so_scope"], "complete")
        self.assertTrue(jobs[0]["shipped_completed"])
        self.assertEqual(jobs[0]["erp_stage_mode"], "completed")

    def test_flatten_dedupes_by_process_sheet(self):
        orders = [
            {
                "sales_order_no": "SO-1",
                "pp_vouchers": [_pp(), _pp(pp_voucher_no="PP-1001B")],
            }
        ]
        jobs = flatten_sales_order_jobs(orders)
        self.assertEqual(len(jobs), 1)
        self.assertEqual(jobs[0]["process_sheet_no"], "APS-1001")

    def test_search_marks_already_flagged(self):
        jobs = flatten_sales_order_jobs(
            [{"sales_order_no": "SO-1", "pp_vouchers": [_pp(), _pp(process_sheet_no="NPS-22", pp_voucher_no="PP-22", inventory_code="AA-1")]}]
        )
        hits = search_jobs(jobs, "APS", flagged_keys={"APS-1001"})
        self.assertEqual(len(hits), 1)
        self.assertTrue(hits[0]["already_flagged"])
        other = search_jobs(jobs, "NPS-22", flagged_keys={"APS-1001"})
        self.assertEqual(len(other), 1)
        self.assertFalse(other[0]["already_flagged"])

    def test_list_flag_candidates_filters_by_ps_type(self):
        jobs = flatten_sales_order_jobs(
            [{
                "sales_order_no": "SO-1",
                "pp_vouchers": [
                    _pp(),
                    _pp(process_sheet_no="NPS26-0374", pp_voucher_no="PP-N1", inventory_code="88D012"),
                    _pp(process_sheet_no="PPS26-9", pp_voucher_no="PP-P1", inventory_code="PP-PART"),
                ],
            }]
        )

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_jobs", return_value=jobs):
            with patch("planning.first_article_service.planner_db") as planner_db:
                planner_db.return_value.__enter__.return_value = FakeCon()
                payload = list_flag_candidates(ps_type_filter="NPS")

        self.assertEqual(payload["matched"], 1)
        self.assertEqual(payload["rows"][0]["process_sheet_no"], "NPS26-0374")
        self.assertEqual(payload["rows"][0]["ps_type"], "NPS")
        types = {item["ps_type"]: item["count"] for item in payload["types"]}
        self.assertEqual(types.get("NPS"), 1)
        self.assertEqual(types.get("APS"), 1)
        self.assertEqual(types.get("PPS"), 1)

    def test_list_flag_candidates_can_filter_historical(self):
        jobs = flatten_sales_order_jobs(
            [{"sales_order_no": "SO-1", "pp_vouchers": [_pp()]}],
            so_scope="active",
        ) + flatten_sales_order_jobs(
            [{"sales_order_no": "SO-H", "pp_vouchers": [_pp(process_sheet_no="AP526-0151", pp_voucher_no="PP-H")]}],
            so_scope="complete",
        )

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_jobs", return_value=jobs):
            with patch("planning.first_article_service.planner_db") as planner_db:
                planner_db.return_value.__enter__.return_value = FakeCon()
                payload = list_flag_candidates(scope_filter="complete")

        self.assertEqual(payload["matched"], 1)
        self.assertEqual(payload["rows"][0]["process_sheet_no"], "AP526-0151")
        self.assertEqual(payload["scope"], "complete")

    def test_search_includes_complete_sales_orders(self):
        payload = {
            "active": [],
            "complete": [{
                "sales_order_no": "SO-H",
                "customer_name": "Coway",
                "pp_vouchers": [_pp(process_sheet_no="AP526-0151", pp_voucher_no="PP-H")],
            }],
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_payload", return_value=payload):
            with patch("planning.first_article_service.planner_db") as planner_db:
                planner_db.return_value.__enter__.return_value = FakeCon()
                with patch("planning.first_article_service._search_jobs_from_pp_cache", return_value=[]):
                    hits = search_flag_candidates("AP526")

        self.assertEqual(len(hits), 1)
        self.assertEqual(hits[0]["process_sheet_no"], "AP526-0151")
        self.assertEqual(hits[0]["so_scope"], "complete")

    def test_flag_process_sheets_rejects_empty(self):
        with self.assertRaisesRegex(ValueError, "at least one"):
            flag_process_sheets([])
        with self.assertRaisesRegex(ValueError, "at least one"):
            flag_process_sheets([{"process_sheet_no": "  "}])

    def test_live_job_map_uses_stale_cache_without_erp_rebuild(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "customer_name": "Coway",
                    "pp_vouchers": [_pp()],
                }
            ]
        }
        with patch("planning.erp_route_cache.get", return_value=payload):
            with patch("planning.sales_orders_route._fetch_sales_orders") as fetch:
                mapping = _live_job_map()
                fetch.assert_not_called()
        self.assertIn("APS-1001", mapping)
        self.assertEqual(mapping["APS-1001"]["part_no"], "8816-01")
        self.assertEqual(mapping["APS-1001"]["coway_proposed_edd"], "")

    def test_list_live_map_does_not_rebuild_when_cache_empty(self):
        with patch("planning.erp_route_cache.get", return_value=None):
            with patch("planning.sales_orders_route._fetch_sales_orders") as fetch:
                mapping = _live_job_map()
                fetch.assert_not_called()
        self.assertEqual(mapping, {})

    def test_search_uses_lite_rebuild_when_cache_empty(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "customer_name": "Coway",
                    "pp_vouchers": [_pp()],
                }
            ]
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.erp_route_cache.get", return_value=None):
            with patch(
                "planning.sales_orders_route._fetch_sales_orders",
                return_value=payload,
            ) as fetch:
                with patch("planning.first_article_service.planner_db") as planner_db:
                    planner_db.return_value.__enter__.return_value = FakeCon()
                    hits = search_flag_candidates("APS")
        fetch.assert_called_once_with(refresh=False, active_only=True, lite=True)
        self.assertEqual(len(hits), 1)
        self.assertEqual(hits[0]["process_sheet_no"], "APS-1001")

    def test_lookup_job_rebuilds_lite_when_cache_empty(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-9",
                    "customer_name": "Coway",
                    "pp_vouchers": [_pp(coway_proposed_edd="2026-08-20")],
                }
            ]
        }
        with patch("planning.erp_route_cache.get", return_value=None):
            with patch(
                "planning.sales_orders_route._fetch_sales_orders",
                return_value=payload,
            ) as fetch:
                job = lookup_sales_order_job("APS-1001")
        fetch.assert_called_once_with(refresh=False, active_only=True, lite=True)
        self.assertEqual(job["coway_proposed_edd"], "2026-08-20")

    def test_list_new_part_rows_filters_to_new_parts(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "first_posted_datetime": "2026-08-12",
                    "pp_vouchers": [
                        _pp(is_new_part=True, material_subcon="2026-09-01", bom_code="SMP-MAT-01"),
                        _pp(
                            process_sheet_no="NPS-22",
                            pp_voucher_no="PP-22",
                            inventory_code="AA-1",
                            is_new_part=False,
                        ),
                    ],
                }
            ]
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_payload", return_value=payload):
            with patch("planning.first_article_service.planner_db") as planner_db:
                with patch(
                    "planning.first_article_service._lookup_parts_with_bom_materials",
                    return_value={"8816-01"},
                ):
                    planner_db.return_value.__enter__.return_value = FakeCon()
                    rows = list_new_part_rows(allow_rebuild=False)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["process_sheet_no"], "APS-1001")
        self.assertTrue(rows[0]["is_new_part"])
        self.assertEqual(rows[0]["ps_type"], "APS")
        self.assertEqual(rows[0]["posted_date"], "2026-08-12")
        self.assertEqual(rows[0]["material_date"], "2026-09-01")
        self.assertFalse(rows[0]["bom_updated"])
        self.assertEqual(rows[0]["bom_code"], "SMP-MAT-01")
        self.assertTrue(rows[0]["has_bom"])
        self.assertEqual(rows[0]["program_pic_ids"], [])
        self.assertEqual(rows[0]["program_pics"], [])
        self.assertFalse(rows[0]["is_exception"])

    def test_list_new_part_rows_has_bom_follows_erp_materials_not_pp_route(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "pp_vouchers": [
                        _pp(is_new_part=True, bom_code=""),
                        _pp(
                            process_sheet_no="NPS-22",
                            pp_voucher_no="PP-22",
                            inventory_code="AA-1",
                            is_new_part=True,
                            bom_code="FLOW-1",
                        ),
                    ],
                }
            ]
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_payload", return_value=payload):
            with patch("planning.first_article_service.planner_db") as planner_db:
                with patch(
                    "planning.first_article_service._lookup_parts_with_bom_materials",
                    return_value={"8816-01"},
                ) as lookup:
                    planner_db.return_value.__enter__.return_value = FakeCon()
                    rows = list_new_part_rows(allow_rebuild=False)

        by_part = {row["part_no"]: row for row in rows}
        self.assertEqual(set(by_part), {"8816-01", "AA-1"})
        self.assertEqual(by_part["8816-01"]["bom_code"], "")
        self.assertTrue(by_part["8816-01"]["has_bom"])
        self.assertEqual(by_part["AA-1"]["bom_code"], "FLOW-1")
        self.assertFalse(by_part["AA-1"]["has_bom"])
        lookup.assert_called_once()

    def test_list_new_part_rows_includes_manual_exceptions(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "first_posted_datetime": "2026-08-12",
                    "pp_vouchers": [
                        _pp(is_new_part=True),
                        _pp(
                            process_sheet_no="NPS-22",
                            pp_voucher_no="PP-22",
                            inventory_code="AA-1",
                            description="Housing",
                            is_new_part=False,
                        ),
                    ],
                }
            ]
        }
        saved = {
            "NPS-22": {
                "process_sheet_no": "NPS-22",
                "pp_voucher_no": "PP-22",
                "bom_updated": False,
                "remarks": "Track anyway",
                "program_finish_at": "",
                "program_pic_ids": [],
                "is_exception": True,
            },
            "PPS26-0001": {
                "process_sheet_no": "PPS26-0001",
                "pp_voucher_no": "",
                "bom_updated": False,
                "remarks": "",
                "program_finish_at": "",
                "program_pic_ids": [],
                "is_exception": True,
            },
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_payload", return_value=payload):
            with patch("planning.first_article_service.planner_db") as planner_db:
                with patch(
                    "planning.first_article_service._lookup_parts_with_bom_materials",
                    return_value=set(),
                ):
                    with patch(
                        "planning.first_article_service._new_part_tracker_map",
                        return_value=saved,
                    ):
                        with patch(
                            "planning.first_article_service._lookup_jobs_from_pp_cache",
                            return_value={},
                        ):
                            with patch(
                                "planning.first_article_service._apply_stage_overlay_to_rows",
                            ):
                                planner_db.return_value.__enter__.return_value = FakeCon()
                                rows = list_new_part_rows(allow_rebuild=False)

        by_ps = {row["process_sheet_no"]: row for row in rows}
        self.assertEqual(set(by_ps), {"APS-1001", "NPS-22", "PPS26-0001"})
        self.assertTrue(by_ps["APS-1001"]["is_new_part"])
        self.assertFalse(by_ps["APS-1001"]["is_exception"])
        self.assertTrue(by_ps["NPS-22"]["is_exception"])
        self.assertFalse(by_ps["NPS-22"]["is_new_part"])
        self.assertEqual(by_ps["NPS-22"]["part_no"], "AA-1")
        self.assertEqual(by_ps["NPS-22"]["remarks"], "Track anyway")
        self.assertTrue(by_ps["PPS26-0001"]["is_exception"])
        self.assertEqual(by_ps["PPS26-0001"]["ps_type"], "PPS")
        self.assertFalse(by_ps["PPS26-0001"]["in_sales_orders"])

    def test_is_complete_status_from_wo_so_or_shipped(self):
        self.assertFalse(_is_complete_status({"so_scope": "active", "erp_stage_mode": "open"}))
        self.assertTrue(_is_complete_status({"shipped_completed": True}))
        self.assertTrue(_is_complete_status({"so_scope": "complete"}))
        self.assertTrue(_is_complete_status({"erp_stage_mode": "completed"}))

    def test_list_new_part_rows_splits_active_and_history(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "first_posted_datetime": "2026-08-12",
                    "pp_vouchers": [
                        _pp(is_new_part=True, remarks="live new"),
                        _pp(
                            process_sheet_no="NPS-DONE",
                            pp_voucher_no="PP-DONE",
                            inventory_code="DONE-1",
                            description="Finished housing",
                            is_new_part=True,
                            shipped_completed=True,
                        ),
                    ],
                }
            ],
            "complete": [
                {
                    "sales_order_no": "SO-H",
                    "first_posted_datetime": "2026-07-01",
                    "pp_vouchers": [
                        _pp(
                            process_sheet_no="MPS-OLD",
                            pp_voucher_no="PP-OLD",
                            inventory_code="OLD-1",
                            description="Old valve",
                            is_new_part=True,
                        ),
                    ],
                }
            ],
        }
        saved = {
            "NPS-22": {
                "process_sheet_no": "NPS-22",
                "pp_voucher_no": "PP-22",
                "bom_updated": False,
                "remarks": "Track anyway",
                "program_finish_at": "",
                "program_pic_ids": [],
                "is_exception": True,
            },
            "NPS-DONE": {
                "process_sheet_no": "NPS-DONE",
                "pp_voucher_no": "PP-DONE",
                "bom_updated": False,
                "remarks": "CAM signed off",
                "program_finish_at": "2026-08-01",
                "program_pic_ids": [],
                "is_exception": False,
            },
            "APS-GONE": {
                "process_sheet_no": "APS-GONE",
                "pp_voucher_no": "",
                "bom_updated": False,
                "remarks": "Kept after S/O dropped",
                "program_finish_at": "2026-06-15",
                "program_pic_ids": [],
                "is_exception": False,
            },
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        with patch("planning.first_article_service._sales_order_payload", return_value=payload):
            with patch("planning.first_article_service.planner_db") as planner_db:
                with patch(
                    "planning.first_article_service._lookup_parts_with_bom_materials",
                    return_value=set(),
                ):
                    with patch(
                        "planning.first_article_service._new_part_tracker_map",
                        return_value=saved,
                    ):
                        with patch(
                            "planning.first_article_service._lookup_jobs_from_pp_cache",
                            return_value={},
                        ):
                            with patch(
                                "planning.first_article_service._apply_stage_overlay_to_rows",
                            ):
                                planner_db.return_value.__enter__.return_value = FakeCon()
                                active = list_new_part_rows(allow_rebuild=False, scope="active")
                                history = list_new_part_rows(allow_rebuild=False, scope="history")

        active_ps = {row["process_sheet_no"] for row in active}
        history_ps = {row["process_sheet_no"] for row in history}
        self.assertEqual(active_ps, {"APS-1001", "NPS-22"})
        self.assertEqual(history_ps, {"NPS-DONE", "MPS-OLD", "APS-GONE"})
        self.assertTrue(all(row["list_scope"] == "active" for row in active))
        self.assertTrue(all(row["list_scope"] == "history" for row in history))
        by_hist = {row["process_sheet_no"]: row for row in history}
        self.assertEqual(by_hist["NPS-DONE"]["remarks"], "CAM signed off")
        self.assertEqual(by_hist["NPS-DONE"]["program_finish_at"], "2026-08-01")
        self.assertEqual(by_hist["NPS-DONE"]["part_no"], "DONE-1")
        self.assertEqual(by_hist["APS-GONE"]["remarks"], "Kept after S/O dropped")
        self.assertEqual(by_hist["APS-GONE"]["program_finish_at"], "2026-06-15")
        self.assertEqual(by_hist["MPS-OLD"]["so_scope"], "complete")

    def test_list_new_part_rows_completed_wo_goes_to_history(self):
        payload = {
            "active": [
                {
                    "sales_order_no": "SO-1",
                    "pp_vouchers": [
                        _pp(
                            process_sheet_no="APS-WO",
                            pp_voucher_no="PP-WO",
                            is_new_part=True,
                            erp_stage_mode="completed",
                        ),
                    ],
                }
            ]
        }

        class FakeCon:
            def execute(self, *args, **kwargs):
                return self

            def fetchall(self):
                return []

            def fetchone(self):
                return None

        def overlay(rows_out):
            for row in rows_out:
                row["erp_stage_mode"] = "completed"

        with patch("planning.first_article_service._sales_order_payload", return_value=payload):
            with patch("planning.first_article_service.planner_db") as planner_db:
                with patch(
                    "planning.first_article_service._lookup_parts_with_bom_materials",
                    return_value=set(),
                ):
                    with patch(
                        "planning.first_article_service._apply_stage_overlay_to_rows",
                        side_effect=overlay,
                    ):
                        planner_db.return_value.__enter__.return_value = FakeCon()
                        active = list_new_part_rows(allow_rebuild=False, scope="active")
                        history = list_new_part_rows(allow_rebuild=False, scope="history")

        self.assertEqual(active, [])
        self.assertEqual([row["process_sheet_no"] for row in history], ["APS-WO"])
        self.assertEqual(history[0]["list_scope"], "history")
        self.assertEqual(history[0]["erp_stage_mode"], "completed")

    def test_parse_pic_names_slash_and_comma(self):
        self.assertEqual(_parse_pic_names("Chang Peng/Anand"), ["Chang Peng", "Anand"])
        self.assertEqual(_parse_pic_names("Chang Peng, Anand"), ["Chang Peng", "Anand"])
        self.assertEqual(_parse_pic_names("  Anand  "), ["Anand"])
        self.assertEqual(_parse_pic_names("None"), [])

    def test_parse_check_cell_ok_date_and_note(self):
        self.assertEqual(_parse_check_cell("OK")["tick"], True)
        self.assertEqual(_parse_check_cell("ok")["text"], "")
        self.assertEqual(_parse_check_cell("NA")["text"], "NA")
        self.assertEqual(_parse_check_cell("Est. Wk 31")["text"], "Est. Wk 31")
        self.assertEqual(_parse_check_cell("31/7/2026")["text"], "2026-07-31")
        self.assertEqual(_parse_check_cell(date(2026, 7, 31))["text"], "2026-07-31")
        self.assertFalse(_parse_check_cell("")["tick"])

    def test_resolve_machine_numbers_against_catalog(self):
        catalog = ["CNC 10", "CNC 22", "CNC 30"]
        self.assertEqual(_resolve_machine_codes("22, 30", catalog), ["CNC 22", "CNC 30"])
        self.assertEqual(_resolve_machine_codes("CNC 10, 22", catalog), ["CNC 10", "CNC 22"])
        self.assertEqual(_resolve_machine_codes("15", catalog), ["CNC 15"])

    def test_history_text_joins_pic_names(self):
        pics = {
            4: {"pic_id": 4, "name": "Ananda"},
            7: {"pic_id": 7, "name": "Chuan Heng"},
        }
        self.assertEqual(history_text("program_pic_ids", [7, 4], pics), "Chuan Heng, Ananda")
        self.assertEqual(history_text("remarks", "  Complex Programme.  ", pics), "Complex Programme.")
        self.assertEqual(history_text("program_finish_at", "26/08/2026", pics), "2026-08-26")

    def test_diff_tracked_fields_skips_unchanged_values(self):
        pics = {4: {"pic_id": 4, "name": "Ananda"}}
        changes = diff_tracked_fields(
            {"remarks": "Same", "program_finish_at": "2026-08-26", "program_pic_ids": [4]},
            {"remarks": "Same", "program_finish_at": "26/08/2026", "program_pic_ids": [4]},
            ("remarks", "program_finish_at", "program_pic_ids"),
            pics_by_id=pics,
        )
        self.assertEqual(changes, [])
        changed = diff_tracked_fields(
            {"remarks": "", "program_finish_at": "", "program_pic_ids": []},
            {"remarks": "2 weeks for tooling", "program_finish_at": "2026-09-01", "program_pic_ids": [4]},
            ("remarks", "program_finish_at", "program_pic_ids"),
            pics_by_id=pics,
        )
        by_field = {item["field_name"]: item for item in changed}
        self.assertEqual(by_field["remarks"]["new_value"], "2 weeks for tooling")
        self.assertEqual(by_field["program_finish_at"]["old_value"], "")
        self.assertEqual(by_field["program_finish_at"]["new_value"], "2026-09-01")
        self.assertEqual(by_field["program_pic_ids"]["new_value"], "Ananda")
        self.assertEqual(by_field["program_pic_ids"]["field_label"], "Programme PIC")

    def test_merge_new_part_row_exposes_wo_stage(self):
        job = job_from_sales_order_pp(
            {"sales_order_no": "SO-9"},
            _pp(
                is_new_part=True,
                current_stage_desc="CNC MILLING",
                current_stage_status="I",
                erp_stage_mode="open",
            ),
        )
        merged = _merge_new_part_row(job, None, pics_by_id={})
        self.assertEqual(merged["current_stage_desc"], "CNC MILLING")
        self.assertEqual(merged["current_stage_status"], "I")
        self.assertEqual(merged["current_stage_status_label"], "In process")
        self.assertEqual(merged["history_count"], 0)

    def test_parse_npi_workbook_maps_template_headers(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append([
            "Data Input", "Part No.", "Part Description", "Total Qty", "PO Due Date",
            "Machine (CNC)", "PIC", "Tooling", "Fixture/Jig", "Gauges/CMM", "Remark",
        ])
        ws.append([
            "NPS26-0321", "BB18-KS", "BRACKET", 30, "20/8/2026",
            "22, 30", "Chang Peng/Anand", "OK", "Est. Wk 32", "NA", "Material in",
        ])
        ws.append(["", "skip", "", "", "", "", "", "", "", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        items = parse_npi_import_workbook(buf.getvalue(), "npi.xlsx")
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["process_sheet_no"], "NPS26-0321")
        patch = items[0]["patch"]
        self.assertEqual(patch["machine_codes"], "22, 30")
        self.assertEqual(patch["pic_names"], "Chang Peng/Anand")
        self.assertEqual(patch["tooling"], "OK")
        self.assertEqual(patch["fixture"], "Est. Wk 32")
        self.assertEqual(patch["gauges"], "NA")
        self.assertEqual(patch["remarks"], "Material in")
        self.assertNotIn("part_no", patch)

    def test_import_template_has_expected_headers(self):
        from openpyxl import load_workbook

        payload = build_import_template_bytes()
        wb = load_workbook(io.BytesIO(payload))
        headers = [cell.value for cell in next(wb.active.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers[0], "Data Input")
        self.assertEqual(headers[6], "PIC")
        self.assertEqual(headers[10], "Remark")
        self.assertIn("Notes", wb.sheetnames)


class FirstArticleRouteTests(unittest.TestCase):
    def setUp(self):
        self.app = app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

    def test_flag_requires_process_sheet(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.post("/api/first-article", json={})

        self.assertEqual(response.status_code, 400)
        self.assertIn("process_sheet_no", response.get_json()["error"])

    def test_patch_without_fields_rejected(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.patch("/api/first-article/1", json={})

        self.assertEqual(response.status_code, 400)
        self.assertIn("No editable fields", response.get_json()["error"])

    def test_pic_add_requires_name(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.post("/api/first-article/pics", json={})

        self.assertEqual(response.status_code, 400)
        self.assertIn("PIC name", response.get_json()["error"])

    def test_flag_list_patch_delete(self):
        created = {
            "first_article_id": 3,
            "process_sheet_no": "APS-1001",
            "part_no": "8816-01",
            "coway_proposed_edd": "2026-08-20",
            "remarks": "",
        }
        patched = {**created, "remarks": "FA in progress", "tooling_mode": "text"}

        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.flag_process_sheet",
                return_value=(created, True),
            ) as flag_fn:
                create_res = self.client.post(
                    "/api/first-article",
                    json={"process_sheet_no": "APS-1001"},
                )
            with patch("planning.first_article_route.load_pics", return_value=[]):
                with patch(
                    "planning.first_article_route.list_tracker_rows",
                    return_value=[created],
                ):
                    with patch("planning.first_article_route.planner_db") as planner_db:
                        planner_db.return_value.__enter__.return_value = object()
                        list_res = self.client.get("/api/first-article")
            with patch(
                "planning.first_article_route.update_tracker_row",
                return_value=patched,
            ) as update_fn:
                patch_res = self.client.patch(
                    "/api/first-article/3",
                    json={"remarks": "FA in progress", "tooling_mode": "text"},
                )
            with patch(
                "planning.first_article_route.unflag_process_sheet",
                return_value=True,
            ) as unflag_fn:
                delete_res = self.client.delete("/api/first-article/3")

        self.assertEqual(create_res.status_code, 201)
        self.assertEqual(create_res.get_json()["row"]["process_sheet_no"], "APS-1001")
        flag_fn.assert_called_once()
        self.assertEqual(list_res.status_code, 200)
        self.assertEqual(list_res.get_json()["count"], 1)
        self.assertEqual(patch_res.status_code, 200)
        self.assertEqual(patch_res.get_json()["row"]["remarks"], "FA in progress")
        update_fn.assert_called_once()
        self.assertEqual(delete_res.status_code, 200)
        unflag_fn.assert_called_once_with(3)

    def test_invalid_check_mode_rejected(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.update_tracker_row",
                side_effect=ValueError("tooling_mode must be tick or text"),
            ):
                response = self.client.patch(
                    "/api/first-article/3",
                    json={"tooling_mode": "maybe"},
                )

        self.assertEqual(response.status_code, 400)
        self.assertIn("tick or text", response.get_json()["error"])

    def test_candidates_lists_active_jobs(self):
        payload = {
            "rows": [{
                "process_sheet_no": "NPS26-0374",
                "ps_type": "NPS",
                "already_flagged": False,
            }],
            "types": [{"ps_type": "NPS", "count": 1}],
            "total": 1,
            "matched": 1,
            "truncated": False,
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.list_flag_candidates",
                return_value=payload,
            ) as list_fn:
                response = self.client.get("/api/first-article/candidates?ps_type=NPS")

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(body["matched"], 1)
        self.assertEqual(body["rows"][0]["process_sheet_no"], "NPS26-0374")
        list_fn.assert_called_once()
        self.assertEqual(list_fn.call_args.kwargs.get("ps_type_filter"), "NPS")

    def test_bulk_flag_requires_items(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.post("/api/first-article/bulk", json={})

        self.assertEqual(response.status_code, 400)
        self.assertIn("items", response.get_json()["error"])

    def test_bulk_flag_posts_selected_sheets(self):
        result = {
            "created": [
                {"first_article_id": 1, "process_sheet_no": "NPS26-0374"},
                {"first_article_id": 2, "process_sheet_no": "NPS26-0375"},
            ],
            "already_flagged": [],
            "created_count": 2,
            "already_flagged_count": 0,
            "count": 2,
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.flag_process_sheets",
                return_value=result,
            ) as flag_fn:
                response = self.client.post(
                    "/api/first-article/bulk",
                    json={
                        "items": [
                            {"process_sheet_no": "NPS26-0374"},
                            {"process_sheet_no": "NPS26-0375"},
                        ]
                    },
                )

        self.assertEqual(response.status_code, 201)
        body = response.get_json()
        self.assertEqual(body["created_count"], 2)
        flag_fn.assert_called_once()
        items = flag_fn.call_args.args[0]
        self.assertEqual(len(items), 2)
        self.assertEqual(items[0]["process_sheet_no"], "NPS26-0374")

    def test_new_parts_list_returns_rows(self):
        rows = [{
            "process_sheet_no": "NPS26-0374",
            "part_no": "BBD012702A REV 02",
            "is_new_part": True,
            "posted_date": "2026-08-12",
            "bom_updated": False,
        }]
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.list_new_part_rows",
                return_value=rows,
            ) as list_fn:
                response = self.client.get("/api/first-article/new-parts")

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(body["count"], 1)
        self.assertEqual(body["rows"][0]["process_sheet_no"], "NPS26-0374")
        self.assertEqual(body["scope"], "active")
        list_fn.assert_called_once_with(scope="active")

    def test_new_parts_list_history_scope(self):
        rows = [{
            "process_sheet_no": "NPS-DONE",
            "part_no": "DONE-1",
            "list_scope": "history",
            "remarks": "CAM signed off",
        }]
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.list_new_part_rows",
                return_value=rows,
            ) as list_fn:
                response = self.client.get("/api/first-article/new-parts?scope=history")

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(body["scope"], "history")
        self.assertEqual(body["count"], 1)
        self.assertEqual(body["rows"][0]["remarks"], "CAM signed off")
        list_fn.assert_called_once_with(scope="history")

    def test_new_parts_patch_requires_process_sheet(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.patch("/api/first-article/new-parts", json={"bom_updated": True})

        self.assertEqual(response.status_code, 400)
        self.assertIn("process_sheet_no", response.get_json()["error"])

    def test_new_parts_patch_saves_fields(self):
        saved = {
            "process_sheet_no": "NPS26-0374",
            "bom_updated": True,
            "remarks": "Waiting CAM",
            "program_finish_at": "2026-09-01T16:00",
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.update_new_part_row",
                return_value=saved,
            ) as update_fn:
                response = self.client.patch(
                    "/api/first-article/new-parts",
                    json={
                        "process_sheet_no": "NPS26-0374",
                        "bom_updated": True,
                        "remarks": "Waiting CAM",
                        "program_finish_at": "2026-09-01T16:00",
                        "program_pic_ids": [4],
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["row"]["bom_updated"], True)
        update_fn.assert_called_once()
        payload = update_fn.call_args.args[0]
        self.assertEqual(payload["process_sheet_no"], "NPS26-0374")
        self.assertTrue(payload["bom_updated"])
        self.assertEqual(payload["program_finish_at"], "2026-09-01T16:00")
        self.assertEqual(payload["program_pic_ids"], [4])

    def test_new_parts_exception_requires_process_sheet(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.post("/api/first-article/new-parts", json={})

        self.assertEqual(response.status_code, 400)
        self.assertIn("process_sheet_no", response.get_json()["error"])

    def test_new_parts_exception_adds_row(self):
        saved = {
            "process_sheet_no": "NPS26-0400",
            "part_no": "BB18-01",
            "is_exception": True,
            "is_new_part": False,
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.add_new_part_exception",
                return_value=(saved, True),
            ) as add_fn:
                response = self.client.post(
                    "/api/first-article/new-parts",
                    json={"process_sheet_no": "NPS26-0400"},
                )

        self.assertEqual(response.status_code, 201)
        body = response.get_json()
        self.assertTrue(body["created"])
        self.assertFalse(body["already_on_list"])
        self.assertTrue(body["row"]["is_exception"])
        add_fn.assert_called_once()
        self.assertEqual(add_fn.call_args.args[0]["process_sheet_no"], "NPS26-0400")

    def test_new_parts_exception_already_on_list(self):
        saved = {
            "process_sheet_no": "APS-1001",
            "is_exception": False,
            "is_new_part": True,
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.add_new_part_exception",
                return_value=(saved, False),
            ):
                response = self.client.post(
                    "/api/first-article/new-parts",
                    json={"process_sheet_no": "APS-1001"},
                )

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertFalse(body["created"])
        self.assertTrue(body["already_on_list"])

    def test_new_parts_exception_remove(self):
        result = {
            "process_sheet_no": "NPS26-0400",
            "still_on_list": False,
            "row": None,
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.remove_new_part_exception",
                return_value=result,
            ) as remove_fn:
                response = self.client.delete("/api/first-article/new-parts/NPS26-0400")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["process_sheet_no"], "NPS26-0400")
        self.assertFalse(response.get_json()["still_on_list"])
        remove_fn.assert_called_once_with("NPS26-0400")

    def test_new_parts_exception_remove_missing(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.remove_new_part_exception",
                return_value=None,
            ):
                response = self.client.delete("/api/first-article/new-parts/NPS26-0400")

        self.assertEqual(response.status_code, 404)

    def test_patch_accepts_machine_codes(self):
        saved = {
            "first_article_id": 3,
            "process_sheet_no": "APS-1001",
            "machine_codes": ["CNC 10", "CNC 20"],
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.update_tracker_row",
                return_value=saved,
            ) as update_fn:
                response = self.client.patch(
                    "/api/first-article/3",
                    json={"machine_codes": ["CNC 10", "CNC 20"]},
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["row"]["machine_codes"], ["CNC 10", "CNC 20"])
        update_fn.assert_called_once()
        self.assertEqual(update_fn.call_args.args[1]["machine_codes"], ["CNC 10", "CNC 20"])

    def test_patch_accepts_pic_names_and_check_text(self):
        saved = {
            "first_article_id": 3,
            "process_sheet_no": "APS-1001",
            "pic_ids": [4, 5],
            "tooling_tick": True,
        }
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.update_tracker_row",
                return_value=saved,
            ) as update_fn:
                response = self.client.patch(
                    "/api/first-article/3",
                    json={"pic_names": "Chang Peng/Anand", "tooling": "OK"},
                )

        self.assertEqual(response.status_code, 200)
        payload = update_fn.call_args.args[1]
        self.assertEqual(payload["pic_names"], "Chang Peng/Anand")
        self.assertEqual(payload["tooling"], "OK")

    def test_import_requires_file(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.post("/api/first-article/import")

        self.assertEqual(response.status_code, 400)
        self.assertIn("Excel", response.get_json()["error"])

    def test_import_posts_workbook(self):
        result = {
            "created": [{"first_article_id": 1, "process_sheet_no": "NPS26-0321"}],
            "updated": [],
            "created_count": 1,
            "updated_count": 0,
            "missing_erp": [],
            "missing_erp_count": 0,
            "count": 1,
            "pics": [],
        }
        items = [{"process_sheet_no": "NPS26-0321", "patch": {"tooling": "OK"}}]
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.parse_npi_import_workbook",
                return_value=items,
            ) as parse_fn:
                with patch(
                    "planning.first_article_route.import_tracker_rows",
                    return_value=result,
                ) as import_fn:
                    response = self.client.post(
                        "/api/first-article/import",
                        data={"file": (io.BytesIO(b"fake-xlsx"), "npi.xlsx")},
                        content_type="multipart/form-data",
                    )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.get_json()["created_count"], 1)
        parse_fn.assert_called_once()
        import_fn.assert_called_once_with(items)

    def test_import_template_downloads_xlsx(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.get("/api/first-article/import-template")

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "spreadsheetml",
            response.headers.get("Content-Type", ""),
        )
        self.assertTrue((response.data or b"").startswith(b"PK"))

    def test_page_lives_under_ops(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.get("/ops/first-article")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("NPI/FA Management", html)
        self.assertIn("NPI Tracker", html)
        self.assertIn("id=\"fa-flag-one\"", html)
        self.assertIn("id=\"fa-import-excel\"", html)
        self.assertIn("fa-bulk-scopes", html)
        self.assertIn("id=\"fa-new-exception-search\"", html)
        self.assertIn("Add exception", html)
        self.assertIn("WO / Stage", html)
        self.assertIn("id=\"fa-tab-history\"", html)
        self.assertIn("id=\"fa-panel-history\"", html)
        self.assertIn("id=\"fa-history-table-body\"", html)
        self.assertIn("id=\"fa-history-modal\"", html)
        self.assertNotIn("First Article Tracker", html)
        self.assertNotIn("Flagged jobs", html)

    def test_archive_url_redirects_to_ops(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.get("/archive/first-article", follow_redirects=False)

        self.assertIn(response.status_code, (301, 302, 303, 307, 308))
        self.assertEqual(response.headers.get("Location"), "/ops/first-article")

    def test_history_requires_process_sheet(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            response = self.client.get("/api/first-article/history?source=new_part")

        self.assertEqual(response.status_code, 400)
        self.assertIn("process_sheet_no", response.get_json()["error"])

    def test_history_lists_changes(self):
        rows = [{
            "change_id": 1,
            "source": "new_part",
            "process_sheet_no": "NPS26-8391",
            "field_name": "remarks",
            "field_label": "Remarks",
            "old_value": "",
            "new_value": "2 weeks for tooling",
            "changed_at": "2026-08-26 10:14:00",
        }]
        with patch.dict(os.environ, {"PLANNER_PASSCODE": "", "ADMIN_PASSCODE": ""}):
            with patch(
                "planning.first_article_route.list_change_history",
                return_value=rows,
            ) as list_fn:
                response = self.client.get(
                    "/api/first-article/history?source=new_part&process_sheet_no=NPS26-8391"
                )

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(body["count"], 1)
        self.assertEqual(body["rows"][0]["new_value"], "2 weeks for tooling")
        list_fn.assert_called_once()
        self.assertEqual(list_fn.call_args.kwargs["source"], "new_part")
        self.assertEqual(list_fn.call_args.kwargs["process_sheet_no"], "NPS26-8391")


class FirstArticleSchemaGuardTests(unittest.TestCase):
    def test_ensure_tables_skips_ddl_when_already_ready(self):
        import planning.first_article_service as svc

        class Boom:
            def execute(self, *args, **kwargs):
                raise AssertionError("schema DDL should be skipped")

        previous = svc._tables_ready
        try:
            svc._tables_ready = True
            _ensure_tables(Boom())
        finally:
            svc._tables_ready = previous

    def test_ensure_tables_skips_alter_when_schema_exists(self):
        import planning.first_article_service as svc

        sql = []

        class Con:
            def execute(self, query, params=None):
                sql.append(" ".join(str(query).split()))
                return self

            def fetchall(self):
                return [
                    {"table_name": table, "column_name": column}
                    for table, column in svc._REQUIRED_COLUMNS
                ]

            def fetchone(self):
                return {"ok": 1}

            def commit(self):
                raise AssertionError("should not commit when schema already complete")

        previous = svc._tables_ready
        try:
            svc._tables_ready = False
            _ensure_tables(Con())
            self.assertTrue(svc._tables_ready)
            self.assertTrue(any("information_schema.columns" in query for query in sql))
            self.assertFalse(any("ALTER TABLE" in query for query in sql))
            self.assertFalse(any("CREATE TABLE" in query for query in sql))
        finally:
            svc._tables_ready = previous

    def test_json_error_maps_deadlock_to_retry(self):
        payload, status = json_error(
            RuntimeError("deadlock detected"),
            fallback_status=502,
        )
        self.assertEqual(status, 503)
        self.assertIn("Refresh", payload["error"])
