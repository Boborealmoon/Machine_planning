"""RFQ checker mapping, cycle-time math, Excel ingest, and routes."""
from __future__ import annotations

import io
import os
import unittest
from unittest.mock import Mock, patch

from openpyxl import Workbook

from app import app
from planning.rfq_checker_service import (
    GROQ_DEFAULT_BASE_URL,
    GROQ_DEFAULT_MODEL,
    apply_column_map,
    apply_defaults_to_mapped_lines,
    build_mapped_lines,
    calculate_times,
    format_lead_time,
    group_archive_batches,
    headers_from_source_rows,
    heuristic_column_map,
    heuristic_covers_core_fields,
    infer_cycle_time_from_source,
    invert_field_map,
    list_workbook_sheets,
    llm_status,
    map_columns_with_llm,
    normalize_part_no,
    normalize_sheet_tag,
    parse_named_sheet,
    parse_workbook_bytes,
    parse_yn,
    pick_default_sheet,
    sheet_by_name,
)


def _xlsx_bytes() -> bytes:
    book = Workbook()
    odd = book.active
    odd.title = "RFQ in"
    odd.append(["Item Code", "Qty pcs", "Customer", "Quote No", "CT min", "Notes"])
    odd.append(["BB18-KS0526-28", 104, "OSS", "6000373725", 180, "Thread gauges"])
    odd.append(["NEW-PART-01", 10, "ACME", "6000001", "", ""])
    archive = book.create_sheet("Archive")
    archive.append(
        [
            "Part No.", "RFQ", "Cust.", "Salesperson", "QTY", "Opns", "Assignment",
            "Machines", "Total C/T (mins)", "Machine Hours", "Total Hours", "Days",
            "Lead Time", "Need Tooling?", "Need Fixture?", "Remark",
        ]
    )
    archive.append(
        [
            "BB18-KS0526-28", "6000373725", "OSS", "Daniel", 104, "2TN 1ML", "",
            "22,10,30,29", 180, 312, 312, 31.2, "6-7wks", "N", "N", "Thread gauges",
        ]
    )
    buf = io.BytesIO()
    book.save(buf)
    return buf.getvalue()


class RfqCheckerServiceTests(unittest.TestCase):
    def test_hours_and_lead_time_match_archive_sample(self):
        calc = calculate_times(104, 180)
        self.assertEqual(calc["machine_hours"], 312.0)
        self.assertEqual(calc["total_hours"], 312.0)
        self.assertEqual(calc["days"], 31.2)
        self.assertEqual(calc["lead_time"], "6-7wks")
        second = calculate_times(150, 84)
        self.assertEqual(second["total_hours"], 210.0)
        self.assertEqual(second["days"], 21.0)
        self.assertEqual(second["lead_time"], "4-5wks")

    def test_format_lead_time_and_yn(self):
        self.assertEqual(format_lead_time(0), "")
        self.assertEqual(format_lead_time(5), "1wk")
        self.assertEqual(parse_yn("yes"), "Y")
        self.assertEqual(parse_yn("No"), "N")
        self.assertEqual(normalize_part_no("bb18 ks0526_28"), "BB18-KS0526-28")

    def test_heuristic_maps_varying_and_archive_headers(self):
        varying = heuristic_column_map(["Item Code", "Qty pcs", "Customer", "Quote No", "CT min", "Notes"])
        self.assertEqual(varying["Item Code"], "part_no")
        self.assertEqual(varying["Qty pcs"], "qty")
        self.assertEqual(varying["Quote No"], "rfq")
        self.assertEqual(varying["CT min"], "total_ct_mins")
        self.assertEqual(varying["Notes"], "remark")
        archive = heuristic_column_map(["Part No.", "QTY", "Assignment", "Remark", "Total C/T (mins)"])
        self.assertEqual(archive["Part No."], "part_no")
        self.assertEqual(archive["Assignment"], "assignment")
        self.assertEqual(archive["Total C/T (mins)"], "total_ct_mins")
        self.assertTrue(heuristic_covers_core_fields(archive))
        self.assertFalse(heuristic_covers_core_fields({"Item Code": "part_no"}))
        self.assertFalse(heuristic_covers_core_fields({"Part No.": "part_no", "Assignment": "assignment"}))

    def test_invert_field_map_accepts_either_direction(self):
        self.assertEqual(invert_field_map({"part_no": "Item Code"}), {"Item Code": "part_no"})
        self.assertEqual(invert_field_map({"Item Code": "part_no"}), {"Item Code": "part_no"})

    def test_groq_key_defaults_to_groq_endpoint(self):
        env = {
            "RFQ_LLM_API_KEY": "gsk_test_key",
            "RFQ_LLM_BASE_URL": "",
            "RFQ_LLM_MODEL": "",
            "GROQ_API_KEY": "",
            "OPENAI_API_KEY": "",
        }
        with patch.dict(os.environ, env, clear=False):
            status = llm_status()
        self.assertTrue(status["configured"])
        self.assertEqual(status["provider"], "groq")
        self.assertEqual(status["base_url"], GROQ_DEFAULT_BASE_URL)
        self.assertEqual(status["model"], GROQ_DEFAULT_MODEL)

    def test_groq_openai_turbo_model_is_remapped(self):
        env = {
            "RFQ_LLM_API_KEY": "gsk_test_key",
            "RFQ_LLM_BASE_URL": "",
            "RFQ_LLM_MODEL": "openai/gpt-3.5-turbo",
            "GROQ_API_KEY": "",
            "OPENAI_API_KEY": "",
        }
        with patch.dict(os.environ, env, clear=False):
            status = llm_status()
        self.assertEqual(status["provider"], "groq")
        self.assertEqual(status["model"], GROQ_DEFAULT_MODEL)

    def test_openai_key_keeps_openai_defaults(self):
        env = {
            "RFQ_LLM_API_KEY": "sk-test-key",
            "RFQ_LLM_BASE_URL": "",
            "RFQ_LLM_MODEL": "",
            "GROQ_API_KEY": "",
            "OPENAI_API_KEY": "",
        }
        with patch.dict(os.environ, env, clear=False):
            status = llm_status()
        self.assertEqual(status["provider"], "openai")
        self.assertIn("openai.com", status["base_url"])
        self.assertEqual(status["model"], "gpt-4o-mini")

    def test_map_columns_posts_to_groq_for_gsk_key(self):
        fake = Mock()
        fake.status_code = 200
        fake.text = ""
        fake.raise_for_status.return_value = None
        fake.json.return_value = {
            "choices": [{
                "message": {
                    "content": '{"column_map": {"Item Code": "part_no", "Qty pcs": "qty"}, "notes": "ok"}'
                }
            }]
        }
        env = {
            "RFQ_LLM_API_KEY": "gsk_test_key",
            "RFQ_LLM_BASE_URL": "",
            "RFQ_LLM_MODEL": "",
            "GROQ_API_KEY": "",
            "OPENAI_API_KEY": "",
        }
        with patch.dict(os.environ, env, clear=False):
            with patch("planning.rfq_checker_service.requests.post", return_value=fake) as post:
                result = map_columns_with_llm(
                    ["Item Code", "Qty pcs"],
                    [{"Item Code": "A", "Qty pcs": 1}],
                )
        url = post.call_args.args[0]
        self.assertTrue(url.startswith(GROQ_DEFAULT_BASE_URL))
        self.assertIn("/chat/completions", url)
        self.assertEqual(post.call_args.kwargs["json"]["model"], GROQ_DEFAULT_MODEL)
        self.assertEqual(result["column_map"]["Item Code"], "part_no")
        self.assertEqual(result["column_map"]["Qty pcs"], "qty")
        self.assertEqual(result["model"], GROQ_DEFAULT_MODEL)

    def test_apply_map_calculates_hours_but_not_days(self):
        mapped = apply_column_map(
            {"Item Code": "BB18-KS0526-28", "Qty pcs": "104", "CT min": "180"},
            {"Item Code": "part_no", "Qty pcs": "qty", "CT min": "total_ct_mins"},
        )
        self.assertEqual(mapped["part_no"], "BB18-KS0526-28")
        self.assertEqual(mapped["total_hours"], 312.0)
        self.assertIsNone(mapped["days"])
        self.assertEqual(mapped["lead_time"], "")
        with_schedule = apply_column_map(
            {"Item Code": "BB18-KS0526-28", "Qty pcs": "104", "CT min": "180", "Days": 12, "Lead Time": "3wks"},
            {
                "Item Code": "part_no",
                "Qty pcs": "qty",
                "CT min": "total_ct_mins",
                "Days": "days",
                "Lead Time": "lead_time",
            },
        )
        self.assertEqual(with_schedule["days"], 12.0)
        self.assertEqual(with_schedule["lead_time"], "3wks")
        self.assertEqual(with_schedule["total_hours"], 312.0)

    def test_known_parts_pull_history_new_parts_stay_blank_on_schedule(self):
        lines = build_mapped_lines(
            [
                {"Item Code": "BB18-KS0526-28", "Qty pcs": 104, "CT min": ""},
                {"Item Code": "NEW-PART-01", "Qty pcs": 10, "CT min": ""},
            ],
            {"Item Code": "part_no", "Qty pcs": "qty", "CT min": "total_ct_mins"},
            {
                "BB18-KS0526-28": {
                    "part_no": "BB18-KS0526-28",
                    "opns": "2TN 1ML",
                    "assignment": "MC 22",
                    "machines": "22,10,30,29",
                    "total_ct_mins": 180,
                }
            },
        )
        self.assertEqual(lines[0]["match_status"], "matched")
        self.assertEqual(lines[0]["opns"], "2TN 1ML")
        self.assertEqual(lines[0]["assignment"], "MC 22")
        self.assertIn("assignment", lines[0]["filled_from_history"])
        self.assertEqual(lines[0]["total_hours"], 312.0)
        self.assertIn("total_ct_mins", lines[0]["filled_from_history"])
        self.assertIsNone(lines[0]["days"])
        self.assertEqual(lines[0]["lead_time"], "")
        self.assertEqual(lines[1]["match_status"], "new")
        self.assertIsNone(lines[1]["total_ct_mins"])
        self.assertIsNone(lines[1]["days"])

    def test_infer_cycle_time_sums_operation_columns(self):
        source = {"Part No.": "NEW-1", "QTY": 4, "OP10": 12, "OP20": 8, "OP30": 5}
        inferred = infer_cycle_time_from_source(source, {"Part No.": "part_no", "QTY": "qty"})
        self.assertEqual(inferred, 25.0)
        mapped = apply_column_map(source, {"Part No.": "part_no", "QTY": "qty"})
        self.assertEqual(mapped["total_ct_mins"], 25.0)
        self.assertEqual(mapped["total_hours"], 1.6667)

    def test_group_archive_batches_keeps_upload_order(self):
        grouped = group_archive_batches(
            [
                {
                    "batch_id": 9,
                    "line_id": 1,
                    "part_no": "A",
                    "match_status": "new",
                    "filename": "rfq.xlsx",
                    "sheet_name": "RFQ 3",
                    "batch_status": "archived",
                    "batch_updated_at": "2026-09-01",
                    "qty": 2,
                },
                {
                    "batch_id": 9,
                    "line_id": 2,
                    "part_no": "B",
                    "match_status": "matched",
                    "filename": "rfq.xlsx",
                    "sheet_name": "RFQ 3",
                    "batch_status": "archived",
                    "qty": 4,
                },
                {
                    "batch_id": 8,
                    "line_id": 3,
                    "part_no": "C",
                    "match_status": "new",
                    "filename": "older.xlsx",
                    "sheet_name": "RFQ 1",
                    "batch_status": "draft",
                },
            ]
        )
        self.assertEqual([item["batch_id"] for item in grouped], [9, 8])
        self.assertEqual(grouped[0]["line_count"], 2)
        self.assertEqual(grouped[0]["new_count"], 1)
        self.assertEqual(grouped[0]["matched_count"], 1)
        self.assertEqual(grouped[0]["filename"], "rfq.xlsx")
        self.assertEqual(grouped[1]["status"], "draft")

    def test_parse_workbook_reads_sheets(self):
        sheets = parse_workbook_bytes(_xlsx_bytes(), "rfq.xlsx")
        names = [sheet["name"] for sheet in sheets]
        self.assertIn("RFQ in", names)
        self.assertIn("Archive", names)
        incoming = next(sheet for sheet in sheets if sheet["name"] == "RFQ in")
        self.assertEqual(incoming["row_count"], 2)
        self.assertEqual(incoming["rows"][0]["Item Code"], "BB18-KS0526-28")

    def test_parse_named_sheet_skips_other_tabs(self):
        payload = _xlsx_bytes()
        names = list_workbook_sheets(payload, "rfq.xlsx")
        self.assertEqual(names, ["RFQ in", "Archive"])
        incoming = parse_named_sheet(payload, "rfq.xlsx", "RFQ in")
        self.assertEqual(incoming["name"], "RFQ in")
        self.assertEqual(incoming["row_count"], 2)
        self.assertEqual(incoming["rows"][0]["Item Code"], "BB18-KS0526-28")

    def test_pick_default_sheet_prefers_last_rfq_tab(self):
        sheets = [
            {"name": "Archive", "row_count": 400},
            {"name": "RFQ 1", "row_count": 8},
            {"name": "RFQ 3", "row_count": 22},
        ]
        self.assertEqual(pick_default_sheet(sheets), "RFQ 3")
        self.assertEqual(sheet_by_name(sheets, "rfq-3")["name"], "RFQ 3")

    def test_parse_two_column_sheet_is_not_dropped(self):
        book = Workbook()
        ws = book.active
        ws.title = "RFQ 3"
        ws.append(["Part No.", "QTY"])
        ws.append(["8818-KS1712-02", 12])
        notes = book.create_sheet("Notes")
        notes.append(["Cover"])
        buf = io.BytesIO()
        book.save(buf)
        sheets = parse_workbook_bytes(buf.getvalue(), "rfq.xlsx")
        names = [sheet["name"] for sheet in sheets]
        self.assertEqual(names, ["RFQ 3", "Notes"])
        rfq = sheet_by_name(sheets, "RFQ 3")
        self.assertEqual(rfq["row_count"], 1)
        self.assertEqual(rfq["rows"][0]["Part No."], "8818-KS1712-02")

    def test_headers_from_source_rows_keep_column_order(self):
        headers = headers_from_source_rows(
            [{"source_row": {"Part No.": "A", "QTY": 1, "Cust.": "OSS"}}]
        )
        self.assertEqual(headers, ["Part No.", "QTY", "Cust."])

    def test_sheet_tag_and_overall_defaults(self):
        self.assertEqual(normalize_sheet_tag("nps"), "NPS")
        self.assertEqual(normalize_sheet_tag("APS26-01"), "APS")
        self.assertEqual(normalize_sheet_tag("  "), "")
        lines = apply_defaults_to_mapped_lines(
            [
                {"part_no": "A", "rfq": "old", "customer": "", "salesperson": "Daniel"},
                {"part_no": "B", "rfq": "", "customer": "ACME", "salesperson": ""},
            ],
            {"sheet_tag": "pps", "rfq": "6001", "customer": "OSS"},
        )
        self.assertEqual(lines[0]["sheet_tag"], "PPS")
        self.assertEqual(lines[0]["rfq"], "6001")
        self.assertEqual(lines[0]["customer"], "OSS")
        self.assertEqual(lines[0]["salesperson"], "Daniel")
        self.assertEqual(lines[1]["customer"], "OSS")
        self.assertEqual(lines[1]["salesperson"], "")
        scheduled = apply_defaults_to_mapped_lines(
            [{"part_no": "A", "days": None, "lead_time": ""}],
            {"days": 4, "lead_time": "1wk"},
        )
        self.assertEqual(scheduled[0]["days"], 4.0)
        self.assertEqual(scheduled[0]["lead_time"], "1wk")


class RfqCheckerRouteTests(unittest.TestCase):
    def setUp(self):
        self.app = app
        self.app.config["TESTING"] = True
        self.client = self.app.test_client()

    def test_pages_render(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": ""}):
            library = self.client.get("/archive/rfq-checker")
            upload = self.client.get("/archive/rfq-checker/upload")
        self.assertEqual(library.status_code, 200)
        self.assertIn("RFQ Checker / Tracker", library.get_data(as_text=True))
        self.assertIn("Tracker", library.get_data(as_text=True))
        self.assertIn("By part no.", library.get_data(as_text=True))
        self.assertEqual(upload.status_code, 200)
        self.assertIn("Upload RFQ Excel", upload.get_data(as_text=True))
        self.assertIn("Sheet defaults", upload.get_data(as_text=True))
        self.assertIn("Lead time", upload.get_data(as_text=True))
        self.assertIn("data-rfq-tag=\"APS\"", upload.get_data(as_text=True))

    def test_short_path_redirects(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": ""}):
            response = self.client.get("/rfq-checker")
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/archive/rfq-checker"))

    def test_upload_passes_selected_sheet(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": ""}):
            with patch("planning.rfq_checker_route.create_batch_from_upload") as create:
                create.return_value = {
                    "batch_id": 9,
                    "sheet_name": "RFQ 3",
                    "lines": [],
                    "sheets": [{"name": "RFQ 3", "row_count": 1}],
                    "headers": ["Part No."],
                }
                response = self.client.post(
                    "/api/rfq-checker/upload",
                    data={"sheet": "RFQ 3", "use_llm": "0", "file": (io.BytesIO(_xlsx_bytes()), "rfq.xlsx")},
                    content_type="multipart/form-data",
                )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(create.call_args.kwargs["sheet_name"], "RFQ 3")
        self.assertEqual(create.call_args.kwargs["use_llm"], False)

    def test_upload_requires_file(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": ""}):
            response = self.client.post("/api/rfq-checker/upload")
        self.assertEqual(response.status_code, 400)
        self.assertIn("Excel", response.get_json()["error"])

    def test_meta_reports_llm_status(self):
        with patch.dict(os.environ, {
            "PLANNER_PASSCODE": "",
            "RFQ_LLM_API_KEY": "",
            "GROQ_API_KEY": "",
            "OPENAI_API_KEY": "",
        }):
            response = self.client.get("/api/rfq-checker/meta")
        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertTrue(body["ok"])
        self.assertIn("part_no", body["field_labels"])
        self.assertEqual(body["hours_per_day"], 10)
        self.assertIn("APS", body["sheet_tags"])
        self.assertIn("NPS", body["sheet_tags"])
        self.assertFalse(body["llm"]["configured"])

    def test_part_master_route_ok(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": ""}):
            with patch("planning.rfq_checker_route.list_part_master") as listed:
                listed.return_value = {
                    "ok": True,
                    "count": 1,
                    "rows": [{"part_no": "BB18-KS0526-28", "assignment": "MC 22", "total_ct_mins": 180}],
                    "query": "",
                }
                response = self.client.get("/api/rfq-checker/part-master")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["rows"][0]["assignment"], "MC 22")

    def test_defaults_route_requires_json(self):
        with patch.dict(os.environ, {"PLANNER_PASSCODE": ""}):
            with patch("planning.rfq_checker_route.update_batch_defaults") as update:
                update.return_value = {"batch_id": 3, "sheet_tag": "APS", "lines": []}
                response = self.client.patch(
                    "/api/rfq-checker/batches/3/defaults",
                    json={"sheet_tag": "APS", "customer": "OSS"},
                )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(update.call_args.args[0], 3)
        self.assertEqual(update.call_args.args[1]["customer"], "OSS")


if __name__ == "__main__":
    unittest.main()
