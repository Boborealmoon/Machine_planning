"""RFQ checker: Excel ingest, LLM/heuristic column mapping, cycle-time math, part match."""
from __future__ import annotations

import io
import json
import logging
import math
import os
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import requests
from psycopg2.extras import Json

from db import planner_db_connect_error

from .helpers import one, planner_db, planner_try_savepoint, rows
from .utils import compact_text, parse_number

logger = logging.getLogger(__name__)

HOURS_PER_DAY = 10.0
DAYS_PER_WEEK = 5.0
MAX_UPLOAD_BYTES = 12 * 1024 * 1024
MAX_RFQ_LINES = 2000
SAMPLE_ROWS_FOR_LLM = 8
LLM_TIMEOUT_SEC = 12
SHEET_TAGS = ("APS", "NPS", "PPS", "MPS", "CPS", "SR")
BATCH_DEFAULT_FIELDS = ("rfq", "customer", "salesperson")
BATCH_SCHEDULE_FIELDS = ("days", "lead_time")
OPENAI_DEFAULT_BASE_URL = "https://api.openai.com/v1"
GROQ_DEFAULT_BASE_URL = "https://api.groq.com/openai/v1"
OPENAI_DEFAULT_MODEL = "gpt-4o-mini"
GROQ_DEFAULT_MODEL = "openai/gpt-oss-20b"
_GROQ_OPENAI_MODELS = frozenset({"openai/gpt-oss-20b", "openai/gpt-oss-120b"})

FIXED_FIELDS = (
    "part_no",
    "rfq",
    "customer",
    "salesperson",
    "qty",
    "opns",
    "assignment",
    "machines",
    "total_ct_mins",
    "machine_hours",
    "total_hours",
    "days",
    "lead_time",
    "need_tooling",
    "need_fixture",
    "remark",
)

LINE_PATCH_FIELDS = FIXED_FIELDS

FIELD_LABELS = {
    "part_no": "Part No.",
    "rfq": "RFQ",
    "customer": "Cust.",
    "salesperson": "Salesperson",
    "qty": "QTY",
    "opns": "Opns",
    "assignment": "Assignment",
    "machines": "Machines",
    "total_ct_mins": "Total C/T (mins)",
    "machine_hours": "Machine Hours",
    "total_hours": "Total Hours",
    "days": "Days",
    "lead_time": "Lead Time",
    "need_tooling": "Need Tooling?",
    "need_fixture": "Need Fixture?",
    "remark": "Remark",
}

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "part_no": (
        "part_no", "partno", "part_number", "item_no", "item_code",
        "inventory_code", "stock_code", "pn", "p_n", "part",
    ),
    "rfq": ("rfq", "rfq_no", "rfq_number", "quotation", "quote_no", "quote"),
    "customer": ("cust", "customer", "customer_name", "customer_code", "cust_name"),
    "salesperson": ("salesperson", "sales_person", "sales", "ae", "pic_sales"),
    "qty": ("qty", "quantity", "order_qty", "qty_pcs", "qtypcs"),
    "opns": ("opns", "ops", "operations", "op", "process"),
    "assignment": ("assignment", "assign", "assigned", "assigned_mc", "assign_mc"),
    "machines": ("machines", "machine", "mc", "mc_no", "machine_no", "machine_nos"),
    "total_ct_mins": (
        "total_c_t_mins", "total_ct_mins", "total_ct", "total_c_t", "cycle_time",
        "ct_mins", "ct_min", "c_t", "total_cycle_time", "cycle_time_mins",
        "mins_pc", "min_pc", "minutes_pc", "std_time", "std_mins", "unit_time",
        "time_mins", "cycletime", "ctpc", "mins_per_pc",
    ),
    "machine_hours": ("machine_hours", "mc_hours", "mch_hours"),
    "total_hours": ("total_hours", "hours"),
    "days": ("days", "day"),
    "lead_time": ("lead_time", "lt", "leadtime", "lead_time_wks"),
    "need_tooling": ("need_tooling", "tooling", "need_tool"),
    "need_fixture": ("need_fixture", "fixture", "need_jig", "jig"),
    "remark": ("remark", "remarks", "notes", "comment", "comments"),
}

_TURN_RE = re.compile(r"\b(turn|turning|tn|lathe)\b", re.I)
_MILL_RE = re.compile(r"\b(mill|milling|ml|vmc|hmc)\b", re.I)


def json_error(exc: Exception, *, fallback_status: int = 500):
    friendly = planner_db_connect_error(exc)
    if friendly:
        return {"error": friendly}, 503
    text = str(exc) or exc.__class__.__name__
    lower = text.lower()
    if (
        "timed out" in lower
        or "timeout" in lower
        or "querycanceled" in lower
        or "statement timeout" in lower
        or "read timed out" in lower
    ):
        return {
            "error": (
                "RFQ upload timed out before assignment and cycle times could be saved. "
                "Uncheck “Use LLM to map columns”, or choose a single RFQ sheet instead of "
                "a full Archive workbook."
            )
        }, 504
    return {"error": text}, fallback_status


def llm_status() -> dict[str, Any]:
    key = _llm_api_key()
    if not key:
        return {"configured": False, "model": "", "base_url": "", "provider": ""}
    return {
        "configured": True,
        "model": _llm_model(),
        "base_url": _llm_base_url(),
        "provider": _llm_provider(),
    }


def _llm_api_key() -> str:
    return compact_text(
        os.getenv("RFQ_LLM_API_KEY")
        or os.getenv("GROQ_API_KEY")
        or os.getenv("OPENAI_API_KEY")
    )


def _llm_uses_groq() -> bool:
    base = compact_text(os.getenv("RFQ_LLM_BASE_URL")).lower()
    if "groq.com" in base:
        return True
    if base:
        return False
    return _llm_api_key().startswith("gsk_")


def _llm_provider() -> str:
    if _llm_uses_groq():
        return "groq"
    base = _llm_base_url().lower()
    if "openai.com" in base:
        return "openai"
    return "custom"


def _llm_base_url() -> str:
    explicit = compact_text(os.getenv("RFQ_LLM_BASE_URL")).rstrip("/")
    if explicit:
        return explicit
    if _llm_uses_groq():
        return GROQ_DEFAULT_BASE_URL
    return OPENAI_DEFAULT_BASE_URL


def _llm_model() -> str:
    explicit = compact_text(os.getenv("RFQ_LLM_MODEL"))
    if _llm_uses_groq():
        return _groq_model_name(explicit)
    if explicit:
        return explicit
    return OPENAI_DEFAULT_MODEL


def _groq_model_name(explicit: str) -> str:
    if not explicit:
        return GROQ_DEFAULT_MODEL
    lower = explicit.lower()
    if "gpt-oss" in lower or lower in _GROQ_OPENAI_MODELS:
        return explicit
    if lower.startswith("openai/gpt-") or lower.startswith(("gpt-3", "gpt-4", "gpt-5", "o1-", "o3-", "o4-")):
        logger.warning(
            "RFQ_LLM_MODEL %s is not served by Groq; using %s",
            explicit,
            GROQ_DEFAULT_MODEL,
        )
        return GROQ_DEFAULT_MODEL
    return explicit


def heuristic_covers_core_fields(column_map: dict[str, str] | None) -> bool:
    fields = {compact_text(value) for value in (column_map or {}).values()}
    return "part_no" in fields and "total_ct_mins" in fields


def normalize_part_no(value: Any) -> str:
    text = compact_text(value).upper()
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-{2,}", "-", text)
    return text.strip("-")


def normalize_header(value: Any) -> str:
    text = compact_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_sheet_tag(value: Any) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", compact_text(value).upper())
    if not compact:
        return ""
    for tag in SHEET_TAGS:
        if compact == tag or compact.startswith(tag):
            return tag
    return compact[:16]


def apply_defaults_to_mapped_lines(
    lines: list[dict[str, Any]],
    defaults: dict[str, Any] | None,
    *,
    overwrite: bool = True,
    fields: list[str] | None = None,
) -> list[dict[str, Any]]:
    payload = defaults or {}
    wanted = fields or ["sheet_tag", *BATCH_DEFAULT_FIELDS, *BATCH_SCHEDULE_FIELDS]
    tag = normalize_sheet_tag(payload.get("sheet_tag"))
    out: list[dict[str, Any]] = []
    for line in lines:
        row = dict(line)
        if "sheet_tag" in wanted and ("sheet_tag" in payload or tag):
            row["sheet_tag"] = tag
        for field in BATCH_DEFAULT_FIELDS:
            if field not in wanted:
                continue
            if field not in payload and f"default_{field}" not in payload:
                continue
            value = compact_text(payload.get(field) or payload.get(f"default_{field}"))
            if overwrite or not compact_text(row.get(field)):
                row[field] = value
        if "days" in wanted and ("days" in payload or "default_days" in payload):
            days_val = payload.get("days") if payload.get("days") is not None else payload.get("default_days")
            if overwrite or row.get("days") in (None, ""):
                row["days"] = _to_number(days_val)
        if "lead_time" in wanted and ("lead_time" in payload or "default_lead_time" in payload):
            lead = compact_text(payload.get("lead_time") or payload.get("default_lead_time"))
            if overwrite or not compact_text(row.get("lead_time")):
                row["lead_time"] = lead
        out.append(row)
    return out


def parse_yn(value: Any) -> str:
    text = compact_text(value).lower()
    if not text:
        return ""
    if text in {"y", "yes", "true", "1", "t"}:
        return "Y"
    if text in {"n", "no", "false", "0", "f"}:
        return "N"
    if text.startswith("y"):
        return "Y"
    if text.startswith("n"):
        return "N"
    return compact_text(value).upper()[:1]


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return None
        return number
    text = compact_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    parsed = parse_number(match.group(0), default=float("nan"))
    if parsed != parsed:
        return None
    return float(parsed)


def format_lead_time(days: float) -> str:
    days = float(days or 0)
    if days <= 0:
        return ""
    weeks = days / DAYS_PER_WEEK
    low = max(1, math.floor(weeks))
    high = max(low, math.ceil(weeks - 1e-9))
    if high <= 1:
        return "1wk"
    if low == high:
        return f"{low}wks"
    return f"{low}-{high}wks"


def calculate_times(qty: Any, total_ct_mins: Any, *, total_hours: Any = None) -> dict[str, Any]:
    qty_n = _to_number(qty) or 0.0
    ct_n = _to_number(total_ct_mins) or 0.0
    hours_override = _to_number(total_hours)
    if hours_override is None:
        hours = (qty_n * ct_n / 60.0) if qty_n and ct_n else 0.0
    else:
        hours = hours_override
    days = hours / HOURS_PER_DAY if hours else 0.0
    return {
        "machine_hours": round(hours, 4),
        "total_hours": round(hours, 4),
        "days": round(days, 4),
        "lead_time": format_lead_time(days),
    }


def summarize_opns(op_types: list[str]) -> str:
    turn = 0
    mill = 0
    other: list[str] = []
    for raw in op_types:
        label = compact_text(raw)
        if not label:
            continue
        if _TURN_RE.search(label):
            turn += 1
        elif _MILL_RE.search(label):
            mill += 1
        else:
            other.append(label)
    parts: list[str] = []
    if turn:
        parts.append(f"{turn}TN")
    if mill:
        parts.append(f"{mill}ML")
    if other:
        parts.append(", ".join(other[:4]))
    return " ".join(parts)


def heuristic_column_map(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    used_fields: set[str] = set()
    for header in headers:
        norm = normalize_header(header)
        if not norm:
            continue
        matched = ""
        for field, aliases in _HEADER_ALIASES.items():
            if field in used_fields:
                continue
            collapsed_norm = norm.replace("_", "")
            collapsed_aliases = {alias.replace("_", "") for alias in aliases}
            if norm in aliases or collapsed_norm in collapsed_aliases:
                matched = field
                break
            if any(norm == alias or norm.startswith(alias + "_") for alias in aliases):
                matched = field
                break
        if matched:
            mapping[header] = matched
            used_fields.add(matched)
    return mapping


def invert_field_map(mapping: dict[str, Any] | None) -> dict[str, str]:
    if not mapping:
        return {}
    values = [compact_text(value) for value in mapping.values()]
    if values and all(value in FIXED_FIELDS or value == "" for value in values):
        return {
            compact_text(key): compact_text(value)
            for key, value in mapping.items()
            if compact_text(value) in FIXED_FIELDS
        }
    inverted: dict[str, str] = {}
    for key, value in mapping.items():
        field = compact_text(key)
        header = compact_text(value)
        if field in FIXED_FIELDS and header:
            inverted[header] = field
    return inverted


def apply_hours(mapped: dict[str, Any], *, hours_override: Any = None) -> dict[str, Any]:
    """Fill machine/total hours from qty x C/T. Never touches days or lead time."""
    calc = calculate_times(
        mapped.get("qty"),
        mapped.get("total_ct_mins"),
        total_hours=_to_number(hours_override) if hours_override is not None else None,
    )
    mapped["machine_hours"] = calc["machine_hours"]
    mapped["total_hours"] = calc["total_hours"]
    return mapped


def mapped_field_set(column_map: dict[str, str] | None) -> set[str]:
    return {compact_text(value) for value in (column_map or {}).values() if compact_text(value)}


def infer_cycle_time_from_source(
    source_row: dict[str, Any] | None,
    column_map: dict[str, str] | None = None,
) -> float | None:
    """Guess total C/T from leftover columns (a dedicated CT col, or summed op times)."""
    source = source_row or {}
    used = {normalize_header(header) for header in (column_map or {})}
    dedicated: list[float] = []
    op_times: list[float] = []
    for header, value in source.items():
        if normalize_header(header) in used:
            continue
        number = _to_number(value)
        if number is None or number <= 0 or number > 20000:
            continue
        norm = normalize_header(header)
        collapsed = norm.replace("_", "")
        if any(token in collapsed for token in ("totalct", "cycletime", "ctmins", "ctmin", "minspc", "stdtime", "unittime")):
            dedicated.append(number)
            continue
        if re.search(r"(^op|_op|operation|stage).*(ct|cycle|min|time)|^(op|stage)_?\d+$", norm):
            op_times.append(number)
    if len(dedicated) == 1:
        return dedicated[0]
    if len(op_times) >= 2:
        return round(sum(op_times), 4)
    return None


def apply_column_map(source_row: dict[str, Any], column_map: dict[str, str]) -> dict[str, Any]:
    mapped: dict[str, Any] = {field: "" for field in FIXED_FIELDS}
    present = mapped_field_set(column_map)
    for header, field in (column_map or {}).items():
        if field not in mapped or compact_text(mapped.get(field)):
            continue
        if header in source_row:
            mapped[field] = source_row.get(header)
            continue
        want = normalize_header(header)
        for key, value in source_row.items():
            if normalize_header(key) == want:
                mapped[field] = value
                break
    mapped["need_tooling"] = parse_yn(mapped.get("need_tooling"))
    mapped["need_fixture"] = parse_yn(mapped.get("need_fixture"))
    for text_field in ("part_no", "rfq", "customer", "salesperson", "opns", "assignment", "machines", "lead_time", "remark"):
        mapped[text_field] = compact_text(mapped.get(text_field))
    mapped["qty"] = _to_number(mapped.get("qty"))
    mapped["total_ct_mins"] = _to_number(mapped.get("total_ct_mins"))
    if mapped["total_ct_mins"] is None:
        inferred = infer_cycle_time_from_source(source_row, column_map)
        if inferred is not None:
            mapped["total_ct_mins"] = inferred
    incoming_hours = _to_number(mapped.get("total_hours")) if "total_hours" in present else None
    incoming_mc = _to_number(mapped.get("machine_hours")) if "machine_hours" in present else None
    calc = calculate_times(mapped["qty"], mapped["total_ct_mins"], total_hours=incoming_hours)
    mapped["machine_hours"] = incoming_mc if incoming_mc is not None else calc["machine_hours"]
    mapped["total_hours"] = incoming_hours if incoming_hours is not None else calc["total_hours"]
    mapped["days"] = _to_number(mapped.get("days")) if "days" in present else None
    if "lead_time" not in present:
        mapped["lead_time"] = ""
    return mapped


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and value != value:
        return None
    return value


def parse_workbook_bytes(payload: bytes, filename: str = "") -> list[dict[str, Any]]:
    name = compact_text(filename).lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        return _parse_xls(payload)
    return _parse_xlsx(payload)


def list_workbook_sheets(payload: bytes, filename: str = "") -> list[str]:
    name = compact_text(filename).lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        return [item["name"] for item in _parse_xls(payload)]
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        return [compact_text(title) or "Sheet1" for title in workbook.sheetnames]
    finally:
        workbook.close()


def parse_named_sheet(payload: bytes, filename: str = "", sheet_name: str = "") -> dict[str, Any]:
    name = compact_text(filename).lower()
    wanted = compact_text(sheet_name)
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        sheets = _parse_xls(payload)
        if not sheets:
            raise ValueError("No usable worksheet was found in that workbook.")
        chosen = wanted or pick_default_sheet(sheets)
        sheet = sheet_by_name(sheets, chosen) if chosen else sheets[0]
        if not sheet:
            available = ", ".join(item["name"] for item in sheets) or "none"
            raise ValueError(f"Sheet {chosen!r} was not found. Available: {available}.")
        return sheet
    names = list_workbook_sheets(payload, filename)
    if not names:
        raise ValueError("No usable worksheet was found in that workbook.")
    summaries = [{"name": item, "row_count": 0} for item in names]
    chosen = wanted or pick_default_sheet(summaries)
    match = next(
        (
            item
            for item in names
            if compact_text(item) == compact_text(chosen)
            or normalize_header(item) == normalize_header(chosen)
        ),
        "",
    )
    if not match:
        available = ", ".join(names) or "none"
        raise ValueError(f"Sheet {chosen!r} was not found. Available: {available}.")
    sheet = _parse_xlsx_named(payload, match, read_only=True)
    if int(sheet.get("row_count") or 0) == 0:
        fallback = _parse_xlsx_named(payload, match, read_only=False)
        if int(fallback.get("row_count") or 0) > 0 or fallback.get("headers"):
            return fallback
    return sheet


def _parse_xlsx_named(payload: bytes, sheet_name: str, *, read_only: bool) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=read_only)
    try:
        try:
            ws = workbook[sheet_name]
        except KeyError as exc:
            raise ValueError(f"Sheet {sheet_name!r} was not found.") from exc
        if not read_only:
            try:
                ws.reset_dimensions()
            except Exception:
                pass
        matrix: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append([_serialize_cell(cell) for cell in row])
        return _sheet_from_matrix(ws.title, matrix)
    finally:
        workbook.close()


def _parse_xlsx(payload: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    # Avoid read_only: stale worksheet dimensions can hide every row on later tabs.
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for ws in workbook.worksheets:
            try:
                ws.reset_dimensions()
            except Exception:
                pass
            matrix: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                matrix.append([_serialize_cell(cell) for cell in row])
            sheets.append(_sheet_from_matrix(ws.title, matrix))
    finally:
        workbook.close()
    return sheets


def _parse_xls(payload: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("xlrd is required to read .xls files.") from exc
    book = xlrd.open_workbook(file_contents=payload)
    sheets: list[dict[str, Any]] = []
    for sheet in book.sheets():
        matrix: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            matrix.append(
                [_serialize_cell(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            )
        sheets.append(_sheet_from_matrix(sheet.name, matrix))
    return sheets


def _sheet_from_matrix(title: str, matrix: list[list[Any]]) -> dict[str, Any]:
    name = compact_text(title) or "Sheet1"
    header_idx = _find_header_row(matrix)
    if header_idx is None:
        for idx, row in enumerate(matrix):
            if any(compact_text(cell) for cell in (row or [])):
                header_idx = idx
                break
    if header_idx is None:
        return {"name": name, "headers": [], "rows": [], "row_count": 0}
    raw_headers = matrix[header_idx]
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, cell in enumerate(raw_headers):
        label = compact_text(cell) or f"Column {idx + 1}"
        count = seen.get(label, 0) + 1
        seen[label] = count
        headers.append(label if count == 1 else f"{label} ({count})")
    records: list[dict[str, Any]] = []
    for row in matrix[header_idx + 1 :]:
        if not any(cell not in (None, "") for cell in (row or [])):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            record[header] = row[idx] if idx < len(row) else None
        records.append(record)
    return {
        "name": name,
        "headers": headers,
        "rows": records,
        "row_count": len(records),
    }


def _find_header_row(matrix: list[list[Any]]) -> int | None:
    best_idx = None
    best_score = 0
    limit = min(len(matrix), 25)
    for idx in range(limit):
        row = matrix[idx] or []
        filled = [compact_text(cell) for cell in row if compact_text(cell)]
        if len(filled) < 2:
            continue
        score = len(filled)
        joined = " ".join(normalize_header(cell) for cell in filled)
        if "part" in joined:
            score += 4
        if "rfq" in joined or "qty" in joined:
            score += 2
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def sheet_by_name(sheets: list[dict[str, Any]], name: str) -> dict[str, Any] | None:
    wanted = compact_text(name)
    if not wanted:
        return None
    for item in sheets:
        if compact_text(item.get("name")) == wanted:
            return item
    wanted_norm = normalize_header(wanted)
    for item in sheets:
        if normalize_header(item.get("name")) == wanted_norm:
            return item
    return None


def pick_default_sheet(sheets: list[dict[str, Any]]) -> str:
    if not sheets:
        return ""
    usable = [item for item in sheets if int(item.get("row_count") or 0) > 0] or list(sheets)

    def is_archive(item: dict[str, Any]) -> bool:
        return normalize_header(item.get("name")) == "archive"

    rfq_named = [item for item in usable if "rfq" in normalize_header(item.get("name"))]
    if rfq_named:
        return rfq_named[-1]["name"]
    non_archive = [item for item in usable if not is_archive(item)]
    pool = non_archive or usable
    return max(pool, key=lambda item: int(item.get("row_count") or 0))["name"]


def headers_from_source_rows(lines: list[dict[str, Any]]) -> list[str]:
    for line in lines or []:
        source = line.get("source_row") or {}
        if isinstance(source, str):
            try:
                source = json.loads(source)
            except json.JSONDecodeError:
                source = {}
        if isinstance(source, dict) and source:
            return [str(key) for key in source.keys()]
    return []


def _parse_llm_json(content: str) -> dict[str, Any]:
    text = compact_text(content) or "{}"
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
        text = re.sub(r"\s*```$", "", text)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, re.S)
        if not match:
            raise
        parsed = json.loads(match.group(0))
    return parsed if isinstance(parsed, dict) else {}


def map_columns_with_llm(headers: list[str], sample_rows: list[dict[str, Any]]) -> dict[str, Any]:
    key = _llm_api_key()
    if not key:
        raise RuntimeError("Set RFQ_LLM_API_KEY, GROQ_API_KEY, or OPENAI_API_KEY to use LLM mapping.")
    payload = {
        "model": _llm_model(),
        "temperature": 0,
        "max_tokens": 2000,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "system",
                "content": (
                    "You map spreadsheet columns onto a fixed RFQ tracker schema. "
                    "Return a JSON object with keys column_map and notes. "
                    "column_map maps original Excel header strings to schema field keys. "
                    "Match semantically, not only by exact name. Examples: "
                    "Item Code or Part Number -> part_no; Qty pcs or Quantity -> qty; "
                    "Quote No or RFQ No -> rfq; Cust. or Customer -> customer; "
                    "CT min, Cycle Time, mins/pc, or Total C/T (mins) -> total_ct_mins; "
                    "Hours or Mach hours -> total_hours; Notes or Remarks -> remark. "
                    "total_ct_mins is minutes per piece, never hours. "
                    "Map days only if a column clearly holds calendar/shop days. "
                    "Map lead_time only if a column clearly holds lead time (weeks/days text). "
                    "Do not invent days or lead time from hours. "
                    "Only omit a header if it clearly has no matching field. "
                    "Do not map two headers to the same field. "
                    "Schema field keys: "
                    + ", ".join(FIXED_FIELDS)
                    + "."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "target_fields": FIELD_LABELS,
                        "headers": headers,
                        "sample_rows": sample_rows[:SAMPLE_ROWS_FOR_LLM],
                    },
                    default=str,
                ),
            },
        ],
    }
    url = _llm_base_url().rstrip("/") + "/chat/completions"
    http_headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    response = requests.post(url, json=payload, headers=http_headers, timeout=LLM_TIMEOUT_SEC)
    if response.status_code >= 400 and "response_format" in (response.text or ""):
        payload.pop("response_format", None)
        response = requests.post(url, json=payload, headers=http_headers, timeout=LLM_TIMEOUT_SEC)
    if response.status_code == 401:
        provider = _llm_provider()
        hint = (
            " Groq keys (gsk_...) need RFQ_LLM_BASE_URL=https://api.groq.com/openai/v1."
            if provider != "groq" and key.startswith("gsk_")
            else ""
        )
        raise RuntimeError(
            f"LLM provider rejected the API key at {url}.{hint}"
        )
    if response.status_code >= 400:
        detail = compact_text(response.text)[:400]
        try:
            err_body = response.json()
            detail = compact_text((err_body.get("error") or {}).get("message") or response.text)[:400]
        except Exception:
            pass
        raise RuntimeError(f"LLM mapping request failed ({response.status_code}) at {url}: {detail}")
    response.raise_for_status()
    body = response.json()
    content = ((body.get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
    parsed = _parse_llm_json(content)
    raw_map = parsed.get("column_map") or parsed.get("mapping") or parsed
    if not isinstance(raw_map, dict):
        raw_map = {}
    column_map = invert_field_map(
        {str(map_key): str(value) for map_key, value in raw_map.items() if value not in (None, "")}
    )
    return {
        "column_map": column_map,
        "notes": compact_text(parsed.get("notes")),
        "model": _llm_model(),
    }


def build_mapped_lines(
    source_rows: list[dict[str, Any]],
    column_map: dict[str, str],
    existing_by_part: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    existing_by_part = existing_by_part or {}
    lines: list[dict[str, Any]] = []
    for idx, source in enumerate(source_rows, start=1):
        if not isinstance(source, dict):
            continue
        mapped = apply_column_map(source, column_map)
        if not mapped["part_no"] and not mapped["rfq"] and not mapped["qty"]:
            continue
        key = normalize_part_no(mapped["part_no"])
        existing = existing_by_part.get(key) if key else None
        filled: list[str] = []
        if existing:
            mapped["match_status"] = "matched"
            mapped["matched_part_no"] = existing.get("part_no") or mapped["part_no"]
            for field in ("assignment", "opns", "machines"):
                if not mapped.get(field) and existing.get(field):
                    mapped[field] = existing[field]
                    filled.append(field)
            if mapped.get("total_ct_mins") in (None, "") and existing.get("total_ct_mins") not in (None, ""):
                mapped["total_ct_mins"] = existing["total_ct_mins"]
                filled.append("total_ct_mins")
            if not mapped.get("customer") and existing.get("rfq_customer"):
                mapped["customer"] = existing["rfq_customer"]
                filled.append("customer")
            if "total_ct_mins" in filled:
                calc = calculate_times(mapped.get("qty"), mapped.get("total_ct_mins"))
                mapped["machine_hours"] = calc["machine_hours"]
                mapped["total_hours"] = calc["total_hours"]
        else:
            mapped["match_status"] = "new"
            mapped["matched_part_no"] = mapped["part_no"]
        mapped["line_no"] = idx
        mapped["filled_from_history"] = filled
        mapped["source_row"] = source
        lines.append(mapped)
    return lines


def fill_missing_cycle_times_with_llm(lines: list[dict[str, Any]]) -> str:
    """For new parts still missing C/T, ask the LLM to read leftover columns. Never invent days/lead time."""
    key = _llm_api_key()
    if not key:
        return ""
    gaps = [
        line for line in lines
        if compact_text(line.get("match_status")) == "new"
        and line.get("total_ct_mins") in (None, "")
        and isinstance(line.get("source_row"), dict)
    ]
    if not gaps:
        return ""
    sample = gaps[:40]
    payload = {
        "model": _llm_model(),
        "temperature": 0,
        "max_tokens": 2000,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "system",
                "content": (
                    "You extract machining cycle time from messy RFQ spreadsheet rows. "
                    "Return JSON {rows: [{line_no, total_ct_mins, days, lead_time}]}. "
                    "total_ct_mins is minutes per piece. If several operation times exist, sum them. "
                    "Use null when the row has no cycle-time numbers — do not guess. "
                    "days and lead_time: copy only if the row already contains those values; otherwise null. "
                    "Never convert hours into days. Never invent a schedule."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(
                    {
                        "rows": [
                            {
                                "line_no": line.get("line_no"),
                                "qty": line.get("qty"),
                                "source": line.get("source_row"),
                            }
                            for line in sample
                        ]
                    },
                    default=str,
                ),
            },
        ],
    }
    url = _llm_base_url().rstrip("/") + "/chat/completions"
    http_headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    response = requests.post(url, json=payload, headers=http_headers, timeout=LLM_TIMEOUT_SEC)
    if response.status_code >= 400 and "response_format" in (response.text or ""):
        payload.pop("response_format", None)
        response = requests.post(url, json=payload, headers=http_headers, timeout=LLM_TIMEOUT_SEC)
    if response.status_code >= 400:
        detail = compact_text(response.text)[:300]
        raise RuntimeError(f"LLM cycle-time fill failed ({response.status_code}): {detail}")
    body = response.json()
    content = ((body.get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
    parsed = _parse_llm_json(content)
    raw_rows = parsed.get("rows") or parsed.get("lines") or []
    by_no = {
        int(item.get("line_no") or 0): item
        for item in raw_rows
        if isinstance(item, dict) and item.get("line_no") not in (None, "")
    }
    filled = 0
    for line in lines:
        item = by_no.get(int(line.get("line_no") or 0))
        if not item:
            continue
        ct = _to_number(item.get("total_ct_mins"))
        if ct is not None and line.get("total_ct_mins") in (None, ""):
            line["total_ct_mins"] = ct
            apply_hours(line)
            filled += 1
        if line.get("days") in (None, "") and _to_number(item.get("days")) is not None:
            line["days"] = _to_number(item.get("days"))
        if not compact_text(line.get("lead_time")) and compact_text(item.get("lead_time")):
            line["lead_time"] = compact_text(item.get("lead_time"))
    if not filled:
        return "LLM checked new parts with no C/T column and did not find extra cycle times."
    return f"LLM filled cycle time on {filled} new part{'s' if filled != 1 else ''} from leftover columns."


def _ensure_tables(con) -> None:
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_rfq_batch (
            batch_id              BIGSERIAL    PRIMARY KEY,
            filename              TEXT         NOT NULL DEFAULT '',
            sheet_name            TEXT         NOT NULL DEFAULT '',
            status                TEXT         NOT NULL DEFAULT 'draft',
            llm_used              BOOLEAN      NOT NULL DEFAULT FALSE,
            llm_model             TEXT         NOT NULL DEFAULT '',
            mapping               JSONB        NOT NULL DEFAULT '{}'::jsonb,
            mapping_notes         TEXT         NOT NULL DEFAULT '',
            sheet_tag             TEXT         NOT NULL DEFAULT '',
            default_rfq           TEXT         NOT NULL DEFAULT '',
            default_customer      TEXT         NOT NULL DEFAULT '',
            default_salesperson   TEXT         NOT NULL DEFAULT '',
            default_days          NUMERIC,
            default_lead_time     TEXT         NOT NULL DEFAULT '',
            created_at            TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            updated_at            TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            CONSTRAINT planner_rfq_batch_status_chk
                CHECK (status IN ('draft', 'archived'))
        )
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_rfq_batch_updated_at
            ON public.planner_rfq_batch (updated_at DESC)
        """
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_rfq_line (
            line_id          BIGSERIAL    PRIMARY KEY,
            batch_id         BIGINT       NOT NULL
                REFERENCES public.planner_rfq_batch(batch_id) ON DELETE CASCADE,
            line_no          INTEGER      NOT NULL DEFAULT 0,
            part_no          TEXT         NOT NULL DEFAULT '',
            rfq              TEXT         NOT NULL DEFAULT '',
            customer         TEXT         NOT NULL DEFAULT '',
            salesperson      TEXT         NOT NULL DEFAULT '',
            sheet_tag        TEXT         NOT NULL DEFAULT '',
            qty              NUMERIC,
            opns             TEXT         NOT NULL DEFAULT '',
            assignment       TEXT         NOT NULL DEFAULT '',
            machines         TEXT         NOT NULL DEFAULT '',
            total_ct_mins    NUMERIC,
            machine_hours    NUMERIC,
            total_hours      NUMERIC,
            days             NUMERIC,
            lead_time        TEXT         NOT NULL DEFAULT '',
            need_tooling     TEXT         NOT NULL DEFAULT '',
            need_fixture     TEXT         NOT NULL DEFAULT '',
            remark           TEXT         NOT NULL DEFAULT '',
            match_status     TEXT         NOT NULL DEFAULT 'new',
            matched_part_no  TEXT         NOT NULL DEFAULT '',
            source_row       JSONB        NOT NULL DEFAULT '{}'::jsonb,
            created_at       TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            updated_at       TIMESTAMPTZ  NOT NULL DEFAULT NOW()
        )
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_rfq_line_batch
            ON public.planner_rfq_line (batch_id, line_no)
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_rfq_line_part_no
            ON public.planner_rfq_line (UPPER(TRIM(part_no)))
        """
    )
    for statement in (
        "ALTER TABLE public.planner_rfq_batch ADD COLUMN IF NOT EXISTS sheet_tag TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE public.planner_rfq_batch ADD COLUMN IF NOT EXISTS default_rfq TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE public.planner_rfq_batch ADD COLUMN IF NOT EXISTS default_customer TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE public.planner_rfq_batch ADD COLUMN IF NOT EXISTS default_salesperson TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE public.planner_rfq_batch ADD COLUMN IF NOT EXISTS default_days NUMERIC",
        "ALTER TABLE public.planner_rfq_batch ADD COLUMN IF NOT EXISTS default_lead_time TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE public.planner_rfq_line ADD COLUMN IF NOT EXISTS sheet_tag TEXT NOT NULL DEFAULT ''",
        "CREATE INDEX IF NOT EXISTS idx_rfq_batch_sheet_tag ON public.planner_rfq_batch (UPPER(TRIM(sheet_tag)))",
        "CREATE INDEX IF NOT EXISTS idx_rfq_line_sheet_tag ON public.planner_rfq_line (UPPER(TRIM(sheet_tag)))",
    ):
        con.execute(statement)
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_rfq_part_master (
            part_key         TEXT         PRIMARY KEY,
            part_no          TEXT         NOT NULL DEFAULT '',
            assignment       TEXT         NOT NULL DEFAULT '',
            opns             TEXT         NOT NULL DEFAULT '',
            machines         TEXT         NOT NULL DEFAULT '',
            total_ct_mins    NUMERIC,
            last_rfq         TEXT         NOT NULL DEFAULT '',
            customer         TEXT         NOT NULL DEFAULT '',
            salesperson      TEXT         NOT NULL DEFAULT '',
            sheet_tag        TEXT         NOT NULL DEFAULT '',
            source_batch_id  BIGINT,
            updated_at       TIMESTAMPTZ  NOT NULL DEFAULT NOW()
        )
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_rfq_part_master_part_no
            ON public.planner_rfq_part_master (UPPER(TRIM(part_no)))
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_rfq_part_master_updated
            ON public.planner_rfq_part_master (updated_at DESC)
        """
    )


def _json_ready(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def serialize_line(row: dict[str, Any] | None, *, include_source: bool = False) -> dict[str, Any] | None:
    if not row:
        return None
    out = {key: _json_ready(value) for key, value in dict(row).items()}
    source = out.get("source_row")
    if not include_source:
        out.pop("source_row", None)
    elif isinstance(source, str):
        try:
            out["source_row"] = json.loads(source)
        except json.JSONDecodeError:
            out["source_row"] = {}
    for field in ("qty", "total_ct_mins", "machine_hours", "total_hours", "days"):
        if out.get(field) is not None:
            out[field] = float(out[field])
    return out


def serialize_batch(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    out = {key: _json_ready(value) for key, value in dict(row).items()}
    mapping = out.get("mapping")
    if isinstance(mapping, str):
        try:
            out["mapping"] = json.loads(mapping)
        except json.JSONDecodeError:
            out["mapping"] = {}
    if out.get("default_days") is not None:
        out["default_days"] = float(out["default_days"])
    return out


def lookup_existing_parts(con, part_nos: list[str]) -> dict[str, dict[str, Any]]:
    keys: list[str] = []
    seen: set[str] = set()
    for part in part_nos:
        for variant in (compact_text(part).upper(), normalize_part_no(part)):
            if variant and variant not in seen:
                seen.add(variant)
                keys.append(variant)
    if not keys:
        return {}
    found: dict[str, dict[str, Any]] = {}
    normalized_keys = [normalize_part_no(key) for key in keys if normalize_part_no(key)]

    def from_cycle_times():
        return rows(
            con.execute(
                """
                SELECT
                    TRIM(part_no) AS part_no,
                    MAX(NULLIF(TRIM(part_description), '')) AS part_description,
                    COUNT(*)::INT AS op_count,
                    SUM(COALESCE(NULLIF(cycle_time, 0), ideal_cycle_time, 0)) AS total_ct_mins,
                    COALESCE(
                        STRING_AGG(DISTINCT NULLIF(TRIM(op_type), ''), ' | '
                                   ORDER BY NULLIF(TRIM(op_type), '')),
                        ''
                    ) AS op_types
                FROM public.planner_cycle_time_master
                WHERE UPPER(TRIM(part_no)) = ANY(%s)
                GROUP BY TRIM(part_no)
                """,
                (keys,),
            )
        )

    for item in planner_try_savepoint(con, "rfq_ct", from_cycle_times, default=[]) or []:
        key = normalize_part_no(item.get("part_no"))
        types = [part.strip() for part in compact_text(item.get("op_types")).split("|") if part.strip()]
        found[key] = {
            "part_no": compact_text(item.get("part_no")),
            "part_description": compact_text(item.get("part_description")),
            "op_count": int(item.get("op_count") or 0),
            "total_ct_mins": float(item.get("total_ct_mins") or 0) or None,
            "opns": summarize_opns(types),
            "assignment": "",
            "machines": "",
            "source": "cycle_time_master",
        }

    def from_vouchers():
        return rows(
            con.execute(
                """
                SELECT
                    TRIM(part_no) AS part_no,
                    MAX(NULLIF(TRIM(description), '')) AS part_description,
                    COUNT(DISTINCT ps_id)::INT AS ps_count,
                    MAX(order_date) AS last_order_date
                FROM public.pp_vouchers_cache
                WHERE UPPER(TRIM(part_no)) = ANY(%s)
                GROUP BY TRIM(part_no)
                """,
                (keys,),
            )
        )

    for item in planner_try_savepoint(con, "rfq_ps", from_vouchers, default=[]) or []:
        key = normalize_part_no(item.get("part_no"))
        current = found.setdefault(
            key,
            {
                "part_no": compact_text(item.get("part_no")),
                "part_description": "",
                "op_count": 0,
                "total_ct_mins": None,
                "opns": "",
                "assignment": "",
                "machines": "",
                "source": "process_sheets",
            },
        )
        if not current.get("part_description"):
            current["part_description"] = compact_text(item.get("part_description"))
        current["ps_count"] = int(item.get("ps_count") or 0)
        current["last_order_date"] = _json_ready(item.get("last_order_date"))

    def from_preferred():
        return rows(
            con.execute(
                """
                SELECT
                    TRIM(bv.inventory_code) AS part_no,
                    COALESCE(
                        STRING_AGG(DISTINCT NULLIF(TRIM(os.preferred_machine), ''), ','
                                   ORDER BY NULLIF(TRIM(os.preferred_machine), '')),
                        ''
                    ) AS machines
                FROM public.planner_bom_variation bv
                JOIN public.planner_operation_seq os ON os.bom_id = bv.bom_id
                WHERE UPPER(TRIM(bv.inventory_code)) = ANY(%s)
                GROUP BY TRIM(bv.inventory_code)
                """,
                (keys,),
            )
        )

    for item in planner_try_savepoint(con, "rfq_mc", from_preferred, default=[]) or []:
        key = normalize_part_no(item.get("part_no"))
        current = found.get(key)
        if current and not current.get("machines"):
            current["machines"] = compact_text(item.get("machines"))

    def from_rfq_master():
        return rows(
            con.execute(
                """
                SELECT part_key, part_no, assignment, opns, machines, total_ct_mins,
                       last_rfq, customer, salesperson, sheet_tag
                FROM public.planner_rfq_part_master
                WHERE part_key = ANY(%s) OR UPPER(TRIM(part_no)) = ANY(%s)
                """,
                (normalized_keys or keys, keys),
            )
        )

    for item in planner_try_savepoint(con, "rfq_master", from_rfq_master, default=[]) or []:
        key = normalize_part_no(item.get("part_no") or item.get("part_key"))
        current = found.setdefault(
            key,
            {
                "part_no": compact_text(item.get("part_no")) or key,
                "part_description": "",
                "op_count": 0,
                "total_ct_mins": None,
                "opns": "",
                "assignment": "",
                "machines": "",
                "source": "rfq_part_master",
            },
        )
        if compact_text(item.get("assignment")):
            current["assignment"] = compact_text(item.get("assignment"))
        if compact_text(item.get("opns")):
            current["opns"] = compact_text(item.get("opns"))
        if compact_text(item.get("machines")):
            current["machines"] = compact_text(item.get("machines"))
        if item.get("total_ct_mins") not in (None, ""):
            current["total_ct_mins"] = float(item.get("total_ct_mins") or 0) or None
        current["source"] = "rfq_part_master"
        current["last_rfq"] = compact_text(item.get("last_rfq"))
        current["rfq_customer"] = compact_text(item.get("customer"))
    return found


def list_existing_parts(query: str = "", *, limit: int = 200, offset: int = 0) -> dict[str, Any]:
    needle = compact_text(query)
    like = f"%{needle}%"
    limit = max(1, min(int(limit or 200), 1000))
    offset = max(0, int(offset or 0))
    with planner_db() as con:
        _ensure_tables(con)

        def from_cycle_times():
            sql = """
                SELECT
                    TRIM(part_no) AS part_no,
                    MAX(NULLIF(TRIM(part_description), '')) AS part_description,
                    COUNT(*)::INT AS op_count,
                    SUM(COALESCE(NULLIF(cycle_time, 0), ideal_cycle_time, 0)) AS total_ct_mins,
                    COALESCE(
                        STRING_AGG(DISTINCT NULLIF(TRIM(op_type), ''), ' | '
                                   ORDER BY NULLIF(TRIM(op_type), '')),
                        ''
                    ) AS op_types
                FROM public.planner_cycle_time_master
                WHERE TRIM(part_no) <> ''
            """
            params: list[Any] = []
            if needle:
                sql += " AND (part_no ILIKE %s OR part_description ILIKE %s)"
                params.extend([like, like])
            sql += " GROUP BY TRIM(part_no) ORDER BY TRIM(part_no) LIMIT %s OFFSET %s"
            params.extend([limit, offset])
            return rows(con.execute(sql, tuple(params)))

        cycle_rows = planner_try_savepoint(con, "rfq_list_ct", from_cycle_times, default=[]) or []

        def from_vouchers():
            sql = """
                SELECT
                    TRIM(part_no) AS part_no,
                    MAX(NULLIF(TRIM(description), '')) AS part_description,
                    COUNT(DISTINCT ps_id)::INT AS ps_count,
                    MAX(order_date) AS last_order_date
                FROM public.pp_vouchers_cache
                WHERE TRIM(part_no) <> ''
            """
            params: list[Any] = []
            if needle:
                sql += " AND (part_no ILIKE %s OR description ILIKE %s)"
                params.extend([like, like])
            sql += " GROUP BY TRIM(part_no) ORDER BY MAX(order_date) DESC NULLS LAST, TRIM(part_no) LIMIT %s OFFSET %s"
            params.extend([limit, offset])
            return rows(con.execute(sql, tuple(params)))

        voucher_rows = planner_try_savepoint(con, "rfq_list_ps", from_vouchers, default=[]) or []
        by_key: dict[str, dict[str, Any]] = {}
        source_rows = cycle_rows or voucher_rows
        for item in source_rows:
            key = normalize_part_no(item.get("part_no"))
            types = [part.strip() for part in compact_text(item.get("op_types")).split("|") if part.strip()]
            by_key[key] = {
                "part_no": compact_text(item.get("part_no")),
                "part_description": compact_text(item.get("part_description")),
                "op_count": int(item.get("op_count") or 0),
                "total_ct_mins": float(item.get("total_ct_mins") or 0) or None,
                "opns": summarize_opns(types) if types else compact_text(item.get("opns")),
                "ps_count": int(item.get("ps_count") or 0),
                "last_order_date": _json_ready(item.get("last_order_date")),
                "machines": "",
            }
        if cycle_rows:
            for item in voucher_rows:
                key = normalize_part_no(item.get("part_no"))
                current = by_key.get(key)
                if not current:
                    continue
                current["ps_count"] = int(item.get("ps_count") or 0)
                current["last_order_date"] = _json_ready(item.get("last_order_date"))
                if not current.get("part_description"):
                    current["part_description"] = compact_text(item.get("part_description"))
        parts = list(by_key.values())
        parts.sort(key=lambda row: (row.get("part_no") or "").upper())
        return {"ok": True, "count": len(parts), "rows": parts, "query": needle}


def get_existing_part(part_no: str) -> dict[str, Any] | None:
    key = normalize_part_no(part_no)
    if not key:
        return None
    with planner_db() as con:
        _ensure_tables(con)
        found = lookup_existing_parts(con, [part_no, key])
        summary = found.get(key)
        if not summary:
            return None

        def ops():
            return rows(
                con.execute(
                    """
                    SELECT part_no, bom_code, stage_no, stage_name, op_no, op_type,
                           cycle_time, ideal_cycle_time, set_up_time, program_no
                    FROM public.planner_cycle_time_master
                    WHERE UPPER(TRIM(part_no)) = %s
                    ORDER BY bom_code, stage_no, op_no NULLS LAST
                    """,
                    (key,),
                )
            )

        def sheets():
            return rows(
                con.execute(
                    """
                    SELECT ps_id, part_no, description, total_qty, order_date, due_date, status, bom_code
                    FROM public.pp_vouchers_cache
                    WHERE UPPER(TRIM(part_no)) = %s
                    ORDER BY order_date DESC NULLS LAST, ps_id
                    LIMIT 40
                    """,
                    (key,),
                )
            )

        def archive_lines():
            return rows(
                con.execute(
                    """
                    SELECT l.*, b.filename, b.sheet_name, b.status AS batch_status
                    FROM public.planner_rfq_line l
                    JOIN public.planner_rfq_batch b ON b.batch_id = l.batch_id
                    WHERE UPPER(TRIM(l.part_no)) = %s
                    ORDER BY l.updated_at DESC
                    LIMIT 40
                    """,
                    (key,),
                )
            )

        def profile():
            return one(
                con.execute(
                    """
                    SELECT part_no, assignment, opns, machines, total_ct_mins,
                           last_rfq, customer, salesperson, sheet_tag, updated_at
                    FROM public.planner_rfq_part_master
                    WHERE part_key = %s
                    """,
                    (key,),
                )
            )

        summary["operations"] = [
            {k: _json_ready(v) for k, v in dict(item).items()}
            for item in (planner_try_savepoint(con, "rfq_part_ops", ops, default=[]) or [])
        ]
        summary["process_sheets"] = [
            {k: _json_ready(v) for k, v in dict(item).items()}
            for item in (planner_try_savepoint(con, "rfq_part_ps", sheets, default=[]) or [])
        ]
        summary["rfq_history"] = [
            serialize_line(item)
            for item in (planner_try_savepoint(con, "rfq_part_hist", archive_lines, default=[]) or [])
        ]
        profile_row = planner_try_savepoint(con, "rfq_part_profile", profile, default=None)
        if profile_row:
            summary["rfq_profile"] = {k: _json_ready(v) for k, v in dict(profile_row).items()}
            if summary["rfq_profile"].get("total_ct_mins") is not None:
                summary["rfq_profile"]["total_ct_mins"] = float(summary["rfq_profile"]["total_ct_mins"])
        else:
            summary["rfq_profile"] = {
                "part_no": summary.get("part_no") or key,
                "assignment": compact_text(summary.get("assignment")),
                "opns": compact_text(summary.get("opns")),
                "machines": compact_text(summary.get("machines")),
                "total_ct_mins": summary.get("total_ct_mins"),
            }
        return summary


def group_archive_batches(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[int, dict[str, Any]] = {}
    order: list[int] = []
    for item in items:
        serialized = serialize_line(item) or {}
        bid = int(serialized.get("batch_id") or item.get("batch_id") or 0)
        if not bid:
            continue
        if bid not in groups:
            order.append(bid)
            days = item.get("default_days")
            groups[bid] = {
                "batch_id": bid,
                "filename": compact_text(item.get("filename")),
                "sheet_name": compact_text(item.get("sheet_name")),
                "status": compact_text(item.get("batch_status") or item.get("status")),
                "sheet_tag": compact_text(
                    item.get("batch_sheet_tag")
                    if item.get("batch_sheet_tag") is not None
                    else item.get("sheet_tag")
                ),
                "default_rfq": compact_text(item.get("default_rfq")),
                "default_customer": compact_text(item.get("default_customer")),
                "default_salesperson": compact_text(item.get("default_salesperson")),
                "default_days": float(days) if days not in (None, "") else None,
                "default_lead_time": compact_text(item.get("default_lead_time")),
                "updated_at": _json_ready(item.get("batch_updated_at") or item.get("updated_at")),
                "created_at": _json_ready(item.get("batch_created_at")),
                "mapping_notes": compact_text(item.get("mapping_notes")),
                "lines": [],
            }
        groups[bid]["lines"].append(serialized)
    out: list[dict[str, Any]] = []
    for bid in order:
        batch = groups[bid]
        lines = batch["lines"]
        batch["line_count"] = len(lines)
        batch["new_count"] = sum(1 for line in lines if compact_text(line.get("match_status")) == "new")
        batch["matched_count"] = sum(1 for line in lines if compact_text(line.get("match_status")) == "matched")
        out.append(batch)
    return out


def list_archive(query: str = "", *, limit: int = 40) -> dict[str, Any]:
    needle = compact_text(query)
    like = f"%{needle}%"
    batch_limit = max(1, min(int(limit or 40), 80))
    with planner_db() as con:
        _ensure_tables(con)
        sql = """
            SELECT
                b.batch_id, b.filename, b.sheet_name, b.status, b.sheet_tag,
                b.default_rfq, b.default_customer, b.default_salesperson,
                b.default_days, b.default_lead_time, b.updated_at, b.created_at,
                b.mapping_notes,
                COUNT(l.line_id)::INT AS line_count,
                COUNT(l.line_id) FILTER (WHERE l.match_status = 'new')::INT AS new_count,
                COUNT(l.line_id) FILTER (WHERE l.match_status = 'matched')::INT AS matched_count
            FROM public.planner_rfq_batch b
            LEFT JOIN public.planner_rfq_line l ON l.batch_id = b.batch_id
            WHERE b.status IN ('archived', 'draft')
        """
        params: list[Any] = []
        if needle:
            sql += """
                AND (
                    b.filename ILIKE %s OR b.sheet_name ILIKE %s OR b.sheet_tag ILIKE %s
                    OR b.default_rfq ILIKE %s OR b.default_customer ILIKE %s
                    OR EXISTS (
                        SELECT 1 FROM public.planner_rfq_line x
                        WHERE x.batch_id = b.batch_id
                          AND (
                              x.part_no ILIKE %s OR x.rfq ILIKE %s OR x.customer ILIKE %s
                              OR x.salesperson ILIKE %s OR x.remark ILIKE %s
                              OR x.sheet_tag ILIKE %s OR x.lead_time ILIKE %s
                          )
                    )
                )
            """
            params.extend([like] * 12)
        sql += """
            GROUP BY b.batch_id
            ORDER BY b.updated_at DESC
            LIMIT %s
        """
        params.append(batch_limit)
        batches = []
        for item in rows(con.execute(sql, tuple(params))):
            row = serialize_batch(item) or {}
            row["line_count"] = int(item.get("line_count") or 0)
            row["new_count"] = int(item.get("new_count") or 0)
            row["matched_count"] = int(item.get("matched_count") or 0)
            row["lines"] = []
            batches.append(row)
        return {
            "ok": True,
            "count": sum(item["line_count"] for item in batches),
            "batch_count": len(batches),
            "batches": batches,
            "rows": [],
            "query": needle,
        }


def get_batch(batch_id: int, *, line_limit: int | None = None) -> dict[str, Any] | None:
    with planner_db() as con:
        _ensure_tables(con)
        batch = serialize_batch(
            one(con.execute("SELECT * FROM public.planner_rfq_batch WHERE batch_id = %s", (int(batch_id),)))
        )
        if not batch:
            return None
        counts = one(
            con.execute(
                """
                SELECT
                    COUNT(*)::INT AS line_count,
                    COUNT(*) FILTER (WHERE match_status = 'new')::INT AS new_count,
                    COUNT(*) FILTER (WHERE match_status = 'matched')::INT AS matched_count
                FROM public.planner_rfq_line
                WHERE batch_id = %s
                """,
                (int(batch_id),),
            )
        ) or {}
        sql = """
            SELECT * FROM public.planner_rfq_line
            WHERE batch_id = %s
            ORDER BY line_no, line_id
        """
        params: list[Any] = [int(batch_id)]
        if line_limit:
            sql += " LIMIT %s"
            params.append(max(1, int(line_limit)))
        raw_lines = rows(con.execute(sql, tuple(params)))
        batch["lines"] = [serialize_line(item) for item in raw_lines]
        batch["line_count"] = int(counts.get("line_count") or 0)
        batch["new_count"] = int(counts.get("new_count") or 0)
        batch["matched_count"] = int(counts.get("matched_count") or 0)
        batch["lines_truncated"] = bool(line_limit and batch["line_count"] > len(raw_lines))
        batch["field_labels"] = FIELD_LABELS
        batch["hours_per_day"] = HOURS_PER_DAY
        batch["headers"] = headers_from_source_rows(raw_lines)
        return batch


def _insert_lines(con, batch_id: int, lines: list[dict[str, Any]]) -> None:
    if not lines:
        return
    from psycopg2.extras import execute_values

    values = [
        (
            batch_id,
            int(line.get("line_no") or 0),
            compact_text(line.get("part_no")),
            compact_text(line.get("rfq")),
            compact_text(line.get("customer")),
            compact_text(line.get("salesperson")),
            normalize_sheet_tag(line.get("sheet_tag")),
            line.get("qty"),
            compact_text(line.get("opns")),
            compact_text(line.get("assignment")),
            compact_text(line.get("machines")),
            line.get("total_ct_mins"),
            line.get("machine_hours"),
            line.get("total_hours"),
            line.get("days"),
            compact_text(line.get("lead_time")),
            compact_text(line.get("need_tooling")),
            compact_text(line.get("need_fixture")),
            compact_text(line.get("remark")),
            compact_text(line.get("match_status")) or "new",
            compact_text(line.get("matched_part_no")),
            Json(line.get("source_row") or {}),
        )
        for line in lines
    ]
    cur = con._conn.cursor()
    execute_values(
        cur,
        """
        INSERT INTO public.planner_rfq_line (
            batch_id, line_no, part_no, rfq, customer, salesperson, sheet_tag, qty, opns,
            assignment, machines, total_ct_mins, machine_hours, total_hours, days,
            lead_time, need_tooling, need_fixture, remark, match_status,
            matched_part_no, source_row
        ) VALUES %s
        """,
        values,
        page_size=200,
    )


def upsert_part_master(con, lines: list[dict[str, Any]], *, batch_id: int | None = None) -> int:
    rows_in: list[tuple[Any, ...]] = []
    seen: set[str] = set()
    for line in reversed(lines or []):
        key = normalize_part_no(line.get("part_no"))
        if not key or key in seen:
            continue
        assignment = compact_text(line.get("assignment"))
        opns = compact_text(line.get("opns"))
        machines = compact_text(line.get("machines"))
        ct = _to_number(line.get("total_ct_mins"))
        if not assignment and not opns and not machines and ct is None:
            continue
        seen.add(key)
        rows_in.append(
            (
                key,
                compact_text(line.get("part_no")) or key,
                assignment,
                opns,
                machines,
                ct,
                compact_text(line.get("rfq")),
                compact_text(line.get("customer")),
                compact_text(line.get("salesperson")),
                normalize_sheet_tag(line.get("sheet_tag")),
                int(batch_id) if batch_id else None,
            )
        )
    if not rows_in:
        return 0
    from psycopg2.extras import execute_values

    cur = con._conn.cursor()
    execute_values(
        cur,
        """
        INSERT INTO public.planner_rfq_part_master (
            part_key, part_no, assignment, opns, machines, total_ct_mins,
            last_rfq, customer, salesperson, sheet_tag, source_batch_id
        ) VALUES %s
        ON CONFLICT (part_key) DO UPDATE SET
            part_no = COALESCE(NULLIF(EXCLUDED.part_no, ''), planner_rfq_part_master.part_no),
            assignment = CASE WHEN EXCLUDED.assignment <> '' THEN EXCLUDED.assignment ELSE planner_rfq_part_master.assignment END,
            opns = CASE WHEN EXCLUDED.opns <> '' THEN EXCLUDED.opns ELSE planner_rfq_part_master.opns END,
            machines = CASE WHEN EXCLUDED.machines <> '' THEN EXCLUDED.machines ELSE planner_rfq_part_master.machines END,
            total_ct_mins = COALESCE(EXCLUDED.total_ct_mins, planner_rfq_part_master.total_ct_mins),
            last_rfq = CASE WHEN EXCLUDED.last_rfq <> '' THEN EXCLUDED.last_rfq ELSE planner_rfq_part_master.last_rfq END,
            customer = CASE WHEN EXCLUDED.customer <> '' THEN EXCLUDED.customer ELSE planner_rfq_part_master.customer END,
            salesperson = CASE WHEN EXCLUDED.salesperson <> '' THEN EXCLUDED.salesperson ELSE planner_rfq_part_master.salesperson END,
            sheet_tag = CASE WHEN EXCLUDED.sheet_tag <> '' THEN EXCLUDED.sheet_tag ELSE planner_rfq_part_master.sheet_tag END,
            source_batch_id = COALESCE(EXCLUDED.source_batch_id, planner_rfq_part_master.source_batch_id),
            updated_at = NOW()
        """,
        rows_in,
        page_size=200,
    )
    return len(rows_in)


def list_part_master(query: str = "", *, limit: int = 400) -> dict[str, Any]:
    needle = compact_text(query)
    like = f"%{needle}%"
    limit = max(1, min(int(limit or 400), 2000))
    with planner_db() as con:
        _ensure_tables(con)
        sql = """
            SELECT part_no, assignment, opns, machines, total_ct_mins,
                   last_rfq, customer, salesperson, sheet_tag, updated_at
            FROM public.planner_rfq_part_master
            WHERE TRIM(part_no) <> ''
        """
        params: list[Any] = []
        if needle:
            sql += """
                AND (
                    part_no ILIKE %s OR assignment ILIKE %s OR last_rfq ILIKE %s
                    OR customer ILIKE %s OR opns ILIKE %s OR machines ILIKE %s
                )
            """
            params.extend([like, like, like, like, like, like])
        sql += " ORDER BY updated_at DESC, part_no LIMIT %s"
        params.append(limit)
        items = []
        for item in rows(con.execute(sql, tuple(params))):
            row = {key: _json_ready(value) for key, value in dict(item).items()}
            if row.get("total_ct_mins") is not None:
                row["total_ct_mins"] = float(row["total_ct_mins"])
            items.append(row)
        return {"ok": True, "count": len(items), "rows": items, "query": needle}


def create_batch_from_upload(
    *,
    filename: str,
    payload: bytes,
    sheet_name: str = "",
    use_llm: bool = True,
    sheet_tag: str = "",
    default_rfq: str = "",
    default_customer: str = "",
    default_salesperson: str = "",
) -> dict[str, Any]:
    if not payload:
        raise ValueError("The Excel file is empty.")
    if len(payload) > MAX_UPLOAD_BYTES:
        raise ValueError("Excel file is larger than 12 MB.")
    names = list_workbook_sheets(payload, filename)
    if not names:
        raise ValueError("No usable worksheet was found in that workbook.")
    sheet_summaries = [{"name": item, "row_count": 0, "headers": []} for item in names]
    chosen = compact_text(sheet_name) or pick_default_sheet(sheet_summaries)
    sheet = parse_named_sheet(payload, filename, chosen)
    for item in sheet_summaries:
        if compact_text(item["name"]) == compact_text(sheet.get("name")):
            item["row_count"] = int(sheet.get("row_count") or 0)
            item["headers"] = list(sheet.get("headers") or [])
            break
    truncated = 0
    source_rows = list(sheet.get("rows") or [])
    if len(source_rows) > MAX_RFQ_LINES:
        truncated = len(source_rows) - MAX_RFQ_LINES
        source_rows = source_rows[:MAX_RFQ_LINES]
    column_map = heuristic_column_map(sheet["headers"])
    mapping_notes = "Mapped with header aliases."
    llm_used = False
    llm_model = ""
    skip_llm = heuristic_covers_core_fields(column_map)
    if use_llm and _llm_api_key() and not skip_llm:
        try:
            llm = map_columns_with_llm(sheet["headers"], source_rows)
            if llm.get("column_map"):
                column_map = llm["column_map"]
                mapping_notes = llm.get("notes") or "Mapped with LLM."
                llm_used = True
                llm_model = llm.get("model") or _llm_model()
        except Exception as exc:
            mapping_notes = f"LLM mapping failed; used header aliases. {exc}"
            logger.warning("RFQ LLM mapping fell back to heuristic: %s", exc)
    elif use_llm and skip_llm:
        mapping_notes = (
            "Mapped with header aliases. LLM skipped because Part No. and C/T columns were already found."
        )
    elif use_llm:
        mapping_notes = "LLM is not configured; mapped with header aliases. Set RFQ_LLM_API_KEY."
    if truncated:
        mapping_notes += (
            f" Stored the first {MAX_RFQ_LINES} rows ({truncated} extra rows skipped to avoid timeout)."
        )

    part_nos = [compact_text(apply_column_map(row, column_map).get("part_no")) for row in source_rows]
    with planner_db() as con:
        _ensure_tables(con)
        existing = lookup_existing_parts(con, part_nos)
    lines = build_mapped_lines(source_rows, column_map, existing)
    if use_llm and _llm_api_key():
        try:
            extra = fill_missing_cycle_times_with_llm(lines)
            if extra:
                mapping_notes = f"{mapping_notes} {extra}".strip()
                llm_used = True
                llm_model = llm_model or _llm_model()
        except Exception as exc:
            logger.warning("RFQ LLM cycle-time fill failed: %s", exc)
            mapping_notes += f" LLM cycle-time fill failed. {exc}"
    defaults = {
        "sheet_tag": normalize_sheet_tag(sheet_tag),
        "rfq": compact_text(default_rfq),
        "customer": compact_text(default_customer),
        "salesperson": compact_text(default_salesperson),
    }
    if any(defaults.values()):
        lines = apply_defaults_to_mapped_lines(lines, defaults, overwrite=True)
    with planner_db() as con:
        try:
            con.execute("SET LOCAL statement_timeout = '45s'")
            con.execute("SET LOCAL lock_timeout = '8s'")
        except Exception:
            pass
        _ensure_tables(con)
        inserted = one(
            con.execute(
                """
                INSERT INTO public.planner_rfq_batch (
                    filename, sheet_name, status, llm_used, llm_model, mapping, mapping_notes,
                    sheet_tag, default_rfq, default_customer, default_salesperson
                )
                VALUES (%s, %s, 'draft', %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING *
                """,
                (
                    compact_text(filename),
                    sheet["name"],
                    llm_used,
                    llm_model,
                    Json(column_map),
                    mapping_notes,
                    defaults["sheet_tag"],
                    defaults["rfq"],
                    defaults["customer"],
                    defaults["salesperson"],
                ),
            )
        )
        batch_id = int(inserted["batch_id"])
        _insert_lines(con, batch_id, lines)
        upsert_part_master(con, lines, batch_id=batch_id)
    batch = get_batch(batch_id)
    assert batch is not None
    batch["sheets"] = sheet_summaries
    batch["headers"] = sheet["headers"]
    return batch


def _defaults_from_batch_row(row: dict[str, Any] | None) -> dict[str, Any]:
    data = row or {}
    return {
        "sheet_tag": normalize_sheet_tag(data.get("sheet_tag")),
        "rfq": compact_text(data.get("default_rfq") or data.get("rfq")),
        "customer": compact_text(data.get("default_customer") or data.get("customer")),
        "salesperson": compact_text(data.get("default_salesperson") or data.get("salesperson")),
        "days": _to_number(data.get("default_days") if data.get("default_days") is not None else data.get("days")),
        "lead_time": compact_text(data.get("default_lead_time") or data.get("lead_time")),
    }


def _apply_defaults_sql(con, batch_id: int, defaults: dict[str, Any], fields: list[str]) -> None:
    assignments: list[str] = []
    values: list[Any] = []
    for field in fields:
        if field == "sheet_tag":
            assignments.append("sheet_tag = %s")
            values.append(normalize_sheet_tag(defaults.get("sheet_tag")))
        elif field in BATCH_DEFAULT_FIELDS:
            assignments.append(f"{field} = %s")
            values.append(compact_text(defaults.get(field)))
        elif field == "days":
            assignments.append("days = %s")
            values.append(_to_number(defaults.get("days")))
        elif field == "lead_time":
            assignments.append("lead_time = %s")
            values.append(compact_text(defaults.get("lead_time")))
    if not assignments:
        return
    assignments.append("updated_at = NOW()")
    values.append(int(batch_id))
    con.execute(
        f"UPDATE public.planner_rfq_line SET {', '.join(assignments)} WHERE batch_id = %s",
        tuple(values),
    )


def update_batch_defaults(batch_id: int, patch: dict[str, Any]) -> dict[str, Any]:
    data = patch or {}
    changed: list[str] = []
    next_values: dict[str, Any] = {}
    if "sheet_tag" in data:
        next_values["sheet_tag"] = normalize_sheet_tag(data.get("sheet_tag"))
        changed.append("sheet_tag")
    for field in BATCH_DEFAULT_FIELDS:
        if field in data or f"default_{field}" in data:
            next_values[field] = compact_text(data.get(field) or data.get(f"default_{field}"))
            changed.append(field)
    if "days" in data or "default_days" in data:
        next_values["days"] = _to_number(
            data.get("days") if data.get("days") is not None else data.get("default_days")
        )
        changed.append("days")
    if "lead_time" in data or "default_lead_time" in data:
        next_values["lead_time"] = compact_text(data.get("lead_time") or data.get("default_lead_time"))
        changed.append("lead_time")
    if not changed:
        raise ValueError("No sheet defaults supplied.")
    with planner_db() as con:
        _ensure_tables(con)
        current = one(con.execute("SELECT * FROM public.planner_rfq_batch WHERE batch_id = %s", (int(batch_id),)))
        if not current:
            raise ValueError("RFQ batch not found.")
        merged = _defaults_from_batch_row(current)
        merged.update(next_values)
        con.execute(
            """
            UPDATE public.planner_rfq_batch
            SET sheet_tag = %s, default_rfq = %s, default_customer = %s,
                default_salesperson = %s, default_days = %s, default_lead_time = %s,
                updated_at = NOW()
            WHERE batch_id = %s
            """,
            (
                merged["sheet_tag"],
                merged["rfq"],
                merged["customer"],
                merged["salesperson"],
                merged.get("days"),
                merged.get("lead_time") or "",
                int(batch_id),
            ),
        )
        _apply_defaults_sql(con, int(batch_id), merged, changed)
    result = get_batch(int(batch_id))
    if not result:
        raise ValueError("RFQ batch not found.")
    return result


def remap_batch(batch_id: int, column_map: dict[str, str]) -> dict[str, Any]:
    cleaned = invert_field_map(column_map)
    if not cleaned:
        raise ValueError("column_map is required.")
    with planner_db() as con:
        _ensure_tables(con)
        batch = one(con.execute("SELECT * FROM public.planner_rfq_batch WHERE batch_id = %s", (int(batch_id),)))
        if not batch:
            raise ValueError("RFQ batch not found.")
        source_rows = [
            item.get("source_row") or {}
            for item in rows(
                con.execute(
                    "SELECT source_row FROM public.planner_rfq_line WHERE batch_id = %s ORDER BY line_no, line_id",
                    (int(batch_id),),
                )
            )
        ]
        source_rows = [json.loads(item) if isinstance(item, str) else item for item in source_rows]
        part_nos = [compact_text(apply_column_map(row, cleaned).get("part_no")) for row in source_rows]
        existing = lookup_existing_parts(con, part_nos)
        lines = build_mapped_lines(source_rows, cleaned, existing)
        defaults = _defaults_from_batch_row(batch)
        text_fields = ["sheet_tag", *BATCH_DEFAULT_FIELDS]
        if any(compact_text(defaults.get(field)) for field in text_fields):
            lines = apply_defaults_to_mapped_lines(lines, defaults, overwrite=True, fields=text_fields)
        con.execute("DELETE FROM public.planner_rfq_line WHERE batch_id = %s", (int(batch_id),))
        _insert_lines(con, int(batch_id), lines)
        upsert_part_master(con, lines, batch_id=int(batch_id))
        con.execute(
            """
            UPDATE public.planner_rfq_batch
            SET mapping = %s, mapping_notes = %s, updated_at = NOW()
            WHERE batch_id = %s
            """,
            (Json(cleaned), "Remapped from the column picker.", int(batch_id)),
        )
    result = get_batch(batch_id)
    if not result:
        raise ValueError("RFQ batch not found.")
    return result


def update_line(line_id: int, patch: dict[str, Any]) -> dict[str, Any]:
    data = {key: patch[key] for key in LINE_PATCH_FIELDS if key in patch}
    if not data:
        raise ValueError("No editable fields supplied.")
    if "need_tooling" in data:
        data["need_tooling"] = parse_yn(data.get("need_tooling"))
    if "need_fixture" in data:
        data["need_fixture"] = parse_yn(data.get("need_fixture"))
    for text_field in ("part_no", "rfq", "customer", "salesperson", "opns", "assignment", "machines", "lead_time", "remark"):
        if text_field in data:
            data[text_field] = compact_text(data.get(text_field))
    for num_field in ("qty", "total_ct_mins", "machine_hours", "total_hours", "days"):
        if num_field in data:
            data[num_field] = _to_number(data.get(num_field))
    with planner_db() as con:
        _ensure_tables(con)
        current = one(con.execute("SELECT * FROM public.planner_rfq_line WHERE line_id = %s", (int(line_id),)))
        if not current:
            raise ValueError("RFQ line not found.")
        merged = dict(current)
        merged.update(data)
        recalc_hours = "qty" in data or "total_ct_mins" in data
        hours_override = merged.get("total_hours") if "total_hours" in data and not recalc_hours else None
        calc = calculate_times(merged.get("qty"), merged.get("total_ct_mins"), total_hours=hours_override)
        if recalc_hours or "total_hours" in data:
            if recalc_hours:
                merged["machine_hours"] = calc["machine_hours"]
                merged["total_hours"] = calc["total_hours"]
            elif "machine_hours" not in data:
                merged["machine_hours"] = calc["machine_hours"]
        assignments = []
        values: list[Any] = []
        recalc_fields = set()
        if recalc_hours or "total_hours" in data:
            recalc_fields.update({"machine_hours", "total_hours"})
        for field in LINE_PATCH_FIELDS:
            if field in data or field in recalc_fields:
                assignments.append(f"{field} = %s")
                values.append(merged.get(field))
        if "part_no" in data:
            existing = lookup_existing_parts(con, [merged.get("part_no") or ""])
            key = normalize_part_no(merged.get("part_no"))
            hit = existing.get(key)
            merged["match_status"] = "matched" if hit else "new"
            merged["matched_part_no"] = (hit or {}).get("part_no") or compact_text(merged.get("part_no"))
            assignments.append("match_status = %s")
            values.append(merged["match_status"])
            assignments.append("matched_part_no = %s")
            values.append(merged["matched_part_no"])
        assignments.append("updated_at = NOW()")
        values.append(int(line_id))
        updated = one(
            con.execute(
                f"""
                UPDATE public.planner_rfq_line
                SET {', '.join(assignments)}
                WHERE line_id = %s
                RETURNING *
                """,
                tuple(values),
            )
        )
        con.execute(
            "UPDATE public.planner_rfq_batch SET updated_at = NOW() WHERE batch_id = %s",
            (updated["batch_id"],),
        )
        upsert_part_master(con, [merged], batch_id=int(updated["batch_id"]))
    return serialize_line(updated) or {}


def set_batch_status(batch_id: int, status: str) -> dict[str, Any]:
    status = compact_text(status).lower()
    if status not in {"draft", "archived"}:
        raise ValueError("status must be draft or archived.")
    with planner_db() as con:
        _ensure_tables(con)
        updated = one(
            con.execute(
                """
                UPDATE public.planner_rfq_batch
                SET status = %s, updated_at = NOW()
                WHERE batch_id = %s
                RETURNING *
                """,
                (status, int(batch_id)),
            )
        )
        if updated and status == "archived":
            archived_lines = rows(
                con.execute(
                    "SELECT * FROM public.planner_rfq_line WHERE batch_id = %s",
                    (int(batch_id),),
                )
            )
            upsert_part_master(con, archived_lines, batch_id=int(batch_id))
    if not updated:
        raise ValueError("RFQ batch not found.")
    return get_batch(int(batch_id)) or {}
