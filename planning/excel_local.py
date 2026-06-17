"""Read a local Excel workbook from disk and return sheet rows as JSON-serializable records."""
from __future__ import annotations

import os
import time
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from .utils import compact_text, normalize_column_name, normalize_sheet_name

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    import xlrd
except Exception:
    xlrd = None

_CACHE_TTL_SEC = 60
_CACHE_VERSION = 3
_cache: dict[tuple[str, str, int], tuple[tuple[float, float, int], dict[str, Any]]] = {}

_AKER_SN_MARKERS = {"s/n", "sn", "s_n"}
_AKER_SUB_HEADER_HINTS = {"dia", "id", "thk", "width", "length"}


def _app_root() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def resolve_excel_path(raw_path: str) -> str:
    """Resolve path relative to app root when not absolute."""
    path = compact_text(raw_path)
    if not path:
        return ""
    if not os.path.isabs(path):
        path = os.path.normpath(os.path.join(_app_root(), path))
    return os.path.normpath(path)


def configured_excel_path() -> str:
    return resolve_excel_path(os.getenv("LOCAL_EXCEL_PATH", ""))


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and value != value:
        return None
    if isinstance(value, bool):
        return value
    return value


def _workbook_format(resolved: str) -> str:
    lower = resolved.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return "openpyxl"
    if lower.endswith(".xls"):
        return "xlrd"
    raise ValueError(
        f"Unsupported workbook format for {resolved!r}. Use .xls or .xlsx."
    )


def records_from_rows(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    header_row = rows[0]
    headers = [normalize_column_name(col) for col in header_row]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or not any(cell not in (None, "") for cell in row):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            raw = row[idx] if idx < len(row) else None
            record[header] = _serialize_cell(raw)
        records.append(record)
    return records


def _openpyxl_sheet_records(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    return records_from_rows(rows)


def _xls_cell_value(sheet, book, row_idx: int, col_idx: int) -> Any:
    if xlrd is None:
        return None
    cell_type = sheet.cell_type(row_idx, col_idx)
    if cell_type == xlrd.XL_CELL_EMPTY:
        return None
    if cell_type == xlrd.XL_CELL_DATE:
        raw = sheet.cell_value(row_idx, col_idx)
        return xlrd.xldate.xldate_as_datetime(raw, book.datemode)
    if cell_type == xlrd.XL_CELL_BOOLEAN:
        return bool(sheet.cell_value(row_idx, col_idx))
    raw = sheet.cell_value(row_idx, col_idx)
    if cell_type == xlrd.XL_CELL_NUMBER and isinstance(raw, float):
        if raw == int(raw):
            return int(raw)
    return raw


def _xls_row_texts(sheet, book, row_idx: int, max_cols: int | None = None) -> list[str]:
    limit = max_cols if max_cols is not None else sheet.ncols
    return [compact_text(_xls_cell_value(sheet, book, row_idx, col_idx)) for col_idx in range(limit)]


def _forward_fill(values: list[str]) -> list[str]:
    filled: list[str] = []
    last = ""
    for value in values:
        if value:
            last = value
        filled.append(last)
    return filled


def _unique_header_key(base: str, seen: dict[str, int]) -> str:
    key = normalize_column_name(base) or "column"
    count = seen.get(key, 0) + 1
    seen[key] = count
    if count == 1:
        return key
    return f"{key}_{count}"


def _is_aker_sub_header_row(values: list[str]) -> bool:
    normalized = {normalize_column_name(value) for value in values if value}
    return bool(normalized & _AKER_SUB_HEADER_HINTS)


def _find_aker_header_layout(sheet, book) -> tuple[int, int | None, int | None, int] | None:
    scan_limit = min(sheet.nrows, 20)
    for row_idx in range(scan_limit):
        first = normalize_column_name(_xls_cell_value(sheet, book, row_idx, 0))
        if first not in _AKER_SN_MARKERS:
            continue
        sub_row = row_idx + 1 if row_idx + 1 < sheet.nrows else None
        if sub_row is not None and _is_aker_sub_header_row(_xls_row_texts(sheet, book, sub_row)):
            group_row = row_idx - 1 if row_idx >= 1 else None
            return row_idx, sub_row, group_row, row_idx + 2
        return row_idx, None, row_idx - 1 if row_idx >= 1 else None, row_idx + 1
    return None


def _build_aker_headers(
    sheet,
    book,
    header_row: int,
    sub_row: int | None,
    group_row: int | None,
) -> list[dict[str, str]]:
    max_cols = sheet.ncols
    for col_idx in range(sheet.ncols - 1, -1, -1):
        main = compact_text(_xls_cell_value(sheet, book, header_row, col_idx))
        sub = (
            compact_text(_xls_cell_value(sheet, book, sub_row, col_idx))
            if sub_row is not None
            else ""
        )
        if main or sub:
            max_cols = col_idx + 1
            break

    main_values = _xls_row_texts(sheet, book, header_row, max_cols)
    sub_values = _xls_row_texts(sheet, book, sub_row, max_cols) if sub_row is not None else [""] * max_cols
    group_values = (
        _forward_fill(_xls_row_texts(sheet, book, group_row, max_cols))
        if group_row is not None
        else [""] * max_cols
    )
    main_values = _forward_fill(main_values)

    columns: list[dict[str, str]] = []
    seen: dict[str, int] = {}
    for col_idx in range(max_cols):
        main = main_values[col_idx]
        sub = sub_values[col_idx]
        group = group_values[col_idx]

        if sub:
            label = f"{main} {sub}".strip() if main else sub
        else:
            label = main or f"Column {col_idx + 1}"

        group_norm = normalize_column_name(group)
        if group_norm and "material_transaction_in" in group_norm:
            label = f"In {label}"
        elif group_norm and "material_transaction_out" in group_norm:
            label = f"Out {label}"

        key = _unique_header_key(label, seen)
        columns.append({"key": key, "label": label})
    return columns


def _extract_aker_sheet_meta(sheet, book) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row_idx in range(min(6, sheet.nrows)):
        label = compact_text(_xls_cell_value(sheet, book, row_idx, 0))
        if label.upper().startswith("MATERIAL TYPE"):
            for col_idx in range(1, min(8, sheet.ncols)):
                value = compact_text(_xls_cell_value(sheet, book, row_idx, col_idx))
                if value:
                    meta["material_type"] = value
                    break
        if label.upper().startswith("COLOR CODE"):
            for col_idx in range(1, min(8, sheet.ncols)):
                value = compact_text(_xls_cell_value(sheet, book, row_idx, col_idx))
                if value:
                    meta["color_code"] = value
                    break
    title = compact_text(_xls_cell_value(sheet, book, 0, 0))
    if title:
        meta["title"] = title
    return meta


def _parse_aker_xls_sheet(sheet, book) -> tuple[list[dict[str, str]], list[dict[str, Any]], dict[str, str]]:
    layout = _find_aker_header_layout(sheet, book)
    meta = _extract_aker_sheet_meta(sheet, book)
    if layout is None:
        seen: dict[str, int] = {}
        columns = []
        for col_idx in range(sheet.ncols):
            label = compact_text(_xls_cell_value(sheet, book, 0, col_idx)) or f"Column {col_idx + 1}"
            columns.append({"key": _unique_header_key(label, seen), "label": label})
        records: list[dict[str, Any]] = []
        for row_idx in range(1, sheet.nrows):
            row_values = [_xls_cell_value(sheet, book, row_idx, col_idx) for col_idx in range(sheet.ncols)]
            if not any(value not in (None, "") for value in row_values):
                continue
            record: dict[str, Any] = {}
            for col_idx, col in enumerate(columns):
                if col_idx >= len(row_values):
                    break
                record[col["key"]] = _serialize_cell(row_values[col_idx])
            records.append(record)
        return columns, records, meta

    header_row, sub_row, group_row, data_start = layout
    columns = _build_aker_headers(sheet, book, header_row, sub_row, group_row)
    records: list[dict[str, Any]] = []
    for row_idx in range(data_start, sheet.nrows):
        row_values = [_xls_cell_value(sheet, book, row_idx, col_idx) for col_idx in range(sheet.ncols)]
        if not any(value not in (None, "") for value in row_values):
            continue
        record: dict[str, Any] = {}
        for col_idx, col in enumerate(columns):
            if col_idx >= len(row_values):
                break
            record[col["key"]] = _serialize_cell(row_values[col_idx])
        records.append(record)
    return columns, records, meta


def _xls_sheet_payload(sheet, book) -> dict[str, Any]:
    columns, rows, meta = _parse_aker_xls_sheet(sheet, book)
    return {"columns": columns, "rows": rows, "meta": meta}


def _xls_sheet_records(sheet, book) -> list[dict[str, Any]]:
    return _xls_sheet_payload(sheet, book)["rows"]


def _resolve_sheet_name(sheet_names: list[str], sheet: str | None) -> str:
    sheet_map = {normalize_sheet_name(name): name for name in sheet_names}
    if sheet:
        key = normalize_sheet_name(sheet)
        if key not in sheet_map:
            available = ", ".join(sheet_names)
            raise ValueError(f"Sheet {sheet!r} not found. Available: {available}")
        return sheet_map[key]
    return sheet_names[0]


def _read_openpyxl_workbook(
    resolved: str,
    *,
    sheet: str | None = None,
    all_sheets: bool = False,
) -> dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    workbook = load_workbook(resolved, data_only=True, read_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        if all_sheets:
            sheets = {name: _openpyxl_sheet_records(workbook[name]) for name in sheet_names}
            return {
                "path": resolved,
                "format": "xlsx",
                "sheet_names": sheet_names,
                "sheets": sheets,
            }

        sheet_name = _resolve_sheet_name(sheet_names, sheet)
        rows = _openpyxl_sheet_records(workbook[sheet_name])
        return {
            "path": resolved,
            "format": "xlsx",
            "sheet": sheet_name,
            "sheet_names": sheet_names,
            "count": len(rows),
            "rows": rows,
        }
    finally:
        workbook.close()


def _read_xls_workbook(
    resolved: str,
    *,
    sheet: str | None = None,
    all_sheets: bool = False,
) -> dict[str, Any]:
    if xlrd is None:
        raise RuntimeError("xlrd is not installed. Run: pip install xlrd")

    book = xlrd.open_workbook(resolved, formatting_info=False)
    sheet_names = book.sheet_names()
    if all_sheets:
        sheets: dict[str, Any] = {}
        sheet_meta: dict[str, Any] = {}
        for name in sheet_names:
            payload = _xls_sheet_payload(book.sheet_by_name(name), book)
            sheets[name] = payload["rows"]
            sheet_meta[name] = {
                "columns": payload["columns"],
                "meta": payload["meta"],
                "count": len(payload["rows"]),
            }
        return {
            "path": resolved,
            "format": "xls",
            "sheet_names": sheet_names,
            "sheets": sheets,
            "sheet_meta": sheet_meta,
        }

    sheet_name = _resolve_sheet_name(sheet_names, sheet)
    payload = _xls_sheet_payload(book.sheet_by_name(sheet_name), book)
    return {
        "path": resolved,
        "format": "xls",
        "sheet": sheet_name,
        "sheet_names": sheet_names,
        "count": len(payload["rows"]),
        "rows": payload["rows"],
        "columns": payload["columns"],
        "meta": payload["meta"],
    }


def read_workbook(
    path: str,
    *,
    sheet: str | None = None,
    all_sheets: bool = False,
) -> dict[str, Any]:
    resolved = resolve_excel_path(path)
    if not resolved:
        raise ValueError("Excel path is empty")
    if not os.path.isfile(resolved):
        raise FileNotFoundError(f"Excel file not found: {resolved}")

    fmt = _workbook_format(resolved)
    if fmt == "xlrd":
        return _read_xls_workbook(resolved, sheet=sheet, all_sheets=all_sheets)
    return _read_openpyxl_workbook(resolved, sheet=sheet, all_sheets=all_sheets)


def read_workbook_cached(
    path: str,
    *,
    sheet: str | None = None,
    all_sheets: bool = False,
    refresh: bool = False,
    cache_ttl_sec: int = _CACHE_TTL_SEC,
) -> dict[str, Any]:
    resolved = resolve_excel_path(path)
    if not resolved:
        raise ValueError("Excel path is empty")

    mtime_ns = os.path.getmtime(resolved)
    cache_key = (resolved, sheet or "", 1 if all_sheets else 0)
    now = time.time()
    cached = _cache.get(cache_key)
    if (
        not refresh
        and cached
        and cached[0][0] == mtime_ns
        and cached[0][1] == _CACHE_VERSION
        and now - cached[0][2] < cache_ttl_sec
    ):
        return cached[1]

    payload = read_workbook(resolved, sheet=sheet, all_sheets=all_sheets)
    _cache[cache_key] = ((mtime_ns, _CACHE_VERSION, now), payload)
    return payload
