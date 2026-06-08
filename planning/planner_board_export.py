"""Server-side machine board Excel export — mirrors static/js/scheduler/export.js."""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .helpers import planner_db, rows
from .materials import material_status_map_for_ps_ids
from .process_sheets import (
    due_date_map_for_planner_ps_ids,
    format_planner_ps_id,
    material_in_map_for_planner_ps_ids,
    parse_planner_ps_id,
)
from .queue_visibility import filter_completed_lane_blocks
from .blocks import attach_block_ps_identity, planner_ps_id_from_block_row
from .utils import compact_text, planner_wall_datetime_to_api


def _attach_board_meta_to_blocks(con, blocks):
    if not blocks:
        return
    board_ps_ids = list(
        dict.fromkeys(
            planner_ps_id_from_block_row(row)
            for row in blocks
            if planner_ps_id_from_block_row(row)
        )
    )
    if not board_ps_ids:
        return
    material_in_by_ps = material_in_map_for_planner_ps_ids(con, board_ps_ids)
    due_date_by_ps = due_date_map_for_planner_ps_ids(con, board_ps_ids)
    for row in blocks:
        ps_id = planner_ps_id_from_block_row(row)
        if not ps_id:
            continue
        row["planner_ps_id"] = ps_id
        row["material_in"] = bool(material_in_by_ps.get(ps_id))
        due_text = compact_text(due_date_by_ps.get(ps_id))
        if due_text:
            row["due_date"] = due_text

MACHINE_BOARD_GROUPS = [
    {
        "id": "mpp",
        "label": "MPP",
        "machine_codes": ["CNC 35", "CNC 36"],
    },
    {
        "id": "multiaxis",
        "label": "Multi-Axis",
        "machine_codes": ["CNC 38", "CNC 39", "CNC 40"],
    },
    {
        "id": "turning",
        "label": "Turning",
        "machine_codes": ["CNC 22", "CNC 30", "CNC 31", "CNC 32", "CNC 10", "CNC 15", "CNC 21", "CNC 24", "CNC 27"],
    },
    {
        "id": "milling",
        "label": "Milling",
        "machine_codes": ["CNC 20", "CNC 29", "CNC 25", "CNC 26"],
    },
]

GROUP_ORDER = ["mpp", "multiaxis", "turning", "milling"]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _normalize_machine_code(code: str) -> str:
    return compact_text(code).upper()


def _format_due_for_export(date_text: str) -> str:
    raw = compact_text(date_text)[:10]
    if not raw:
        return ""
    try:
        d = date.fromisoformat(raw)
    except ValueError:
        return raw
    yy = str(d.year)[-2:]
    return f"{d.day} {MONTHS[d.month - 1]} {yy}"


def _format_schedule_for_export(date_time_text: str) -> str:
    raw = compact_text(date_time_text)
    if not raw:
        return ""
    date_fmt = _format_due_for_export(raw[:10])
    if not date_fmt:
        return raw
    if len(raw) >= 16 and raw[10] in {" ", "T"}:
        return f"{date_fmt} {raw[11:16]}"
    return date_fmt


def _block_queued_at(block: dict[str, Any]) -> str:
    return compact_text(
        block.get("visual_start_datetime")
        or block.get("calculated_start_datetime")
        or block.get("predicted_start_at")
        or block.get("anchor_datetime")
    )


def _block_output_at(block: dict[str, Any]) -> str:
    if compact_text(block.get("actual_end_at")):
        return compact_text(block.get("actual_end_at"))
    return compact_text(
        block.get("predicted_end_at")
        or block.get("calculated_end_datetime")
        or block.get("visual_end_datetime")
    )


def _ps_display(block: dict[str, Any], group_ps_id: str = "") -> tuple[str, str]:
    planner_ps_id = compact_text(block.get("planner_ps_id") or group_ps_id or block.get("source_ps_id") or block.get("job_no"))
    if planner_ps_id.startswith("[Temp]"):
        return planner_ps_id, ""
    base, partial = parse_planner_ps_id(planner_ps_id)
    if not base:
        return planner_ps_id, ""
    partial_text = str(int(partial or 1)) if int(partial or 1) > 1 else ""
    return base, partial_text


def _group_blocks_for_machine(blocks: list[dict[str, Any]], machine_id: int) -> list[dict[str, Any]]:
    machine_blocks = [
        block for block in blocks if int(block.get("machine_id") or 0) == int(machine_id)
    ]
    if not machine_blocks:
        return []

    by_key: dict[str, list[dict[str, Any]]] = {}
    for block in machine_blocks:
        group_id = int(block.get("group_id") or 0)
        block_id = int(block.get("block_id") or 0)
        key = f"g:{group_id}" if group_id > 0 else f"s:{block_id}"
        by_key.setdefault(key, []).append(block)

    groups: list[dict[str, Any]] = []
    for members in by_key.values():
        members.sort(
            key=lambda row: (
                int(row.get("queue_position") or 0),
                int(row.get("block_id") or 0),
            )
        )
        leader = members[0]
        group_id = int(leader.get("group_id") or 0)
        ps_id = compact_text(leader.get("planner_ps_id") or leader.get("source_ps_id") or leader.get("job_no"))
        operation_label = compact_text(leader.get("group_label"))
        if not operation_label and len(members) > 1:
            operation_label = " & ".join(
                compact_text(row.get("source_op_no") or row.get("operation_name"))
                for row in members
                if compact_text(row.get("source_op_no") or row.get("operation_name"))
            )
        if not operation_label:
            operation_label = compact_text(leader.get("source_op_no") or leader.get("operation_name"))
        target_qty = max(float(row.get("scheduled_qty") or 0) for row in members)
        groups.append(
            {
                "group_id": group_id,
                "ps_id": ps_id,
                "operation_label": operation_label,
                "leader": leader,
                "blocks": members,
                "target_qty": target_qty,
                "material_status": leader.get("material_status") or {},
            }
        )

    groups.sort(
        key=lambda group: (
            int((group.get("leader") or {}).get("queue_position") or 0),
            int((group.get("leader") or {}).get("block_id") or 0),
        )
    )
    return groups


def _machine_availability_end(groups: list[dict[str, Any]]) -> str:
    ends = [_block_output_at(group.get("leader") or {}) for group in groups]
    ends = [value for value in ends if value]
    return max(ends) if ends else ""


def _build_cell_lines(
    group: dict[str, Any],
    *,
    due_by_ps: dict[str, str],
    material_by_ps: dict[str, dict[str, Any]],
) -> list[str]:
    leader = group.get("leader") or {}
    base, partial = _ps_display(leader, compact_text(group.get("ps_id")))
    ps_display = base if not partial else f"{base} P{partial}"
    ps_due_key = format_planner_ps_id(base, int(partial or 1)) if base else compact_text(group.get("ps_id"))

    op_line = compact_text(group.get("operation_label"))
    if len(group.get("blocks") or []) == 1:
        op_no = compact_text(leader.get("source_op_no"))
        op_name = compact_text(leader.get("operation_name"))
        op_line = " ".join(part for part in (op_no, op_name) if part).strip() or op_line

    detail_parts: list[str] = []
    part_name = compact_text(leader.get("part_name") or leader.get("part_no"))
    if part_name:
        detail_parts.append(part_name)
    elif op_line:
        detail_parts.append(op_line)
    target_qty = float(group.get("target_qty") or leader.get("scheduled_qty") or 0)
    if target_qty > 0:
        qty_text = str(int(target_qty)) if target_qty == int(target_qty) else str(target_qty)
        detail_parts.append(f"{qty_text}ea")

    due_raw = compact_text(due_by_ps.get(ps_due_key) or leader.get("due_date"))
    due_fmt = _format_due_for_export(due_raw)
    start_fmt = _format_schedule_for_export(_block_queued_at(leader))
    end_fmt = _format_schedule_for_export(_block_output_at(leader))

    lines = [
        ps_display or compact_text(group.get("ps_id")),
        ", ".join(detail_parts),
        f"Start {start_fmt}" if start_fmt else "",
        f"End {end_fmt}" if end_fmt else "",
        f"Due Date {due_fmt}" if due_fmt else "",
    ]
    lines = [line for line in lines if compact_text(line)]

    mat = material_by_ps.get(ps_due_key) or group.get("material_status") or {}
    mat_label = compact_text(mat.get("label"))
    if mat_label:
        lines.append(mat_label)
    return lines


def _export_machine_columns(machines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_code = {_normalize_machine_code(row.get("machine_code")): row for row in machines}
    assigned: set[str] = set()
    columns: list[dict[str, Any]] = []

    for spec in MACHINE_BOARD_GROUPS:
        for code in spec.get("machine_codes") or []:
            key = _normalize_machine_code(code)
            machine = by_code.get(key)
            if not machine:
                continue
            columns.append({"machine": machine, "group_label": spec.get("label") or ""})
            assigned.add(key)

    for machine in machines:
        key = _normalize_machine_code(machine.get("machine_code"))
        if key in assigned:
            continue
        columns.append({"machine": machine, "group_label": ""})

    return columns


def _fetch_board_blocks(con) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, str], dict[str, dict[str, Any]]]:
    machines = rows(
        con.execute(
            """
            SELECT machine_id, machine_no AS machine_code, machine_category, shift_profile, active
            FROM planner_machines
            WHERE active = TRUE
            ORDER BY machine_id
            """
        )
    )
    raw_blocks = rows(
        con.execute(
            """
            SELECT b.*, o.job_no, o.operation_name, o.total_qty, o.setup_minutes, o.cycle_minutes_per_qty,
                   o.compatible_machine_group, o.source_ps_id, o.source_op_seq_id AS source_op_seq_id, o.source_op_no,
                   m.machine_no AS machine_code, m.machine_category, m.shift_profile,
                   g.group_label AS group_label, g.group_type AS group_type,
                   qs.predicted_start_at AS qs_predicted_start_at,
                   qs.predicted_end_at AS qs_predicted_end_at,
                   qs.remaining_qty AS qs_remaining_qty,
                   qs.output_qty AS qs_output_qty,
                   qs.good_qty AS qs_good_qty,
                   qs.reject_qty AS qs_reject_qty,
                   qs.schedule_status AS qs_schedule_status
            FROM planner_run_block b
            JOIN planner_operation o ON o.operation_id = b.operation_id
            JOIN planner_machines m ON m.machine_id = b.machine_id
            LEFT JOIN planner_run_block_group g ON g.group_id = b.group_id
            LEFT JOIN planner_machine_queue_state qs ON qs.block_id = b.block_id
            WHERE COALESCE(b.active, TRUE) = TRUE
            ORDER BY b.machine_id, b.queue_position, b.block_id
            """
        )
    )

    blocks: list[dict[str, Any]] = []
    for row in raw_blocks:
        item = dict(row)
        calc_start = planner_wall_datetime_to_api(item.get("calculated_start_datetime"))
        calc_end = planner_wall_datetime_to_api(item.get("calculated_end_datetime"))
        item["calculated_start_datetime"] = calc_start
        item["calculated_end_datetime"] = calc_end
        pred_start = planner_wall_datetime_to_api(item.get("qs_predicted_start_at")) or calc_start
        pred_end = planner_wall_datetime_to_api(item.get("qs_predicted_end_at")) or calc_end
        item["predicted_start_at"] = pred_start
        item["predicted_end_at"] = pred_end
        item["visual_start_datetime"] = pred_start
        item["visual_end_datetime"] = pred_end
        for drop_key in (
            "qs_predicted_start_at",
            "qs_predicted_end_at",
            "qs_remaining_qty",
            "qs_output_qty",
            "qs_good_qty",
            "qs_reject_qty",
            "qs_schedule_status",
        ):
            item.pop(drop_key, None)
        blocks.append(item)

    attach_block_ps_identity(con, blocks)
    _attach_board_meta_to_blocks(con, blocks)
    blocks = filter_completed_lane_blocks(con, blocks)

    board_ps_ids = list(
        dict.fromkeys(
            planner_ps_id_from_block_row(row)
            for row in blocks
            if planner_ps_id_from_block_row(row)
        )
    )
    due_by_ps = due_date_map_for_planner_ps_ids(con, board_ps_ids) if board_ps_ids else {}
    material_by_ps = material_status_map_for_ps_ids(con, board_ps_ids) if board_ps_ids else {}
    for row in blocks:
        ps_id = planner_ps_id_from_block_row(row)
        if ps_id and ps_id in material_by_ps:
            row["material_status"] = material_by_ps[ps_id]

    return machines, blocks, due_by_ps, material_by_ps


def build_planner_board_workbook_bytes(*, snapshot_date: date | None = None) -> tuple[bytes, str]:
    """Return (xlsx bytes, filename) for the current machine board snapshot."""
    snapshot_date = snapshot_date or datetime.now().date()
    with planner_db() as con:
        machines, blocks, due_by_ps, material_by_ps = _fetch_board_blocks(con)

    columns = _export_machine_columns(machines)
    if not columns:
        raise ValueError("No active machines to export")

    queues = []
    for col in columns:
        machine_id = int(col["machine"]["machine_id"])
        groups = _group_blocks_for_machine(blocks, machine_id)
        queues.append({"machine": col["machine"], "groups": groups})
    max_depth = max(1, max(len(queue["groups"]) for queue in queues))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Machine Board"

    thin = Side(style="thin", color="FFD0D5DD")
    thin_border = Border(top=thin, left=thin, bottom=thin, right=thin)
    header_fill = PatternFill("solid", fgColor="FFF3F4F6")

    group_row = sheet.row_dimensions[1]
    machine_row = sheet.row_dimensions[2]
    group_row.height = 22
    machine_row.height = 20

    group_spans: list[dict[str, Any]] = []
    col_index = 1
    for idx, col in enumerate(columns):
        label = col.get("group_label") or ""
        prev = columns[idx - 1] if idx > 0 else None
        if not label:
            group_spans.append({"start": col_index, "end": col_index, "label": ""})
        elif prev and prev.get("group_label") == label:
            group_spans[-1]["end"] = col_index
        else:
            group_spans.append({"start": col_index, "end": col_index, "label": label})

        machine_cell = sheet.cell(row=2, column=col_index, value=compact_text(col["machine"].get("machine_code")))
        machine_cell.font = Font(bold=True, size=11, color="FF1A1C1D")
        machine_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        machine_cell.fill = header_fill
        machine_cell.border = thin_border
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = 22
        col_index += 1

    for span in group_spans:
        if span["end"] > span["start"]:
            sheet.merge_cells(start_row=1, start_column=span["start"], end_row=1, end_column=span["end"])
        cell = sheet.cell(row=1, column=span["start"], value=span.get("label") or "")
        cell.font = Font(bold=True, size=11, color="FF4B5563")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = thin_border

    availability_row = 3
    sheet.row_dimensions[availability_row].height = 18
    for col_idx, queue in enumerate(queues, start=1):
        availability_end = _machine_availability_end(queue["groups"])
        cell = sheet.cell(row=availability_row, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(size=9, color="FF6B7280")
        cell.fill = header_fill
        if availability_end:
            cell.value = f"Queue ends {_format_schedule_for_export(availability_end)}"

    for depth in range(max_depth):
        row_index = 4 + depth
        sheet.row_dimensions[row_index].height = 84
        for col_idx, queue in enumerate(queues, start=1):
            group = queue["groups"][depth] if depth < len(queue["groups"]) else None
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.font = Font(size=10, color="FF1A1C1D")
            if not group:
                cell.value = ""
                continue
            lines = _build_cell_lines(group, due_by_ps=due_by_ps, material_by_ps=material_by_ps)
            cell.value = "\n".join(lines)

    sheet.freeze_panes = "A4"
    workbook.properties.creator = "Production Planner"
    workbook.properties.created = datetime.now()

    buffer = io.BytesIO()
    workbook.save(buffer)
    filename = f"machine-production-board-{snapshot_date.isoformat()}.xlsx"
    return buffer.getvalue(), filename
