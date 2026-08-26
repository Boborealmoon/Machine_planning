"""First Article Tracker - flagged process sheets, PIC roster, S/O live fields."""
from __future__ import annotations

import io
import json
import logging
import re
import threading
from datetime import date, datetime
from typing import Any

from db import planner_db_connect_error
from .helpers import one, planner_db, rows
from .staged_erp import serialize_row
from .utils import compact_text

logger = logging.getLogger(__name__)

CHECK_TEXT_MODES = ("tick", "text")
CHECK_TEXT_FIELDS = ("tooling", "fixture", "gauges")
_SEARCH_LIMIT = 25
_CANDIDATE_LIMIT = 1500
_BULK_FLAG_LIMIT = 200
_IMPORT_LIMIT = 500
_MAX_IMPORT_BYTES = 12 * 1024 * 1024
IMPORT_TEMPLATE_COLUMNS = (
    "Data Input",
    "Part No.",
    "Part Description",
    "Total Qty",
    "PO Due Date",
    "Machine (CNC)",
    "PIC",
    "Tooling",
    "Fixture/Jig",
    "Gauges/CMM",
    "Remark",
)
_IMPORT_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "process_sheet_no": (
        "datainput", "datainputprocesssheetno", "datainputprocesssheet",
        "processsheetno", "processsheet", "processsheetnumber",
        "psno", "psnumber", "psn",
    ),
    "part_no": ("partno", "partnumber"),
    "part_description": ("partdescription", "description", "partdesc"),
    "total_qty": ("totalqty", "qty", "quantity"),
    "po_due_date": ("poduedate", "duedate"),
    "machine_codes": ("machinecnc", "machine", "machines", "cnc", "cncmachine"),
    "pic_names": ("pic", "personincharge", "pics"),
    "tooling": ("tooling", "tools", "tool"),
    "fixture": ("fixturejig", "fixture", "jig"),
    "gauges": ("gaugescmm", "gauges", "cmm", "gauge", "gaugecmm"),
    "remarks": ("remark", "remarks", "notes", "note", "comment"),
}
_IMPORT_PATCH_FIELDS = ("machine_codes", "pic_names", "tooling", "fixture", "gauges", "remarks")
_PIC_SPLIT_RE = re.compile(r"[/,;|]+")
_MACHINE_NUM_RE = re.compile(r"(\d+)\s*$")
_CHECK_READY_VALUES = frozenset({
    "ok", "okay", "ready", "yes", "y", "done", "complete", "completed", "true", "tick",
})
_BLANK_CELL_VALUES = frozenset({"", "-", "—", "–", "none", "nil", "null"})
_MONTH_NAME_TO_NUM = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
    "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
_PS_TYPE_ORDER = ("APS", "NPS", "MPS", "PPS", "CPS", "SR", "OTHER")
_ROW_SELECT = """
    first_article_id, process_sheet_no, pp_voucher_no, pic_ids, machine_codes,
    tooling_mode, tooling_tick, tooling_text,
    fixture_mode, fixture_tick, fixture_text,
    gauges_mode, gauges_tick, gauges_text,
    remarks, created_at, updated_at
"""
_HISTORY_STATUSES = frozenset({
    "history", "completed", "complete", "closed", "h", "c",
})
_STAGE_STATUS_LABELS = {"I": "In process", "R": "Released", "P": "Pending", "C": "Completed"}


_NEW_PART_ROW_SELECT = """
    process_sheet_no, pp_voucher_no, bom_updated, remarks, program_finish_at,
    program_pic_ids, is_exception, created_at, updated_at
"""
_NEW_PART_PATCH_FIELDS = ("bom_updated", "remarks", "program_finish_at", "program_pic_ids")
_HISTORY_SOURCES = ("new_part", "flagged")
_HISTORY_FIELDS_NEW_PART = ("remarks", "program_finish_at", "program_pic_ids")
_HISTORY_FIELDS_FLAGGED = ("remarks", "pic_ids")
_HISTORY_FIELD_LABELS = {
    "remarks": "Remarks",
    "program_finish_at": "Programme estimated finish",
    "program_pic_ids": "Programme PIC",
    "pic_ids": "PIC",
}
_HISTORY_LIMIT = 200
_SCHEMA_LOCK_KEY = 874512031
_REQUIRED_COLUMNS = (
    ("planner_first_article_pic", "pic_id"),
    ("planner_first_article", "machine_codes"),
    ("planner_first_article_new_part", "program_pic_ids"),
    ("planner_first_article_new_part", "is_exception"),
    ("planner_first_article_change_log", "change_id"),
)
_tables_ready = False
_tables_lock = threading.Lock()


def _schema_complete(con) -> bool:
    fetched = rows(
        con.execute(
            """
            SELECT table_name, column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND (
                    (table_name = 'planner_first_article_pic' AND column_name = 'pic_id')
                 OR (table_name = 'planner_first_article' AND column_name = 'machine_codes')
                 OR (table_name = 'planner_first_article_new_part'
                     AND column_name IN ('program_pic_ids', 'is_exception'))
                 OR (table_name = 'planner_first_article_change_log' AND column_name = 'change_id')
              )
            """
        )
    )
    have = {
        (compact_text(item.get("table_name")), compact_text(item.get("column_name")))
        for item in fetched
    }
    return all(pair in have for pair in _REQUIRED_COLUMNS)


def _add_column_if_missing(con, table: str, column: str, ddl: str) -> None:
    found = one(
        con.execute(
            """
            SELECT 1 AS ok
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = %s
              AND column_name = %s
            """,
            (table, column),
        )
    )
    if found:
        return
    con.execute(ddl)


def _try_commit(con) -> None:
    commit = getattr(con, "commit", None)
    if callable(commit):
        commit()


def _ensure_tables(con) -> None:
    global _tables_ready
    if _tables_ready:
        return
    with _tables_lock:
        if _tables_ready:
            return
        if _schema_complete(con):
            _tables_ready = True
            return
        cur = con.execute("SELECT pg_advisory_xact_lock(%s)", (_SCHEMA_LOCK_KEY,))
        fetchone = getattr(cur, "fetchone", None)
        if callable(fetchone):
            fetchone()
        if _schema_complete(con):
            _tables_ready = True
            return
        _migrate_first_article_schema(con)
        _try_commit(con)
        _tables_ready = True


def _migrate_first_article_schema(con) -> None:
    con.execute("SET LOCAL lock_timeout = '5s'")
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_first_article_pic (
            pic_id     BIGSERIAL    PRIMARY KEY,
            name       TEXT         NOT NULL,
            active     BOOLEAN      NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ  NOT NULL DEFAULT NOW()
        )
        """
    )
    con.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_fa_pic_name_active_unique
            ON public.planner_first_article_pic (LOWER(TRIM(name)))
            WHERE active = TRUE
        """
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_first_article (
            first_article_id BIGSERIAL    PRIMARY KEY,
            process_sheet_no TEXT         NOT NULL,
            pp_voucher_no    TEXT         NOT NULL DEFAULT '',
            pic_ids          BIGINT[]     NOT NULL DEFAULT '{}',
            tooling_mode     TEXT         NOT NULL DEFAULT 'tick',
            tooling_tick     BOOLEAN      NOT NULL DEFAULT FALSE,
            tooling_text     TEXT         NOT NULL DEFAULT '',
            fixture_mode     TEXT         NOT NULL DEFAULT 'tick',
            fixture_tick     BOOLEAN      NOT NULL DEFAULT FALSE,
            fixture_text     TEXT         NOT NULL DEFAULT '',
            gauges_mode      TEXT         NOT NULL DEFAULT 'tick',
            gauges_tick      BOOLEAN      NOT NULL DEFAULT FALSE,
            gauges_text      TEXT         NOT NULL DEFAULT '',
            remarks          TEXT         NOT NULL DEFAULT '',
            machine_codes    TEXT[],
            created_at       TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            updated_at       TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            CONSTRAINT planner_first_article_tooling_mode_chk
                CHECK (tooling_mode IN ('tick', 'text')),
            CONSTRAINT planner_first_article_fixture_mode_chk
                CHECK (fixture_mode IN ('tick', 'text')),
            CONSTRAINT planner_first_article_gauges_mode_chk
                CHECK (gauges_mode IN ('tick', 'text'))
        )
        """
    )
    con.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_fa_process_sheet_unique
            ON public.planner_first_article (LOWER(TRIM(process_sheet_no)))
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_fa_updated_at
            ON public.planner_first_article (updated_at DESC)
        """
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_first_article_new_part (
            process_sheet_no   TEXT         PRIMARY KEY,
            pp_voucher_no      TEXT         NOT NULL DEFAULT '',
            bom_updated        BOOLEAN      NOT NULL DEFAULT FALSE,
            remarks            TEXT         NOT NULL DEFAULT '',
            program_finish_at  TEXT         NOT NULL DEFAULT '',
            program_pic_ids    BIGINT[]     NOT NULL DEFAULT '{}',
            is_exception       BOOLEAN      NOT NULL DEFAULT FALSE,
            created_at         TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            updated_at         TIMESTAMPTZ  NOT NULL DEFAULT NOW()
        )
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_fa_new_part_updated_at
            ON public.planner_first_article_new_part (updated_at DESC)
        """
    )
    _add_column_if_missing(
        con,
        "planner_first_article_new_part",
        "program_pic_ids",
        """
        ALTER TABLE public.planner_first_article_new_part
            ADD COLUMN IF NOT EXISTS program_pic_ids BIGINT[] NOT NULL DEFAULT '{}'
        """,
    )
    _add_column_if_missing(
        con,
        "planner_first_article_new_part",
        "is_exception",
        """
        ALTER TABLE public.planner_first_article_new_part
            ADD COLUMN IF NOT EXISTS is_exception BOOLEAN NOT NULL DEFAULT FALSE
        """,
    )
    _add_column_if_missing(
        con,
        "planner_first_article",
        "machine_codes",
        """
        ALTER TABLE public.planner_first_article
            ADD COLUMN IF NOT EXISTS machine_codes TEXT[]
        """,
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS public.planner_first_article_change_log (
            change_id          BIGSERIAL    PRIMARY KEY,
            source             TEXT         NOT NULL,
            process_sheet_no   TEXT         NOT NULL,
            first_article_id   BIGINT,
            field_name         TEXT         NOT NULL,
            old_value          TEXT         NOT NULL DEFAULT '',
            new_value          TEXT         NOT NULL DEFAULT '',
            changed_at         TIMESTAMPTZ  NOT NULL DEFAULT NOW(),
            CONSTRAINT planner_first_article_change_log_source_chk
                CHECK (source IN ('new_part', 'flagged'))
        )
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_fa_change_log_ps_at
            ON public.planner_first_article_change_log (LOWER(TRIM(process_sheet_no)), changed_at DESC)
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_fa_change_log_source_ps
            ON public.planner_first_article_change_log (source, LOWER(TRIM(process_sheet_no)), changed_at DESC)
        """
    )


def _ps_base(value: Any) -> str:
    return compact_text(value).split("::")[0]


def _ps_key(value: Any) -> str:
    return _ps_base(value).upper()


def _job_ps_type(job: dict[str, Any] | None) -> str:
    from .so_outstanding_balance_service import ps_type

    if not job:
        return ""
    return ps_type(job.get("process_sheet_no") or job.get("pp_voucher_no"))


def _job_search_blob(job: dict[str, Any]) -> str:
    return " ".join(
        compact_text(job.get(field)).upper()
        for field in (
            "process_sheet_no",
            "pp_voucher_no",
            "part_no",
            "part_description",
            "sales_order_no",
            "customer_name",
            "ps_type",
            "so_scope",
            "current_stage_desc",
            "erp_last_stage_desc",
            "machine_cnc",
        )
    )


def _date_text(value: Any) -> str:
    text = compact_text(value)
    return text[:10] if text else ""


def _parse_bool(value: Any, *, field: str) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value in (0, 1):
            return bool(value)
        raise ValueError(f"{field} must be a boolean")
    text = compact_text(value).lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"", "0", "false", "no", "off"}:
        return False
    raise ValueError(f"{field} must be a boolean")


def _parse_program_finish(value: Any) -> str:
    text = compact_text(value)
    if not text:
        return ""
    iso_head = text.replace(" ", "T")
    if len(iso_head) >= 10 and iso_head[4] == "-" and iso_head[7] == "-":
        candidate = iso_head[:10]
        try:
            return date.fromisoformat(candidate).isoformat()
        except ValueError as exc:
            raise ValueError("program_finish_at must be a valid date") from exc
    token = text.split()[0].split("T")[0]
    for sep in ("/", "-", "."):
        parts = token.split(sep)
        if len(parts) != 3 or not all(part.isdigit() for part in parts):
            continue
        day_s, month_s, year_s = parts
        year = int(year_s)
        if year < 100:
            year += 2000
        try:
            return date(year, int(month_s), int(day_s)).isoformat()
        except ValueError as exc:
            raise ValueError("program_finish_at must be a valid date") from exc
    raise ValueError("program_finish_at must be YYYY-MM-DD or DD/MM/YYYY")


def _parse_material_subcon(raw: Any) -> tuple[bool, str, str]:
    text = compact_text(raw)
    if not text:
        return False, "", ""
    if text.upper() == "ARRIVED":
        return True, "", ""
    iso = _date_text(text) if len(text) >= 10 and text[4] == "-" else ""
    if iso and iso == text[:10] and (len(text) == 10 or text[10] in "T "):
        return False, iso, ""
    dmy = compact_text(raw)
    parts = dmy.replace("-", "/").split("/")
    if len(parts) == 3 and all(part.isdigit() for part in parts):
        day, month, year = (int(parts[0]), int(parts[1]), int(parts[2]))
        if year < 100:
            year += 2000
        if 1 <= day <= 31 and 1 <= month <= 12:
            iso = f"{year:04d}-{month:02d}-{day:02d}"
            return False, iso, ""
    return False, "", text


def _material_display(arrived: bool, material_date: str, legacy: str) -> str:
    if arrived:
        return "Arrived"
    if material_date:
        return material_date
    return compact_text(legacy)


def _qty_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return compact_text(value) or None
    if number != number:
        return None
    if number == int(number):
        return int(number)
    return number


def _machine_list(pp: dict[str, Any]) -> list[str]:
    machines: list[str] = []
    seen: set[str] = set()

    def _add(codes: Any) -> None:
        if not isinstance(codes, (list, tuple)):
            code = compact_text(codes)
            if code and code.upper() not in seen:
                seen.add(code.upper())
                machines.append(code)
            return
        for item in codes:
            code = compact_text(item)
            if code and code.upper() not in seen:
                seen.add(code.upper())
                machines.append(code)

    _add(pp.get("queued_machines"))
    by_partial = pp.get("queued_machines_by_partial") or {}
    if isinstance(by_partial, dict):
        for codes in by_partial.values():
            _add(codes)
    for partial in pp.get("partials") or []:
        _add(partial.get("queued_machines"))
    return machines


def _parse_machine_codes(raw: Any) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, str):
        text = compact_text(raw)
        if not text:
            return []
        if text.startswith("["):
            try:
                raw = json.loads(text)
            except json.JSONDecodeError as exc:
                raise ValueError("machine_codes must be a list of machine names") from exc
        else:
            raw = [part for part in text.replace(";", ",").split(",") if compact_text(part)]
    if not isinstance(raw, (list, tuple)):
        raw = [raw]
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        code = compact_text(item)
        key = code.upper()
        if not code or key in seen:
            continue
        seen.add(key)
        out.append(code)
    return out


def _machine_number(code: str) -> str:
    match = _MACHINE_NUM_RE.search(compact_text(code))
    return match.group(1) if match else ""


def _resolve_machine_codes(raw: Any, catalog: list[str] | None = None) -> list[str]:
    parsed = _parse_machine_codes(raw)
    catalog_list = [compact_text(item) for item in (catalog or []) if compact_text(item)]
    by_upper = {item.upper(): item for item in catalog_list}
    by_num: dict[str, str] = {}
    for item in catalog_list:
        number = _machine_number(item)
        if number and number not in by_num:
            by_num[number] = item
    out: list[str] = []
    seen: set[str] = set()
    for code in parsed:
        resolved = by_upper.get(code.upper())
        if not resolved:
            number = _machine_number(code)
            if number and by_num.get(number):
                resolved = by_num[number]
            elif number and compact_text(code) == number:
                resolved = f"CNC {number}"
            else:
                resolved = code
        key = resolved.upper()
        if key in seen:
            continue
        seen.add(key)
        out.append(resolved)
    return out


def _norm_import_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", compact_text(value).lower())


def _import_field_for_header(header: Any) -> str:
    needle = _norm_import_header(header)
    if not needle:
        return ""
    for field, aliases in _IMPORT_HEADER_ALIASES.items():
        if needle == field.replace("_", "") or needle in aliases:
            return field
    return ""


def _parse_pic_names(raw: Any) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        parts = [compact_text(item) for item in raw]
    else:
        text = compact_text(raw)
        if not text:
            return []
        parts = [compact_text(part) for part in _PIC_SPLIT_RE.split(text)]
    out: list[str] = []
    seen: set[str] = set()
    for name in parts:
        key = name.lower()
        if not name or key in _BLANK_CELL_VALUES or key in {"na", "n/a"} or key in seen:
            continue
        seen.add(key)
        out.append(name)
    return out


def _resolve_pic_names(con, names: list[str]) -> list[int]:
    out: list[int] = []
    seen: set[int] = set()
    for name in names:
        pic, _created = add_pic(con, name)
        pic_id = int(pic.get("pic_id") or 0)
        if pic_id <= 0 or pic_id in seen:
            continue
        seen.add(pic_id)
        out.append(pic_id)
    return out


def _parse_month_name_date(text: str) -> str:
    match = re.match(
        r"^(\d{1,2})[ \-./]+([A-Za-z]{3,9})(?:[ \-./]+(\d{2,4}))?$",
        compact_text(text),
    )
    if not match:
        return ""
    day = int(match.group(1))
    month = _MONTH_NAME_TO_NUM.get(match.group(2).lower())
    if not month or day < 1 or day > 31:
        return ""
    year_s = compact_text(match.group(3))
    year = date.today().year if not year_s else int(year_s)
    if year < 100:
        year += 2000
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return ""


def _parse_flexible_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = compact_text(value)
    if not text:
        return ""
    iso_head = text.replace(" ", "T")
    if len(iso_head) >= 10 and iso_head[4] == "-" and iso_head[7] == "-":
        candidate = iso_head[:10]
        try:
            return date.fromisoformat(candidate).isoformat()
        except ValueError:
            return ""
    named = _parse_month_name_date(text)
    if named:
        return named
    token = text.split()[0].split("T")[0]
    for sep in ("/", "-", "."):
        parts = token.split(sep)
        if len(parts) != 3 or not all(part.isdigit() for part in parts):
            continue
        first, second, year_s = parts
        year = int(year_s)
        if year < 100:
            year += 2000
        day_n, month_n = int(first), int(second)
        if month_n > 12 and day_n <= 12:
            day_n, month_n = month_n, day_n
        try:
            return date(year, month_n, day_n).isoformat()
        except ValueError:
            return ""
    return ""


def _cell_is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and compact_text(value).lower() in _BLANK_CELL_VALUES:
        return True
    if isinstance(value, float) and value != value:
        return True
    return False


def _parse_check_cell(raw: Any) -> dict[str, Any]:
    if raw is None or _cell_is_blank(raw):
        return {"tick": False, "mode": "tick", "text": ""}
    if isinstance(raw, bool):
        return {"tick": raw, "mode": "tick", "text": ""}
    if isinstance(raw, (datetime, date)):
        iso = _parse_flexible_date(raw)
        return {"tick": False, "mode": "text", "text": iso}
    text = compact_text(raw)
    if text.lower() in _CHECK_READY_VALUES:
        return {"tick": True, "mode": "tick", "text": ""}
    iso = _parse_flexible_date(raw)
    if iso:
        return {"tick": False, "mode": "text", "text": iso}
    return {"tick": False, "mode": "text" if text else "tick", "text": text}


def _serialize_import_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value != value:
        return None
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def _map_import_headers(header_row: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    seen: set[str] = set()
    for index, header in enumerate(header_row or []):
        field = _import_field_for_header(header)
        if not field or field in seen:
            continue
        seen.add(field)
        mapping[index] = field
    return mapping


def _sheet_matrix_from_xlsx(payload: bytes) -> list[tuple[str, list[list[Any]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheets: list[tuple[str, list[list[Any]]]] = []
    try:
        for ws in workbook.worksheets:
            matrix: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                matrix.append([_serialize_import_cell(cell) for cell in row])
            sheets.append((compact_text(ws.title) or "Sheet", matrix))
    finally:
        workbook.close()
    return sheets


def _sheet_matrix_from_xls(payload: bytes) -> list[tuple[str, list[list[Any]]]]:
    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("xlrd is required to read .xls files.") from exc
    book = xlrd.open_workbook(file_contents=payload)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for sheet in book.sheets():
        matrix: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            matrix.append(
                [_serialize_import_cell(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            )
        sheets.append((compact_text(sheet.name) or "Sheet", matrix))
    return sheets


def _items_from_sheet_matrix(matrix: list[list[Any]]) -> list[dict[str, Any]]:
    header_index = -1
    mapping: dict[int, str] = {}
    for index, row in enumerate(matrix[:15]):
        candidate = _map_import_headers(row)
        if "process_sheet_no" in candidate.values() and len(candidate) >= 1:
            header_index = index
            mapping = candidate
            break
    if header_index < 0:
        return []
    items: list[dict[str, Any]] = []
    for offset, row in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        parsed: dict[str, Any] = {}
        for col, field in mapping.items():
            if col >= len(row):
                continue
            parsed[field] = row[col]
        process_sheet_no = _ps_base(parsed.get("process_sheet_no"))
        if not process_sheet_no:
            continue
        patch: dict[str, Any] = {}
        for field in _IMPORT_PATCH_FIELDS:
            if field not in parsed or _cell_is_blank(parsed.get(field)):
                continue
            patch[field] = parsed.get(field)
        items.append({
            "process_sheet_no": process_sheet_no,
            "patch": patch,
            "source_row": offset,
        })
    return items


def parse_npi_import_workbook(payload: bytes, filename: str = "") -> list[dict[str, Any]]:
    if not payload:
        raise ValueError("The Excel file is empty")
    if len(payload) > _MAX_IMPORT_BYTES:
        raise ValueError("Excel file is larger than 12 MB")
    name = compact_text(filename).lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        sheets = _sheet_matrix_from_xls(payload)
    else:
        sheets = _sheet_matrix_from_xlsx(payload)
    best: list[dict[str, Any]] = []
    for _title, matrix in sheets:
        items = _items_from_sheet_matrix(matrix)
        if len(items) > len(best):
            best = items
        if items and any(item.get("patch") for item in items):
            return items
    if not best:
        raise ValueError(
            "Could not find a Data Input / process sheet column. "
            "Use the NPI Tracker template headers."
        )
    return best


def build_import_template_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    workbook = Workbook()
    ws = workbook.active
    ws.title = "NPI Tracker"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="334155")
    lock_fill = PatternFill("solid", fgColor="E2E8F0")
    for index, header in enumerate(IMPORT_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(index)].width = 18 if index < 11 else 28
    ws.auto_filter.ref = f"A1:{get_column_letter(len(IMPORT_TEMPLATE_COLUMNS))}1"
    ws.freeze_panes = "A2"
    ws["A1"].comment = Comment(
        "Process sheet number. Part details autofill from ERP on import.",
        "NPI Tracker",
    )
    for col in range(2, 6):
        ws.cell(row=2, column=col).fill = lock_fill
    notes = workbook.create_sheet("Notes")
    notes["A1"] = "How to import"
    notes["A1"].font = Font(bold=True, size=14)
    notes["A3"] = "1. Put the process sheet number in Data Input. Part No, Description, Qty and PO Due Date autofill from ERP."
    notes["A4"] = "2. PIC: type names separated by / or comma (e.g. Chang Peng/Anand). Unknown names are added to the PIC list."
    notes["A5"] = "3. Machine (CNC): comma-separated names or numbers (e.g. 22, 30 or CNC 10, CNC 20)."
    notes["A6"] = "4. Tooling / Fixture/Jig / Gauges/CMM: OK (ready), NA, a date, or a note such as Est. Wk 31."
    notes["A7"] = "5. Empty cells leave the current tracker value unchanged. Re-importing updates matching process sheets."
    notes.column_dimensions["A"].width = 120
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _stage_status_label(value: Any) -> str:
    code = compact_text(value).upper()
    if not code:
        return ""
    return _STAGE_STATUS_LABELS.get(code, compact_text(value))


def _empty_stage() -> dict[str, Any]:
    return {
        "current_stage_no": None,
        "current_stage_desc": "",
        "current_stage_status": "",
        "erp_stage_mode": "unassigned",
        "erp_last_stage_desc": "",
        "erp_wo_stage_count": 0,
    }


def _stage_from_item(item: dict[str, Any] | None) -> dict[str, Any]:
    row = item or {}
    mode = compact_text(row.get("erp_stage_mode")).lower() or "unassigned"
    desc = compact_text(row.get("current_stage_desc"))
    status = compact_text(row.get("current_stage_status"))
    if desc or status:
        mode = "open"
    try:
        stage_no = int(row.get("current_stage_no")) if row.get("current_stage_no") is not None else None
    except (TypeError, ValueError):
        stage_no = None
    try:
        wo_count = int(row.get("erp_wo_stage_count") or 0)
    except (TypeError, ValueError):
        wo_count = 0
    return {
        "current_stage_no": stage_no,
        "current_stage_desc": desc,
        "current_stage_status": status,
        "erp_stage_mode": mode,
        "erp_last_stage_desc": compact_text(row.get("erp_last_stage_desc")),
        "erp_wo_stage_count": wo_count,
    }


def _stage_from_pp(pp: dict[str, Any] | None) -> dict[str, Any]:
    row = pp or {}
    items = list(row.get("partials") or [])
    items.append(row)
    best = _empty_stage()
    for item in items:
        stage = _stage_from_item(item)
        if stage["current_stage_desc"] or stage["current_stage_status"] or stage["erp_stage_mode"] == "open":
            return stage
        if best["erp_stage_mode"] == "unassigned":
            best = stage
        if stage["erp_stage_mode"] == "completed":
            best = stage
    if row.get("shipped_completed") and best["erp_stage_mode"] == "unassigned":
        best["erp_stage_mode"] = "completed"
    return best


def _stage_from_overlay(overlay: dict[tuple[str, int], dict[str, Any]], process_sheet_no: str) -> dict[str, Any] | None:
    base = _ps_base(process_sheet_no).upper()
    if not base or not overlay:
        return None
    matches = [
        dict(value)
        for (ps_base, _partial), value in overlay.items()
        if compact_text(ps_base).upper() == base
    ]
    if not matches:
        return None
    return _stage_from_pp({"partials": matches})


def _is_history_status(value: Any) -> bool:
    return compact_text(value).lower() in _HISTORY_STATUSES


def _looks_like_ps(query: str) -> bool:
    text = compact_text(query).upper()
    if len(text) < 4:
        return False
    if any(text.startswith(prefix) for prefix in ("APS", "NPS", "MPS", "PPS", "CPS", "SR", "AP", "NP", "MP", "PP", "CP")):
        return True
    return "-" in text and any(ch.isdigit() for ch in text)


def _coway_edd(pp: dict[str, Any]) -> str:
    edd = _date_text(pp.get("coway_proposed_edd"))
    if edd:
        return edd
    for partial in pp.get("partials") or []:
        edd = _date_text(partial.get("coway_proposed_edd"))
        if edd:
            return edd
    return ""


def job_from_sales_order_pp(
    order: dict[str, Any],
    pp: dict[str, Any],
    *,
    so_scope: str = "active",
) -> dict[str, Any] | None:
    process_sheet_no = _ps_base(pp.get("process_sheet_no") or pp.get("pp_voucher_no"))
    pp_voucher_no = compact_text(pp.get("pp_voucher_no"))
    if not process_sheet_no and not pp_voucher_no:
        return None
    machines = _machine_list(pp)
    posted = _date_text(
        order.get("first_posted_datetime")
        or pp.get("order_date")
        or order.get("order_date")
        or pp.get("posted_date")
    )
    material_subcon = compact_text(pp.get("material_subcon"))
    arrived, material_date, material_legacy = _parse_material_subcon(material_subcon)
    bom_code = compact_text(pp.get("bom_code"))
    scope = compact_text(so_scope).lower() or "active"
    shipped = bool(pp.get("shipped_completed")) or scope == "complete"
    stage = _stage_from_pp(pp)
    if shipped and stage["erp_stage_mode"] == "unassigned":
        stage["erp_stage_mode"] = "completed"
    job = {
        "process_sheet_no": process_sheet_no or pp_voucher_no,
        "pp_voucher_no": pp_voucher_no,
        "part_no": compact_text(pp.get("inventory_code") or pp.get("part_no")),
        "part_description": compact_text(pp.get("description") or pp.get("part_description")),
        "total_qty": _qty_value(pp.get("pp_qty") if pp.get("pp_qty") is not None else pp.get("total_qty")),
        "po_due_date": _date_text(pp.get("due_date") or pp.get("po_due_date")),
        "posted_date": posted,
        "queued_machines": machines,
        "machine_cnc": ", ".join(machines),
        "coway_proposed_edd": _coway_edd(pp),
        "sales_order_no": compact_text(order.get("sales_order_no") or pp.get("source_voucher_no")),
        "customer_name": compact_text(order.get("customer_name")),
        "is_new_part": bool(pp.get("is_new_part")),
        "bom_code": bom_code,
        "has_bom": False,
        "material_subcon": material_subcon,
        "material_arrived": arrived,
        "material_date": material_date,
        "material_legacy": material_legacy,
        "material_display": _material_display(arrived, material_date, material_legacy),
        "so_scope": "complete" if shipped else "active",
        "shipped_completed": shipped,
        "from_erp_cache": False,
        **stage,
    }
    job["ps_type"] = _job_ps_type(job) or "OTHER"
    return job


def _lookup_parts_with_bom_materials(part_nos: list[Any] | tuple[Any, ...] | None) -> set[str]:
    """Has BOM = the part has leaf material lines in ERP inventory_bom_listing."""
    from .bom_materials import parts_with_leaf_bom_materials, parts_with_leaf_bom_materials_planner
    from .staged_erp import live_query

    codes = sorted({compact_text(part) for part in (part_nos or []) if compact_text(part)})
    if not codes:
        return set()

    def db_query(sql, params=(), fetchall=False):
        if not fetchall:
            return None
        return live_query(sql, params, timeout_ms=15000)

    try:
        return parts_with_leaf_bom_materials(db_query, codes)
    except Exception:
        logger.exception("ERP leaf BOM lookup failed; using planner material_per_bom")
    try:
        with planner_db() as con:
            return parts_with_leaf_bom_materials_planner(con, codes)
    except Exception:
        logger.exception("planner leaf BOM lookup failed")
        return set()


def _apply_has_bom(jobs: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    job_rows = list(jobs or [])
    found = _lookup_parts_with_bom_materials([job.get("part_no") for job in job_rows])
    for job in job_rows:
        job["has_bom"] = compact_text(job.get("part_no")) in found
    return job_rows


def flatten_sales_order_jobs(
    orders: list[dict[str, Any]] | None,
    *,
    so_scope: str = "active",
) -> list[dict[str, Any]]:
    jobs: list[dict[str, Any]] = []
    seen: set[str] = set()
    for order in orders or []:
        for pp in order.get("pp_vouchers") or []:
            job = job_from_sales_order_pp(order, pp, so_scope=so_scope)
            if not job:
                continue
            key = _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
            if not key or key in seen:
                continue
            seen.add(key)
            jobs.append(job)
    return jobs


def index_jobs_by_ps(jobs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for job in jobs:
        for raw in (job.get("process_sheet_no"), job.get("pp_voucher_no")):
            key = _ps_key(raw)
            if key and key not in indexed:
                indexed[key] = job
    return indexed


def search_jobs(
    jobs: list[dict[str, Any]],
    query: str,
    *,
    flagged_keys: set[str] | None = None,
    limit: int = _SEARCH_LIMIT,
) -> list[dict[str, Any]]:
    needle = compact_text(query).upper()
    if not needle:
        return []
    flagged = flagged_keys or set()
    cap = max(1, min(int(limit or _SEARCH_LIMIT), 50))
    hits: list[dict[str, Any]] = []
    for job in jobs:
        if needle not in _job_search_blob(job):
            continue
        hits.append(_decorate_candidate(job, flagged))
        if len(hits) >= cap:
            break
    return hits


def _decorate_candidate(job: dict[str, Any], flagged_keys: set[str]) -> dict[str, Any]:
    out = dict(job)
    key = _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
    out["already_flagged"] = key in flagged_keys
    out["ps_type"] = compact_text(out.get("ps_type")) or _job_ps_type(out)
    return out


def list_flag_candidates(
    *,
    query: str = "",
    ps_type_filter: str = "",
    scope_filter: str = "",
    limit: int = _CANDIDATE_LIMIT,
) -> dict[str, Any]:
    jobs = _sales_order_jobs(allow_rebuild=True)
    with planner_db() as con:
        _ensure_tables(con)
        flagged = _flagged_keys(con)
    decorated = [_decorate_candidate(job, flagged) for job in jobs]
    wanted_scope = compact_text(scope_filter).lower()
    if wanted_scope in {"active", "complete"}:
        decorated = [
            job
            for job in decorated
            if compact_text(job.get("so_scope")).lower() == wanted_scope
        ]
    type_counts: dict[str, int] = {}
    for job in decorated:
        label = compact_text(job.get("ps_type")) or "OTHER"
        type_counts[label] = type_counts.get(label, 0) + 1

    needle = compact_text(query).upper()
    wanted = compact_text(ps_type_filter).upper()
    cap = max(1, min(int(limit or _CANDIDATE_LIMIT), 2500))
    hits: list[dict[str, Any]] = []
    truncated = False
    for job in decorated:
        kind = compact_text(job.get("ps_type")) or "OTHER"
        if wanted and wanted != kind:
            continue
        if needle and needle not in _job_search_blob(job):
            continue
        hits.append(job)
        if len(hits) >= cap:
            truncated = True
            break

    types = sorted(
        type_counts,
        key=lambda label: (
            _PS_TYPE_ORDER.index(label) if label in _PS_TYPE_ORDER else 99,
            label,
        ),
    )
    return {
        "rows": hits,
        "types": [{"ps_type": label, "count": type_counts[label]} for label in types],
        "total": len(decorated),
        "matched": len(hits),
        "truncated": truncated,
        "scope": wanted_scope or "all",
    }


def _parse_pic_ids(raw: Any) -> list[int]:
    if raw is None:
        return []
    if isinstance(raw, str):
        text = compact_text(raw)
        if not text:
            return []
        if text.startswith("["):
            try:
                raw = json.loads(text)
            except json.JSONDecodeError as exc:
                raise ValueError("pic_ids must be a list of ids") from exc
        else:
            raw = [part for part in text.split(",") if compact_text(part)]
    if not isinstance(raw, (list, tuple)):
        raw = [raw]
    out: list[int] = []
    seen: set[int] = set()
    for item in raw:
        try:
            pic_id = int(item)
        except (TypeError, ValueError) as exc:
            raise ValueError("pic_ids must be integers") from exc
        if pic_id <= 0 or pic_id in seen:
            continue
        seen.add(pic_id)
        out.append(pic_id)
    return out


def _parse_mode(value: Any, *, field: str) -> str:
    mode = compact_text(value).lower()
    if mode not in CHECK_TEXT_MODES:
        raise ValueError(f"{field} must be tick or text")
    return mode


def _serialize_pic(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    out = serialize_row(dict(row))
    out["pic_id"] = int(out.get("pic_id") or 0)
    out["name"] = compact_text(out.get("name"))
    out["active"] = bool(out.get("active", True))
    return out


def _pics_for_ids(pics_by_id: dict[int, dict[str, Any]], pic_ids: list[int]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for pic_id in pic_ids:
        pic = pics_by_id.get(pic_id)
        if pic:
            out.append({"pic_id": pic["pic_id"], "name": pic["name"]})
    return out


def _pic_display(pics_by_id: dict[int, dict[str, Any]] | None, pic_ids: Any) -> str:
    names = [
        compact_text(pic.get("name"))
        for pic in _pics_for_ids(pics_by_id or {}, _parse_pic_ids(pic_ids))
        if compact_text(pic.get("name"))
    ]
    return ", ".join(names)


def history_field_label(field_name: str) -> str:
    key = compact_text(field_name)
    return _HISTORY_FIELD_LABELS.get(key, key.replace("_", " ").strip() or key)


def history_text(field_name: str, value: Any, pics_by_id: dict[int, dict[str, Any]] | None = None) -> str:
    field = compact_text(field_name)
    if field in {"program_pic_ids", "pic_ids"}:
        return _pic_display(pics_by_id, value)
    if field == "program_finish_at":
        try:
            return _parse_program_finish(value)
        except ValueError:
            return compact_text(value)
    return compact_text(value)


def diff_tracked_fields(
    before: dict[str, Any] | None,
    after: dict[str, Any] | None,
    fields: tuple[str, ...] | list[str],
    *,
    pics_by_id: dict[int, dict[str, Any]] | None = None,
) -> list[dict[str, str]]:
    previous = before or {}
    current = after or {}
    out: list[dict[str, str]] = []
    for field in fields:
        old_value = history_text(field, previous.get(field), pics_by_id)
        new_value = history_text(field, current.get(field), pics_by_id)
        if old_value == new_value:
            continue
        out.append({
            "field_name": compact_text(field),
            "field_label": history_field_label(field),
            "old_value": old_value,
            "new_value": new_value,
        })
    return out


def _insert_change_logs(
    con,
    *,
    source: str,
    process_sheet_no: str,
    first_article_id: int | None,
    changes: list[dict[str, str]],
) -> None:
    kind = compact_text(source).lower()
    ps = _ps_base(process_sheet_no)
    if kind not in _HISTORY_SOURCES or not ps or not changes:
        return
    for change in changes:
        field_name = compact_text(change.get("field_name"))
        if not field_name:
            continue
        con.execute(
            """
            INSERT INTO planner_first_article_change_log (
                source, process_sheet_no, first_article_id, field_name, old_value, new_value
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                kind,
                ps,
                int(first_article_id) if first_article_id else None,
                field_name,
                compact_text(change.get("old_value")),
                compact_text(change.get("new_value")),
            ),
        )


def _history_count_map(con, source: str, keys: list[str] | None) -> dict[str, int]:
    wanted = sorted({_ps_key(item) for item in (keys or []) if _ps_key(item)})
    kind = compact_text(source).lower()
    if not wanted or kind not in _HISTORY_SOURCES:
        return {}
    fetched = rows(
        con.execute(
            """
            SELECT UPPER(TRIM(process_sheet_no)) AS ps_key, COUNT(*) AS n
            FROM planner_first_article_change_log
            WHERE source = %s
              AND UPPER(TRIM(process_sheet_no)) = ANY(%s)
            GROUP BY UPPER(TRIM(process_sheet_no))
            """,
            (kind, wanted),
        )
    )
    return {_ps_key(row.get("ps_key")): int(row.get("n") or 0) for row in fetched or []}


def _serialize_change_log(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    out = serialize_row(dict(row))
    field_name = compact_text(out.get("field_name"))
    return {
        "change_id": int(out.get("change_id") or 0),
        "source": compact_text(out.get("source")),
        "process_sheet_no": _ps_base(out.get("process_sheet_no")),
        "first_article_id": int(out.get("first_article_id") or 0) or None,
        "field_name": field_name,
        "field_label": history_field_label(field_name),
        "old_value": compact_text(out.get("old_value")),
        "new_value": compact_text(out.get("new_value")),
        "changed_at": compact_text(out.get("changed_at")),
    }


def list_change_history(*, source: str, process_sheet_no: str, limit: int = _HISTORY_LIMIT) -> list[dict[str, Any]]:
    kind = compact_text(source).lower()
    ps = _ps_base(process_sheet_no)
    if kind not in _HISTORY_SOURCES:
        raise ValueError("source must be new_part or flagged")
    if not ps:
        raise ValueError("process_sheet_no is required")
    cap = max(1, min(int(limit or _HISTORY_LIMIT), 500))
    with planner_db() as con:
        _ensure_tables(con)
        fetched = rows(
            con.execute(
                """
                SELECT change_id, source, process_sheet_no, first_article_id,
                       field_name, old_value, new_value, changed_at
                FROM planner_first_article_change_log
                WHERE source = %s
                  AND LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
                ORDER BY changed_at DESC, change_id DESC
                LIMIT %s
                """,
                (kind, ps, cap),
            )
        )
    return [item for item in (_serialize_change_log(row) for row in fetched) if item]


def _decorate_stage_labels(row: dict[str, Any] | None) -> dict[str, Any]:
    out = row or {}
    out["current_stage_status_label"] = _stage_status_label(out.get("current_stage_status"))
    return out


def _serialize_tracker_row(
    row: dict[str, Any] | None,
    *,
    live: dict[str, Any] | None = None,
    pics_by_id: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    if not row:
        return None
    out = serialize_row(dict(row))
    out["first_article_id"] = int(out.get("first_article_id") or 0)
    out["process_sheet_no"] = _ps_base(out.get("process_sheet_no"))
    out["pp_voucher_no"] = compact_text(out.get("pp_voucher_no"))
    pic_ids = _parse_pic_ids(out.get("pic_ids"))
    out["pic_ids"] = pic_ids
    out["pics"] = _pics_for_ids(pics_by_id or {}, pic_ids)
    for prefix in CHECK_TEXT_FIELDS:
        out[f"{prefix}_mode"] = compact_text(out.get(f"{prefix}_mode")).lower() or "tick"
        out[f"{prefix}_tick"] = bool(out.get(f"{prefix}_tick"))
        out[f"{prefix}_text"] = compact_text(out.get(f"{prefix}_text"))
    out["remarks"] = compact_text(out.get("remarks"))

    live = live or {}
    saved_machines = out.get("machine_codes")
    live_machines = list(live.get("queued_machines") or [])
    if saved_machines is None:
        machines = _parse_machine_codes(live_machines)
    else:
        machines = _parse_machine_codes(saved_machines)
    out["machine_codes"] = machines
    out["machine_cnc"] = ", ".join(machines)
    out["queued_machines"] = live_machines or machines

    out["part_no"] = compact_text(live.get("part_no"))
    out["part_description"] = compact_text(live.get("part_description"))
    out["total_qty"] = live.get("total_qty")
    out["po_due_date"] = _date_text(live.get("po_due_date"))
    out["coway_proposed_edd"] = _date_text(live.get("coway_proposed_edd"))
    out["sales_order_no"] = compact_text(live.get("sales_order_no"))
    out["customer_name"] = compact_text(live.get("customer_name"))
    out["so_scope"] = compact_text(live.get("so_scope"))
    out["shipped_completed"] = bool(live.get("shipped_completed"))
    out["from_erp_cache"] = bool(live.get("from_erp_cache"))
    out["in_sales_orders"] = bool(live) and not out["from_erp_cache"]
    stage = _stage_from_item(live) if live else _empty_stage()
    out["current_stage_no"] = stage["current_stage_no"]
    out["current_stage_desc"] = stage["current_stage_desc"]
    out["current_stage_status"] = stage["current_stage_status"]
    out["erp_stage_mode"] = stage["erp_stage_mode"]
    out["erp_last_stage_desc"] = stage["erp_last_stage_desc"]
    out["erp_wo_stage_count"] = stage["erp_wo_stage_count"]
    out["current_stage_status_label"] = _stage_status_label(stage["current_stage_status"])
    out["history_count"] = int(out.get("history_count") or 0)
    return out


def load_pics(con) -> list[dict[str, Any]]:
    _ensure_tables(con)
    fetched = rows(
        con.execute(
            """
            SELECT pic_id, name, active, created_at
            FROM planner_first_article_pic
            WHERE active = TRUE
            ORDER BY LOWER(TRIM(name)), pic_id
            """
        )
    )
    return [pic for pic in (_serialize_pic(item) for item in fetched) if pic]


def add_pic(con, name: str) -> tuple[dict[str, Any], bool]:
    _ensure_tables(con)
    clean = compact_text(name)
    if not clean:
        raise ValueError("PIC name is required")

    existing = one(
        con.execute(
            """
            SELECT pic_id, name, active, created_at
            FROM planner_first_article_pic
            WHERE active = TRUE
              AND LOWER(TRIM(name)) = LOWER(TRIM(%s))
            ORDER BY pic_id
            LIMIT 1
            """,
            (clean,),
        )
    )
    if existing:
        serialized = _serialize_pic(dict(existing))
        return serialized or {}, False

    inactive = one(
        con.execute(
            """
            SELECT pic_id, name, active, created_at
            FROM planner_first_article_pic
            WHERE active = FALSE
              AND LOWER(TRIM(name)) = LOWER(TRIM(%s))
            ORDER BY pic_id
            LIMIT 1
            """,
            (clean,),
        )
    )
    if inactive:
        row = one(
            con.execute(
                """
                UPDATE planner_first_article_pic
                SET active = TRUE, name = %s
                WHERE pic_id = %s
                RETURNING pic_id, name, active, created_at
                """,
                (clean, int(inactive["pic_id"])),
            )
        )
        return _serialize_pic(dict(row) if row else {}) or {}, True

    row = one(
        con.execute(
            """
            INSERT INTO planner_first_article_pic (name)
            VALUES (%s)
            RETURNING pic_id, name, active, created_at
            """,
            (clean,),
        )
    )
    return _serialize_pic(dict(row) if row else {}) or {}, True


def delete_pic(con, pic_id: int) -> dict[str, Any] | None:
    _ensure_tables(con)
    row = one(
        con.execute(
            """
            SELECT pic_id, name
            FROM planner_first_article_pic
            WHERE pic_id = %s AND active = TRUE
            """,
            (int(pic_id),),
        )
    )
    if not row:
        return None
    name = compact_text(row.get("name"))
    ids = [
        int(item["pic_id"])
        for item in rows(
            con.execute(
                """
                SELECT pic_id
                FROM planner_first_article_pic
                WHERE active = TRUE
                  AND LOWER(TRIM(name)) = LOWER(TRIM(%s))
                """,
                (name,),
            )
        )
    ]
    if not ids:
        return None
    con.execute(
        """
        UPDATE planner_first_article
        SET pic_ids = COALESCE(
                ARRAY(
                    SELECT x FROM UNNEST(pic_ids) AS x
                    WHERE x <> ALL(%s)
                ),
                '{}'::bigint[]
            ),
            updated_at = NOW()
        WHERE pic_ids && %s
        """,
        (ids, ids),
    )
    con.execute(
        """
        UPDATE planner_first_article_new_part
        SET program_pic_ids = COALESCE(
                ARRAY(
                    SELECT x FROM UNNEST(program_pic_ids) AS x
                    WHERE x <> ALL(%s)
                ),
                '{}'::bigint[]
            ),
            updated_at = NOW()
        WHERE program_pic_ids && %s
        """,
        (ids, ids),
    )
    cur = con.execute(
        """
        UPDATE planner_first_article_pic
        SET active = FALSE
        WHERE pic_id = ANY(%s) AND active = TRUE
        """,
        (ids,),
    )
    return {"name": name, "removed_count": int(getattr(cur, "rowcount", 0) or 0)}


def _pics_by_id(con) -> dict[int, dict[str, Any]]:
    return {int(pic["pic_id"]): pic for pic in load_pics(con)}


def _validate_pic_ids(con, pic_ids: list[int]) -> list[int]:
    if not pic_ids:
        return []
    known = {int(pic["pic_id"]) for pic in load_pics(con)}
    unknown = [pic_id for pic_id in pic_ids if pic_id not in known]
    if unknown:
        raise ValueError("Unknown PIC id")
    return pic_ids


def _peek_complete_orders() -> list[dict[str, Any]]:
    try:
        from .erp_route_cache import get as cache_get
        from .sales_orders_route import _sales_orders_cache_key
    except Exception:
        logger.exception("first article complete-order cache import failed")
        return []
    try:
        cached = cache_get(_sales_orders_cache_key("complete"), ttl_sec=0)
    except Exception:
        logger.exception("first article complete-order cache read failed")
        return []
    if isinstance(cached, dict):
        return list(cached.get("complete") or [])
    return []


def _peek_cached_sales_orders() -> dict[str, Any]:
    """Read S/O cache only. Never trigger a live ERP rebuild."""
    try:
        from .erp_route_cache import get as cache_get
        from .sales_orders_route import _sales_orders_cache_key
    except Exception:
        logger.exception("first article sales-order cache import failed")
        return {}

    payload: dict[str, Any] = {}
    for lite in (False, True):
        try:
            cached = cache_get(_sales_orders_cache_key("active", lite=lite), ttl_sec=0)
        except Exception:
            logger.exception("first article sales-order cache read failed")
            continue
        if isinstance(cached, dict) and (cached.get("active") or not payload):
            payload = dict(cached)
            if payload.get("active"):
                break
    complete = _peek_complete_orders()
    if complete:
        if not payload:
            payload = {"active": [], "complete": complete}
        else:
            payload["complete"] = complete
    return payload if isinstance(payload, dict) else {}


def _sales_order_payload(*, allow_rebuild: bool = False) -> dict[str, Any]:
    payload = _peek_cached_sales_orders()
    if not payload.get("active") and allow_rebuild:
        try:
            from .sales_orders_route import _fetch_sales_orders

            # Staged/lite rebuild only. Full live S/O build can stall the tracker
            # for minutes while COMAIN is busy or another worker holds the cache lock.
            rebuilt = _fetch_sales_orders(refresh=False, active_only=True, lite=True) or {}
            if isinstance(rebuilt, dict):
                payload = dict(rebuilt)
                if not payload.get("complete"):
                    payload["complete"] = _peek_complete_orders()
        except Exception:
            logger.exception("first article sales-order lookup failed")
    return payload if isinstance(payload, dict) else {}


def _ensure_new_part_flags(orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
    needs = False
    for order in orders or []:
        for pp in order.get("pp_vouchers") or []:
            if "is_new_part" not in pp:
                needs = True
                break
        if needs:
            break
    if not needs:
        return orders
    try:
        from .sales_orders_route import _apply_new_part_overlay

        _apply_new_part_overlay(orders)
    except Exception:
        logger.exception("first article new-part overlay failed")
    return orders


def _sales_order_jobs(*, allow_rebuild: bool = False) -> list[dict[str, Any]]:
    payload = _sales_order_payload(allow_rebuild=allow_rebuild)
    jobs = flatten_sales_order_jobs(list(payload.get("active") or []), so_scope="active")
    seen = {
        _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
        for job in jobs
    }
    for job in flatten_sales_order_jobs(list(payload.get("complete") or []), so_scope="complete"):
        key = _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
        if not key or key in seen:
            continue
        seen.add(key)
        jobs.append(job)
    return jobs


def _job_from_pp_cache_row(row: dict[str, Any]) -> dict[str, Any] | None:
    process_sheet_no = _ps_base(row.get("process_sheet_no") or row.get("ps_id"))
    if not process_sheet_no:
        return None
    historical = _is_history_status(row.get("status") or row.get("erp_status"))
    stage = _stage_from_item(row)
    if historical and stage["erp_stage_mode"] == "unassigned" and not stage["current_stage_desc"]:
        stage["erp_stage_mode"] = "completed"
    job = {
        "process_sheet_no": process_sheet_no,
        "pp_voucher_no": compact_text(row.get("pp_voucher_no")),
        "part_no": compact_text(row.get("part_no") or row.get("inventory_code")),
        "part_description": compact_text(row.get("part_description") or row.get("description") or row.get("main_desc")),
        "total_qty": _qty_value(row.get("total_qty") if row.get("total_qty") is not None else row.get("pp_qty")),
        "po_due_date": _date_text(row.get("po_due_date") or row.get("due_date")),
        "posted_date": _date_text(row.get("order_date") or row.get("posted_date")),
        "queued_machines": [],
        "machine_cnc": "",
        "coway_proposed_edd": "",
        "sales_order_no": compact_text(row.get("sales_order_no") or row.get("source_voucher_no")),
        "customer_name": compact_text(row.get("customer_name")),
        "is_new_part": False,
        "bom_code": compact_text(row.get("bom_code")),
        "has_bom": False,
        "material_subcon": "",
        "material_arrived": False,
        "material_date": "",
        "material_legacy": "",
        "material_display": "",
        "so_scope": "complete" if historical else "active",
        "shipped_completed": historical,
        "from_erp_cache": True,
        **stage,
    }
    job["ps_type"] = _job_ps_type(job) or "OTHER"
    return job


def _merge_live_job(primary: dict[str, Any] | None, extra: dict[str, Any] | None) -> dict[str, Any] | None:
    if not extra:
        return primary
    if not primary:
        return extra
    out = dict(primary)
    for field in (
        "part_no",
        "part_description",
        "po_due_date",
        "posted_date",
        "sales_order_no",
        "customer_name",
        "coway_proposed_edd",
        "current_stage_desc",
        "current_stage_status",
        "erp_last_stage_desc",
        "pp_voucher_no",
        "bom_code",
    ):
        if not compact_text(out.get(field)) and compact_text(extra.get(field)):
            out[field] = extra.get(field)
    if out.get("total_qty") in (None, "") and extra.get("total_qty") not in (None, ""):
        out["total_qty"] = extra.get("total_qty")
    if not out.get("queued_machines") and extra.get("queued_machines"):
        out["queued_machines"] = list(extra.get("queued_machines") or [])
        out["machine_cnc"] = compact_text(extra.get("machine_cnc")) or ", ".join(out["queued_machines"])
    if compact_text(out.get("erp_stage_mode")) in {"", "unassigned"} and compact_text(extra.get("erp_stage_mode")):
        out["erp_stage_mode"] = extra.get("erp_stage_mode")
        out["current_stage_no"] = extra.get("current_stage_no")
        out["erp_wo_stage_count"] = extra.get("erp_wo_stage_count") or out.get("erp_wo_stage_count")
    return out


def _index_jobs(jobs) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for job in jobs or []:
        if not job:
            continue
        key = _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
        if key:
            out[key] = job
    return out


def _lookup_jobs_from_pp_cache(keys: list[str] | None) -> dict[str, dict[str, Any]]:
    wanted = sorted({_ps_key(item) for item in (keys or []) if _ps_key(item)})
    if not wanted:
        return {}
    fetched: list[dict[str, Any]] = []
    try:
        with planner_db() as con:
            fetched = rows(
                con.execute(
                    """
                    SELECT
                        UPPER(TRIM(split_part(ps_id, '::', 1))) AS ps_key,
                        MAX(split_part(ps_id, '::', 1)) AS process_sheet_no,
                        MAX(part_no) AS part_no,
                        MAX(description) AS part_description,
                        MAX(total_qty) AS total_qty,
                        MIN(due_date) AS po_due_date,
                        MAX(order_date) AS order_date,
                        MAX(source_voucher_no) AS sales_order_no,
                        MAX(status) AS status
                    FROM pp_vouchers_cache
                    WHERE UPPER(TRIM(ps_id)) = ANY(%s)
                       OR UPPER(TRIM(split_part(ps_id, '::', 1))) = ANY(%s)
                    GROUP BY UPPER(TRIM(split_part(ps_id, '::', 1)))
                    """,
                    (wanted, wanted),
                )
            )
    except Exception:
        logger.exception("first article pp_vouchers_cache lookup failed")
        fetched = []
    out = _index_jobs(_job_from_pp_cache_row(dict(row)) for row in fetched or [])
    missing = [key for key in wanted if key not in out]
    if missing:
        out.update(_lookup_jobs_from_process_sheet_info(missing))
    _fill_part_descriptions(out)
    return out


def _lookup_jobs_from_process_sheet_info(keys: list[str]) -> dict[str, dict[str, Any]]:
    wanted = sorted({_ps_key(item) for item in keys if _ps_key(item)})
    if not wanted:
        return {}
    try:
        with planner_db() as con:
            fetched = rows(
                con.execute(
                    """
                    SELECT
                        UPPER(TRIM(process_sheet_no)) AS ps_key,
                        MAX(process_sheet_no) AS process_sheet_no,
                        MAX(pp_voucher_no) AS pp_voucher_no,
                        MAX(inventory_code) AS part_no,
                        MAX(COALESCE(total_qty, 0)) AS total_qty
                    FROM mfg_process_sheet_info
                    WHERE UPPER(TRIM(process_sheet_no)) = ANY(%s)
                    GROUP BY UPPER(TRIM(process_sheet_no))
                    """,
                    (wanted,),
                )
            )
    except Exception:
        logger.exception("first article mfg_process_sheet_info lookup failed")
        return {}
    return _index_jobs(_job_from_pp_cache_row(dict(row)) for row in fetched or [])


def _fill_part_descriptions(jobs_by_ps: dict[str, dict[str, Any]]) -> None:
    codes = sorted({
        compact_text(job.get("part_no"))
        for job in jobs_by_ps.values()
        if compact_text(job.get("part_no")) and not compact_text(job.get("part_description"))
    })
    if not codes:
        return
    try:
        with planner_db() as con:
            fetched = rows(
                con.execute(
                    """
                    SELECT inventory_code, main_desc
                    FROM part_desc
                    WHERE inventory_code = ANY(%s)
                    """,
                    (codes,),
                )
            )
    except Exception:
        logger.exception("first article part description lookup failed")
        return
    by_code = {
        compact_text(row.get("inventory_code")): compact_text(row.get("main_desc"))
        for row in fetched or []
        if compact_text(row.get("inventory_code"))
    }
    for job in jobs_by_ps.values():
        desc = by_code.get(compact_text(job.get("part_no")))
        if desc:
            job["part_description"] = desc


def _search_jobs_from_pp_cache(query: str, *, flagged_keys: set[str], limit: int) -> list[dict[str, Any]]:
    needle = compact_text(query).upper()
    if len(needle) < 2:
        return []
    pattern = f"%{needle}%"
    cap = max(1, min(int(limit or _SEARCH_LIMIT), 50))
    try:
        with planner_db() as con:
            fetched = rows(
                con.execute(
                    """
                    SELECT
                        split_part(ps_id, '::', 1) AS process_sheet_no,
                        part_no,
                        description AS part_description,
                        total_qty,
                        due_date AS po_due_date,
                        order_date,
                        source_voucher_no AS sales_order_no,
                        status
                    FROM pp_vouchers_cache
                    WHERE UPPER(TRIM(split_part(ps_id, '::', 1))) LIKE %s
                       OR UPPER(TRIM(COALESCE(part_no, ''))) LIKE %s
                    ORDER BY ps_id, pp_partial_no NULLS FIRST
                    LIMIT 80
                    """,
                    (pattern, pattern),
                )
            )
    except Exception:
        logger.exception("first article pp_vouchers_cache search failed")
        return []
    hits: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in fetched or []:
        job = _job_from_pp_cache_row(dict(row))
        if not job:
            continue
        key = _ps_key(job.get("process_sheet_no"))
        if not key or key in seen:
            continue
        seen.add(key)
        hits.append(_decorate_candidate(job, flagged_keys))
        if len(hits) >= cap:
            break
    return hits


def load_machine_catalog() -> list[str]:
    try:
        from .machines import fetch_machines

        with planner_db() as con:
            fetched = fetch_machines(con)
    except Exception:
        logger.exception("first article machine catalog failed")
        return []
    out: list[str] = []
    seen: set[str] = set()
    for row in fetched or []:
        code = compact_text((row or {}).get("machine_no") or (row or {}).get("machine_code"))
        key = code.upper()
        if not code or key in seen:
            continue
        seen.add(key)
        out.append(code)
    return out


def _live_job_map(*, allow_rebuild: bool = False) -> dict[str, dict[str, Any]]:
    return index_jobs_by_ps(_sales_order_jobs(allow_rebuild=allow_rebuild))


def _flagged_keys(con) -> set[str]:
    fetched = rows(
        con.execute(
            """
            SELECT process_sheet_no, pp_voucher_no
            FROM planner_first_article
            """
        )
    )
    keys: set[str] = set()
    for row in fetched:
        for raw in (row.get("process_sheet_no"), row.get("pp_voucher_no")):
            key = _ps_key(raw)
            if key:
                keys.add(key)
    return keys


def list_tracker_rows(*, live_by_ps: dict[str, dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    if live_by_ps is not None:
        live_map = dict(live_by_ps)
    else:
        try:
            live_map = _live_job_map()
        except Exception:
            logger.exception("first article live map failed")
            live_map = {}
    with planner_db() as con:
        _ensure_tables(con)
        pics = _pics_by_id(con)
        fetched = rows(
            con.execute(
                f"""
                SELECT {_ROW_SELECT}
                FROM planner_first_article
                ORDER BY updated_at DESC, first_article_id DESC
                """
            )
        )
        all_keys = [
            key
            for key in (
                _ps_key(row.get("process_sheet_no") or row.get("pp_voucher_no"))
                for row in fetched
            )
            if key
        ]
        history_counts = _history_count_map(con, "flagged", all_keys)
    cache_map = _lookup_jobs_from_pp_cache(all_keys)
    for key, job in cache_map.items():
        live_map[key] = _merge_live_job(live_map.get(key), job) or job
    out: list[dict[str, Any]] = []
    for row in fetched:
        key = _ps_key(row.get("process_sheet_no") or row.get("pp_voucher_no"))
        serialized = _serialize_tracker_row(row, live=live_map.get(key), pics_by_id=pics)
        if serialized:
            serialized["history_count"] = int(history_counts.get(key) or 0)
            out.append(serialized)
    _apply_stage_overlay_to_rows(out)
    return out


def _apply_stage_overlay_to_rows(rows_out: list[dict[str, Any]]) -> None:
    ps_nos = [compact_text(row.get("process_sheet_no")) for row in rows_out if compact_text(row.get("process_sheet_no"))]
    if not ps_nos:
        return
    try:
        from .sales_orders_route import _load_stage_overlay

        overlay = _load_stage_overlay(ps_nos, live=False)
    except Exception:
        logger.exception("first article stage overlay failed")
        return
    for row in rows_out:
        stage = _stage_from_overlay(overlay, compact_text(row.get("process_sheet_no")))
        if not stage:
            continue
        if stage["current_stage_desc"] or stage["erp_stage_mode"] != "unassigned":
            row.update(stage)
            row["current_stage_status_label"] = _stage_status_label(stage["current_stage_status"])


def lookup_sales_order_job(process_sheet_no: str, pp_voucher_no: str = "") -> dict[str, Any] | None:
    live_map = _live_job_map(allow_rebuild=True)
    for raw in (process_sheet_no, pp_voucher_no):
        job = live_map.get(_ps_key(raw))
        if job:
            return job
    extra = _lookup_jobs_from_pp_cache([process_sheet_no, pp_voucher_no])
    for raw in (process_sheet_no, pp_voucher_no):
        job = extra.get(_ps_key(raw))
        if job:
            return job
    return None


def search_flag_candidates(query: str, *, limit: int = _SEARCH_LIMIT) -> list[dict[str, Any]]:
    if not compact_text(query):
        return []
    jobs = _sales_order_jobs(allow_rebuild=True)
    with planner_db() as con:
        _ensure_tables(con)
        flagged = _flagged_keys(con)
    hits = search_jobs(jobs, query, flagged_keys=flagged, limit=limit)
    cap = max(1, min(int(limit or _SEARCH_LIMIT), 50))
    if len(hits) >= cap:
        return hits
    if hits and not _looks_like_ps(query):
        return hits
    seen = {
        _ps_key(hit.get("process_sheet_no") or hit.get("pp_voucher_no"))
        for hit in hits
    }
    for extra in _search_jobs_from_pp_cache(query, flagged_keys=flagged, limit=limit):
        key = _ps_key(extra.get("process_sheet_no") or extra.get("pp_voucher_no"))
        if not key or key in seen:
            continue
        seen.add(key)
        hits.append(extra)
        if len(hits) >= cap:
            break
    return hits


def _upsert_flagged_row(
    con,
    process_sheet_no: str,
    pp_voucher_no: str,
    *,
    machine_codes: list[str] | None = None,
) -> tuple[Any, bool]:
    existing = one(
        con.execute(
            f"""
            SELECT {_ROW_SELECT}
            FROM planner_first_article
            WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (process_sheet_no,),
        )
    )
    if existing:
        row = existing
        if pp_voucher_no and compact_text(row.get("pp_voucher_no")) != pp_voucher_no:
            row = one(
                con.execute(
                    f"""
                    UPDATE planner_first_article
                    SET pp_voucher_no = %s, updated_at = NOW()
                    WHERE first_article_id = %s
                    RETURNING {_ROW_SELECT}
                    """,
                    (pp_voucher_no, int(row["first_article_id"])),
                )
            ) or existing
        return row, False
    row = one(
        con.execute(
            f"""
            INSERT INTO planner_first_article (process_sheet_no, pp_voucher_no, machine_codes, updated_at)
            VALUES (%s, %s, %s, NOW())
            RETURNING {_ROW_SELECT}
            """,
            (process_sheet_no, pp_voucher_no, machine_codes),
        )
    )
    return row, True


def _normalize_flag_item(raw: Any) -> tuple[str, str] | None:
    if isinstance(raw, str):
        process_sheet_no = _ps_base(raw)
        pp_voucher_no = ""
    elif isinstance(raw, dict):
        process_sheet_no = _ps_base(raw.get("process_sheet_no") or raw.get("pp_voucher_no"))
        pp_voucher_no = compact_text(raw.get("pp_voucher_no"))
    else:
        return None
    if not process_sheet_no:
        return None
    return process_sheet_no, pp_voucher_no


def flag_process_sheet(data: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    parsed = _normalize_flag_item(data if isinstance(data, dict) else {})
    if not parsed:
        raise ValueError("process_sheet_no is required")
    process_sheet_no, pp_voucher_no = parsed
    live = lookup_sales_order_job(process_sheet_no, pp_voucher_no)
    if live:
        process_sheet_no = compact_text(live.get("process_sheet_no")) or process_sheet_no
        pp_voucher_no = compact_text(live.get("pp_voucher_no")) or pp_voucher_no
    seed_machines = _parse_machine_codes((live or {}).get("queued_machines") or (live or {}).get("machine_cnc"))

    with planner_db() as con:
        _ensure_tables(con)
        row, created = _upsert_flagged_row(
            con,
            process_sheet_no,
            pp_voucher_no,
            machine_codes=seed_machines or None,
        )
        pics = _pics_by_id(con)
    serialized = _serialize_tracker_row(row, live=live or {}, pics_by_id=pics)
    if not serialized:
        raise RuntimeError("Failed to flag process sheet")
    return serialized, created


def flag_process_sheets(items: list[Any] | None) -> dict[str, Any]:
    parsed: list[tuple[str, str]] = []
    seen: set[str] = set()
    for raw in items or []:
        item = _normalize_flag_item(raw)
        if not item:
            continue
        key = _ps_key(item[0])
        if key in seen:
            continue
        seen.add(key)
        parsed.append(item)
    if not parsed:
        raise ValueError("Select at least one process sheet")
    if len(parsed) > _BULK_FLAG_LIMIT:
        raise ValueError(f"Select at most {_BULK_FLAG_LIMIT} process sheets")

    live_map = _live_job_map(allow_rebuild=True)
    created_rows: list[dict[str, Any]] = []
    already: list[dict[str, Any]] = []
    with planner_db() as con:
        _ensure_tables(con)
        pics = _pics_by_id(con)
        for process_sheet_no, pp_voucher_no in parsed:
            live = (
                live_map.get(_ps_key(process_sheet_no))
                or live_map.get(_ps_key(pp_voucher_no))
                or {}
            )
            if live:
                process_sheet_no = compact_text(live.get("process_sheet_no")) or process_sheet_no
                pp_voucher_no = compact_text(live.get("pp_voucher_no")) or pp_voucher_no
            seed_machines = _parse_machine_codes(live.get("queued_machines") or live.get("machine_cnc"))
            row, created = _upsert_flagged_row(
                con,
                process_sheet_no,
                pp_voucher_no,
                machine_codes=seed_machines or None,
            )
            serialized = _serialize_tracker_row(row, live=live or {}, pics_by_id=pics)
            if not serialized:
                continue
            if created:
                created_rows.append(serialized)
            else:
                already.append(serialized)
    return {
        "created": created_rows,
        "already_flagged": already,
        "created_count": len(created_rows),
        "already_flagged_count": len(already),
        "count": len(created_rows) + len(already),
    }


def _apply_tracker_patch(
    con,
    current: dict[str, Any],
    data: dict[str, Any],
    *,
    catalog: list[str] | None = None,
) -> dict[str, Any]:
    if "pic_names" in data:
        current["pic_ids"] = _resolve_pic_names(con, _parse_pic_names(data.get("pic_names")))
    elif "pic_ids" in data:
        current["pic_ids"] = _validate_pic_ids(con, _parse_pic_ids(data.get("pic_ids")))
    if "machine_codes" in data:
        current["machine_codes"] = _resolve_machine_codes(data.get("machine_codes"), catalog)
    if "remarks" in data:
        current["remarks"] = compact_text(data.get("remarks"))
    for prefix in CHECK_TEXT_FIELDS:
        if prefix in data:
            parsed = _parse_check_cell(data.get(prefix))
            current[f"{prefix}_mode"] = parsed["mode"]
            current[f"{prefix}_tick"] = parsed["tick"]
            current[f"{prefix}_text"] = parsed["text"]
            continue
        mode_key = f"{prefix}_mode"
        tick_key = f"{prefix}_tick"
        text_key = f"{prefix}_text"
        if mode_key in data:
            current[mode_key] = _parse_mode(data.get(mode_key), field=mode_key)
        if tick_key in data:
            current[tick_key] = bool(data.get(tick_key))
        if text_key in data:
            current[text_key] = compact_text(data.get(text_key))
    return current


def _write_tracker_row(con, first_article_id: int, current: dict[str, Any]):
    return one(
        con.execute(
            f"""
            UPDATE planner_first_article
            SET pic_ids = %s,
                machine_codes = %s,
                tooling_mode = %s,
                tooling_tick = %s,
                tooling_text = %s,
                fixture_mode = %s,
                fixture_tick = %s,
                fixture_text = %s,
                gauges_mode = %s,
                gauges_tick = %s,
                gauges_text = %s,
                remarks = %s,
                updated_at = NOW()
            WHERE first_article_id = %s
            RETURNING {_ROW_SELECT}
            """,
            (
                current.get("pic_ids") or [],
                current.get("machine_codes"),
                compact_text(current.get("tooling_mode")) or "tick",
                bool(current.get("tooling_tick")),
                compact_text(current.get("tooling_text")),
                compact_text(current.get("fixture_mode")) or "tick",
                bool(current.get("fixture_tick")),
                compact_text(current.get("fixture_text")),
                compact_text(current.get("gauges_mode")) or "tick",
                bool(current.get("gauges_tick")),
                compact_text(current.get("gauges_text")),
                compact_text(current.get("remarks")),
                int(first_article_id),
            ),
        )
    )


def import_tracker_rows(items: list[Any] | None) -> dict[str, Any]:
    parsed: list[dict[str, Any]] = []
    last_by_key: dict[str, dict[str, Any]] = {}
    duplicate_count = 0
    for raw in items or []:
        if not isinstance(raw, dict):
            continue
        process_sheet_no = _ps_base(raw.get("process_sheet_no") or raw.get("pp_voucher_no"))
        if not process_sheet_no:
            continue
        key = _ps_key(process_sheet_no)
        item = {
            "process_sheet_no": process_sheet_no,
            "pp_voucher_no": compact_text(raw.get("pp_voucher_no")),
            "patch": dict(raw.get("patch") or {}),
            "source_row": raw.get("source_row"),
        }
        if key in last_by_key:
            duplicate_count += 1
        last_by_key[key] = item
    parsed = list(last_by_key.values())
    if not parsed:
        raise ValueError("No process sheet numbers found in the Excel file")
    if len(parsed) > _IMPORT_LIMIT:
        raise ValueError(f"Import at most {_IMPORT_LIMIT} process sheets")

    live_map = _live_job_map(allow_rebuild=True)
    cache_map = _lookup_jobs_from_pp_cache([item["process_sheet_no"] for item in parsed])
    for key, job in cache_map.items():
        live_map[key] = _merge_live_job(live_map.get(key), job) or job
    catalog = load_machine_catalog()

    created_rows: list[dict[str, Any]] = []
    updated_rows: list[dict[str, Any]] = []
    missing_erp: list[str] = []
    errors: list[dict[str, Any]] = []
    with planner_db() as con:
        _ensure_tables(con)
        pics = _pics_by_id(con)
        for item in parsed:
            process_sheet_no = item["process_sheet_no"]
            pp_voucher_no = item["pp_voucher_no"]
            try:
                live = live_map.get(_ps_key(process_sheet_no)) or live_map.get(_ps_key(pp_voucher_no)) or {}
                if live:
                    process_sheet_no = compact_text(live.get("process_sheet_no")) or process_sheet_no
                    pp_voucher_no = compact_text(live.get("pp_voucher_no")) or pp_voucher_no
                else:
                    missing_erp.append(process_sheet_no)
                patch = dict(item.get("patch") or {})
                if "machine_codes" in patch:
                    seed_machines = _resolve_machine_codes(patch.get("machine_codes"), catalog)
                else:
                    seed_machines = _parse_machine_codes(live.get("queued_machines") or live.get("machine_cnc"))
                row, created = _upsert_flagged_row(
                    con,
                    process_sheet_no,
                    pp_voucher_no,
                    machine_codes=seed_machines or None,
                )
                current = dict(row or {})
                if patch:
                    before = dict(current)
                    current = _apply_tracker_patch(con, current, patch, catalog=catalog)
                    if "pic_names" in patch:
                        pics = _pics_by_id(con)
                    changes = diff_tracked_fields(
                        before,
                        current,
                        _HISTORY_FIELDS_FLAGGED,
                        pics_by_id=pics,
                    )
                    row = _write_tracker_row(con, int(current["first_article_id"]), current) or current
                    _insert_change_logs(
                        con,
                        source="flagged",
                        process_sheet_no=process_sheet_no,
                        first_article_id=int(current["first_article_id"]),
                        changes=changes,
                    )
                serialized = _serialize_tracker_row(row, live=live or {}, pics_by_id=pics)
                if not serialized:
                    continue
                if created:
                    created_rows.append(serialized)
                else:
                    updated_rows.append(serialized)
            except Exception as exc:
                logger.exception("first article import row failed")
                errors.append({
                    "process_sheet_no": process_sheet_no,
                    "source_row": item.get("source_row"),
                    "error": str(exc),
                })
        pics = _pics_by_id(con)
    return {
        "created": created_rows,
        "updated": updated_rows,
        "created_count": len(created_rows),
        "updated_count": len(updated_rows),
        "missing_erp": missing_erp,
        "missing_erp_count": len(missing_erp),
        "duplicate_count": duplicate_count,
        "error_count": len(errors),
        "errors": errors,
        "count": len(created_rows) + len(updated_rows),
        "pics": list(pics.values()),
    }


def update_tracker_row(first_article_id: int, data: dict[str, Any]) -> dict[str, Any] | None:
    with planner_db() as con:
        _ensure_tables(con)
        existing = one(
            con.execute(
                f"""
                SELECT {_ROW_SELECT}
                FROM planner_first_article
                WHERE first_article_id = %s
                """,
                (int(first_article_id),),
            )
        )
        if not existing:
            return None
        current = _apply_tracker_patch(
            con,
            dict(existing),
            data,
            catalog=load_machine_catalog() if "machine_codes" in data else None,
        )
        pics = _pics_by_id(con)
        changes = diff_tracked_fields(
            dict(existing),
            current,
            _HISTORY_FIELDS_FLAGGED,
            pics_by_id=pics,
        )
        row = _write_tracker_row(con, int(first_article_id), current)
        _insert_change_logs(
            con,
            source="flagged",
            process_sheet_no=compact_text((row or current).get("process_sheet_no")),
            first_article_id=int(first_article_id),
            changes=changes,
        )
        history_count = int(
            _history_count_map(
                con,
                "flagged",
                [compact_text((row or current).get("process_sheet_no"))],
            ).get(_ps_key((row or current).get("process_sheet_no"))) or 0
        )
    live = lookup_sales_order_job(
        compact_text((row or {}).get("process_sheet_no")),
        compact_text((row or {}).get("pp_voucher_no")),
    )
    serialized = _serialize_tracker_row(row, live=live or {}, pics_by_id=pics)
    if serialized:
        serialized["history_count"] = history_count
        _apply_stage_overlay_to_rows([serialized])
    return serialized


def unflag_process_sheet(first_article_id: int) -> bool:
    with planner_db() as con:
        _ensure_tables(con)
        row = one(
            con.execute(
                """
                DELETE FROM planner_first_article
                WHERE first_article_id = %s
                RETURNING first_article_id
                """,
                (int(first_article_id),),
            )
        )
    return bool(row)


def _serialize_new_part_saved(
    row: dict[str, Any] | None,
    *,
    pics_by_id: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not row:
        return {
            "bom_updated": False,
            "remarks": "",
            "program_finish_at": "",
            "program_pic_ids": [],
            "program_pics": [],
            "is_exception": False,
        }
    out = serialize_row(dict(row))
    pic_ids = _parse_pic_ids(out.get("program_pic_ids"))
    return {
        "bom_updated": bool(out.get("bom_updated")),
        "remarks": compact_text(out.get("remarks")),
        "program_finish_at": compact_text(out.get("program_finish_at")),
        "program_pic_ids": pic_ids,
        "program_pics": _pics_for_ids(pics_by_id or {}, pic_ids),
        "is_exception": bool(out.get("is_exception")),
    }


def _blank_new_part_job(process_sheet_no: str, pp_voucher_no: str = "") -> dict[str, Any]:
    job = {
        "process_sheet_no": process_sheet_no,
        "pp_voucher_no": compact_text(pp_voucher_no),
        "part_no": "",
        "part_description": "",
        "total_qty": None,
        "po_due_date": "",
        "posted_date": "",
        "queued_machines": [],
        "machine_cnc": "",
        "coway_proposed_edd": "",
        "sales_order_no": "",
        "customer_name": "",
        "is_new_part": False,
        "bom_code": "",
        "has_bom": False,
        "material_subcon": "",
        "material_arrived": False,
        "material_date": "",
        "material_legacy": "",
        "material_display": "",
        "so_scope": "",
        "shipped_completed": False,
        "from_erp_cache": False,
        **_empty_stage(),
    }
    job["ps_type"] = _job_ps_type(job) or "OTHER"
    return job


def _is_complete_status(row: dict[str, Any] | None) -> bool:
    item = row or {}
    if bool(item.get("shipped_completed")):
        return True
    if compact_text(item.get("so_scope")).lower() == "complete":
        return True
    return compact_text(item.get("erp_stage_mode")).lower() == "completed"


def _merge_new_part_row(
    job: dict[str, Any],
    saved: dict[str, Any] | None,
    *,
    pics_by_id: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    out = dict(job)
    tracker = _serialize_new_part_saved(saved, pics_by_id=pics_by_id)
    out.update(tracker)
    if job.get("is_new_part"):
        out["in_sales_orders"] = True
    elif job.get("from_erp_cache"):
        out["in_sales_orders"] = False
    else:
        out["in_sales_orders"] = compact_text(job.get("so_scope")).lower() in {"active", "complete"}
    if not compact_text(out.get("ps_type")):
        out["ps_type"] = _job_ps_type(out) or "OTHER"
    _decorate_stage_labels(out)
    out["history_count"] = int(out.get("history_count") or 0)
    out["list_scope"] = "history" if _is_complete_status(out) else "active"
    return out


def _new_part_tracker_map(con) -> dict[str, dict[str, Any]]:
    fetched = rows(
        con.execute(
            f"""
            SELECT {_NEW_PART_ROW_SELECT}
            FROM planner_first_article_new_part
            """
        )
    )
    out: dict[str, dict[str, Any]] = {}
    for row in fetched:
        key = _ps_key(row.get("process_sheet_no") or row.get("pp_voucher_no"))
        if key:
            out[key] = dict(row)
    return out


def _live_jobs_for_new_parts(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]]]:
    active_orders = list(payload.get("active") or [])
    complete_orders = list(payload.get("complete") or [])
    _ensure_new_part_flags(active_orders)
    _ensure_new_part_flags(complete_orders)
    active_jobs = flatten_sales_order_jobs(active_orders, so_scope="active")
    live_map = index_jobs_by_ps(active_jobs)
    complete_jobs = flatten_sales_order_jobs(complete_orders, so_scope="complete")
    for job in complete_jobs:
        key = _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
        if key and key not in live_map:
            live_map[key] = job
    new_active = [job for job in active_jobs if job and job.get("is_new_part")]
    new_complete = [job for job in complete_jobs if job and job.get("is_new_part")]
    return new_active, new_complete, live_map


def _remember_new_part_keys(con, jobs: list[dict[str, Any]], *, already: set[str] | None = None) -> None:
    known = already or set()
    seen: set[str] = set()
    for job in jobs or []:
        process_sheet_no = _ps_base(job.get("process_sheet_no") or job.get("pp_voucher_no"))
        key = _ps_key(process_sheet_no)
        if not process_sheet_no or key in seen or key in known:
            continue
        seen.add(key)
        con.execute(
            """
            INSERT INTO planner_first_article_new_part (process_sheet_no, pp_voucher_no)
            VALUES (%s, %s)
            ON CONFLICT (process_sheet_no) DO UPDATE
            SET pp_voucher_no = CASE
                    WHEN EXCLUDED.pp_voucher_no <> '' THEN EXCLUDED.pp_voucher_no
                    ELSE planner_first_article_new_part.pp_voucher_no
                END
            """,
            (process_sheet_no, compact_text(job.get("pp_voucher_no"))),
        )


def list_new_part_rows(*, allow_rebuild: bool = True, scope: str = "active") -> list[dict[str, Any]]:
    wanted = compact_text(scope).lower() or "active"
    if wanted not in {"active", "history"}:
        wanted = "active"
    payload = _sales_order_payload(allow_rebuild=allow_rebuild)
    new_active, new_complete, live_map = _live_jobs_for_new_parts(payload)
    with planner_db() as con:
        _ensure_tables(con)
        pics = _pics_by_id(con)
        saved_map = _new_part_tracker_map(con)
        count_keys = sorted({
            key
            for key in (
                _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
                for job in (*new_active, *new_complete)
            )
            if key
        } | set(saved_map.keys()))
        history_counts = _history_count_map(con, "new_part", count_keys)

    assembled: list[tuple[dict[str, Any], dict[str, Any] | None]] = []
    seen: set[str] = set()

    def _add(job: dict[str, Any], saved: dict[str, Any] | None) -> None:
        key = _ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))
        if not key or key in seen:
            return
        seen.add(key)
        assembled.append((job, saved))

    for job in new_active:
        _add(job, saved_map.get(_ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))))
    for job in new_complete:
        _add(job, saved_map.get(_ps_key(job.get("process_sheet_no") or job.get("pp_voucher_no"))))

    leftover = [key for key in saved_map if key and key not in seen]
    cache_map = _lookup_jobs_from_pp_cache(leftover) if leftover else {}
    for key in leftover:
        saved = saved_map.get(key) or {}
        job = (
            live_map.get(key)
            or cache_map.get(key)
            or _blank_new_part_job(
                compact_text(saved.get("process_sheet_no")) or key,
                compact_text(saved.get("pp_voucher_no")),
            )
        )
        _add(job, saved)

    _apply_has_bom([job for job, _saved in assembled])
    out = []
    for job, saved in assembled:
        merged = _merge_new_part_row(job, saved, pics_by_id=pics)
        key = _ps_key(merged.get("process_sheet_no") or merged.get("pp_voucher_no"))
        merged["history_count"] = int(history_counts.get(key) or 0)
        out.append(merged)
    _apply_stage_overlay_to_rows(out)
    for row in out:
        complete = _is_complete_status(row)
        live_new = bool(row.get("is_new_part")) and compact_text(row.get("so_scope")).lower() != "complete"
        if complete:
            row["list_scope"] = "history"
        elif live_new or row.get("is_exception"):
            row["list_scope"] = "active"
        else:
            row["list_scope"] = "history"

    remembered = [row for row in out if row.get("list_scope") == "history"]
    missing = [
        row
        for row in remembered
        if _ps_key(row.get("process_sheet_no") or row.get("pp_voucher_no")) not in saved_map
    ]
    if missing:
        try:
            with planner_db() as con:
                _ensure_tables(con)
                _remember_new_part_keys(con, missing, already=set(saved_map))
        except Exception:
            logger.exception("first article history remember failed")

    scoped = [row for row in out if compact_text(row.get("list_scope")) == wanted]
    scoped.sort(
        key=lambda row: (
            compact_text(row.get("posted_date")),
            compact_text(row.get("process_sheet_no")).upper(),
        ),
        reverse=True,
    )
    return scoped


def update_new_part_row(data: dict[str, Any]) -> dict[str, Any]:
    process_sheet_no = _ps_base(data.get("process_sheet_no") or data.get("pp_voucher_no"))
    if not process_sheet_no:
        raise ValueError("process_sheet_no is required")
    pp_voucher_no = compact_text(data.get("pp_voucher_no"))
    live = lookup_sales_order_job(process_sheet_no, pp_voucher_no) or {}
    if live:
        process_sheet_no = compact_text(live.get("process_sheet_no")) or process_sheet_no
        pp_voucher_no = compact_text(live.get("pp_voucher_no")) or pp_voucher_no

    with planner_db() as con:
        _ensure_tables(con)
        existing = one(
            con.execute(
                f"""
                SELECT {_NEW_PART_ROW_SELECT}
                FROM planner_first_article_new_part
                WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
                LIMIT 1
                """,
                (process_sheet_no,),
            )
        )
        current = dict(existing) if existing else {
            "process_sheet_no": process_sheet_no,
            "pp_voucher_no": pp_voucher_no,
            "bom_updated": False,
            "remarks": "",
            "program_finish_at": "",
            "program_pic_ids": [],
            "is_exception": False,
        }
        if pp_voucher_no:
            current["pp_voucher_no"] = pp_voucher_no
        if "bom_updated" in data:
            current["bom_updated"] = _parse_bool(data.get("bom_updated"), field="bom_updated")
        if "remarks" in data:
            current["remarks"] = compact_text(data.get("remarks"))
        if "program_finish_at" in data:
            current["program_finish_at"] = _parse_program_finish(data.get("program_finish_at"))
        if "program_pic_ids" in data:
            current["program_pic_ids"] = _validate_pic_ids(con, _parse_pic_ids(data.get("program_pic_ids")))
        pics = _pics_by_id(con)
        changes = diff_tracked_fields(
            dict(existing) if existing else {
                "remarks": "",
                "program_finish_at": "",
                "program_pic_ids": [],
            },
            current,
            _HISTORY_FIELDS_NEW_PART,
            pics_by_id=pics,
        )
        row = one(
            con.execute(
                f"""
                INSERT INTO planner_first_article_new_part (
                    process_sheet_no, pp_voucher_no, bom_updated, remarks,
                    program_finish_at, program_pic_ids, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (process_sheet_no) DO UPDATE
                SET pp_voucher_no = EXCLUDED.pp_voucher_no,
                    bom_updated = EXCLUDED.bom_updated,
                    remarks = EXCLUDED.remarks,
                    program_finish_at = EXCLUDED.program_finish_at,
                    program_pic_ids = EXCLUDED.program_pic_ids,
                    updated_at = NOW()
                RETURNING {_NEW_PART_ROW_SELECT}
                """,
                (
                    process_sheet_no,
                    compact_text(current.get("pp_voucher_no")),
                    bool(current.get("bom_updated")),
                    compact_text(current.get("remarks")),
                    compact_text(current.get("program_finish_at")),
                    current.get("program_pic_ids") or [],
                ),
            )
        )
        _insert_change_logs(
            con,
            source="new_part",
            process_sheet_no=process_sheet_no,
            first_article_id=None,
            changes=changes,
        )
        history_count = int(
            _history_count_map(con, "new_part", [process_sheet_no]).get(_ps_key(process_sheet_no)) or 0
        )
    job = live or _blank_new_part_job(process_sheet_no, compact_text(current.get("pp_voucher_no")))
    if not live:
        job["is_new_part"] = not bool(current.get("is_exception"))
    _apply_has_bom([job])
    merged = _merge_new_part_row(job, dict(row) if row else current, pics_by_id=pics)
    merged["history_count"] = history_count
    _apply_stage_overlay_to_rows([merged])
    complete = _is_complete_status(merged)
    live_new = bool(merged.get("is_new_part")) and compact_text(merged.get("so_scope")).lower() != "complete"
    if complete:
        merged["list_scope"] = "history"
    elif live_new or merged.get("is_exception"):
        merged["list_scope"] = "active"
    else:
        merged["list_scope"] = "history"
    return merged


def _upsert_new_part_exception(
    con,
    process_sheet_no: str,
    pp_voucher_no: str,
) -> tuple[Any, bool]:
    existing = one(
        con.execute(
            f"""
            SELECT {_NEW_PART_ROW_SELECT}
            FROM planner_first_article_new_part
            WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
            LIMIT 1
            """,
            (process_sheet_no,),
        )
    )
    if existing and existing.get("is_exception"):
        row = existing
        if pp_voucher_no and compact_text(row.get("pp_voucher_no")) != pp_voucher_no:
            row = one(
                con.execute(
                    f"""
                    UPDATE planner_first_article_new_part
                    SET pp_voucher_no = %s, updated_at = NOW()
                    WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
                    RETURNING {_NEW_PART_ROW_SELECT}
                    """,
                    (pp_voucher_no, process_sheet_no),
                )
            ) or existing
        return row, False
    row = one(
        con.execute(
            f"""
            INSERT INTO planner_first_article_new_part (
                process_sheet_no, pp_voucher_no, is_exception, updated_at
            )
            VALUES (%s, %s, TRUE, NOW())
            ON CONFLICT (process_sheet_no) DO UPDATE
            SET pp_voucher_no = CASE
                    WHEN EXCLUDED.pp_voucher_no <> '' THEN EXCLUDED.pp_voucher_no
                    ELSE planner_first_article_new_part.pp_voucher_no
                END,
                is_exception = TRUE,
                updated_at = NOW()
            RETURNING {_NEW_PART_ROW_SELECT}
            """,
            (process_sheet_no, pp_voucher_no),
        )
    )
    return row, True


def add_new_part_exception(data: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    parsed = _normalize_flag_item(data if isinstance(data, dict) else {})
    if not parsed:
        raise ValueError("process_sheet_no is required")
    process_sheet_no, pp_voucher_no = parsed
    live = lookup_sales_order_job(process_sheet_no, pp_voucher_no) or {}
    if live:
        process_sheet_no = compact_text(live.get("process_sheet_no")) or process_sheet_no
        pp_voucher_no = compact_text(live.get("pp_voucher_no")) or pp_voucher_no

    already_new = bool(
        live.get("is_new_part")
        and compact_text(live.get("so_scope")).lower() != "complete"
    )
    with planner_db() as con:
        _ensure_tables(con)
        pics = _pics_by_id(con)
        if already_new:
            saved = one(
                con.execute(
                    f"""
                    SELECT {_NEW_PART_ROW_SELECT}
                    FROM planner_first_article_new_part
                    WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
                    LIMIT 1
                    """,
                    (process_sheet_no,),
                )
            )
            job = live
            _apply_has_bom([job])
            serialized = _merge_new_part_row(job, dict(saved) if saved else None, pics_by_id=pics)
            return serialized, False
        row, created = _upsert_new_part_exception(con, process_sheet_no, pp_voucher_no)
    job = live or _blank_new_part_job(process_sheet_no, pp_voucher_no)
    _apply_has_bom([job])
    serialized = _merge_new_part_row(job, dict(row) if row else None, pics_by_id=pics)
    if not serialized:
        raise RuntimeError("Failed to add new-part exception")
    return serialized, created


def remove_new_part_exception(process_sheet_no: str) -> dict[str, Any] | None:
    key = _ps_base(process_sheet_no)
    if not key:
        raise ValueError("process_sheet_no is required")
    with planner_db() as con:
        _ensure_tables(con)
        existing = one(
            con.execute(
                f"""
                SELECT {_NEW_PART_ROW_SELECT}
                FROM planner_first_article_new_part
                WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
                LIMIT 1
                """,
                (key,),
            )
        )
        if not existing or not existing.get("is_exception"):
            return None
        row = one(
            con.execute(
                f"""
                UPDATE planner_first_article_new_part
                SET is_exception = FALSE, updated_at = NOW()
                WHERE LOWER(TRIM(process_sheet_no)) = LOWER(TRIM(%s))
                RETURNING {_NEW_PART_ROW_SELECT}
                """,
                (key,),
            )
        )
        pics = _pics_by_id(con)
    live = lookup_sales_order_job(
        compact_text((row or existing).get("process_sheet_no")),
        compact_text((row or existing).get("pp_voucher_no")),
    ) or {}
    still_on_list = bool(live.get("is_new_part") and compact_text(live.get("so_scope")).lower() != "complete")
    job = live or _blank_new_part_job(
        compact_text((row or existing).get("process_sheet_no")) or key,
        compact_text((row or existing).get("pp_voucher_no")),
    )
    if still_on_list:
        _apply_has_bom([job])
        merged = _merge_new_part_row(job, dict(row) if row else dict(existing), pics_by_id=pics)
    else:
        merged = None
    return {
        "process_sheet_no": compact_text((row or existing).get("process_sheet_no")) or key,
        "still_on_list": still_on_list,
        "row": merged,
    }


def json_error(exc: Exception, *, fallback_status: int = 500):
    friendly = planner_db_connect_error(exc)
    if friendly:
        return {"error": friendly}, 503
    text = str(exc) or exc.__class__.__name__
    if "deadlock detected" in text.lower():
        return {"error": "The first article tracker is busy. Refresh and try again."}, 503
    return {"error": text}, fallback_status
