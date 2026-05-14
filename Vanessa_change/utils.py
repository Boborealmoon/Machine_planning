from __future__ import annotations

import re
from datetime import date, datetime

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None


def compact_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_number(value, default=0.0):
    try:
        if value is None or value == "":
            return float(default)
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def parse_nullable_number(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_date_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none", "null"}:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def normalize_column_name(value):
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def normalize_sheet_name(value):
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def first_nonempty(row, *keys):
    for key in keys:
        value = compact_text(row.get(key))
        if value:
            return value
    return ""


def workbook_partial_qty(row, total_key="total_qty", partial_key="partial_qty"):
    total = parse_number(row.get(total_key), 0)
    partial = parse_number(row.get(partial_key), 0)
    if partial > 0:
        return partial
    return total


def excel_sheet_records(ws):
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [normalize_column_name(col) for col in header_row]
    records = []
    for row in rows_iter:
        if row is None or not any(cell not in (None, "") for cell in row):
            continue
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else None
        records.append(record)
    return records


def load_trial_workbook(file_storage):
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Run pip install openpyxl first.")
    workbook = load_workbook(file_storage, data_only=True, read_only=True)
    sheet_map = {normalize_sheet_name(name): name for name in workbook.sheetnames}
    required = ["ProcessSheets", "BOM_op_stage"]
    missing = [sheet for sheet in required if normalize_sheet_name(sheet) not in sheet_map]
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing)}")
    return {
        "ProcessSheets": excel_sheet_records(workbook[sheet_map[normalize_sheet_name("ProcessSheets")]]),
        "BOM_op_stage": excel_sheet_records(workbook[sheet_map[normalize_sheet_name("BOM_op_stage")]]),
        "Material_Per_BOM": excel_sheet_records(
            workbook[sheet_map[normalize_sheet_name("Material (Per BOM)")]]
        )
        if normalize_sheet_name("Material (Per BOM)") in sheet_map
        else [],
    }


def load_trial_active_sheet(file_storage):
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Run pip install openpyxl first.")
    workbook = load_workbook(file_storage, data_only=True, read_only=True)
    sheet_map = {normalize_sheet_name(name): name for name in workbook.sheetnames}
    matched = next((sheet_map[key] for key in ("activeorders", "activeorder", "active") if key in sheet_map), None)
    if not matched:
        raise ValueError("Active sheet workbook must contain a sheet named Active Orders / ActiveOrders.")
    return excel_sheet_records(workbook[matched])


def validate_cycle_minutes(total_qty, scheduled_qty, cycle_minutes):
    qty = max(parse_number(total_qty, 0), parse_number(scheduled_qty, 0))
    cycle = parse_number(cycle_minutes, 0)
    if qty > 0 and cycle <= 0:
        return "Cycle Minutes / Qty must be greater than 0 when quantity is greater than 0"
    return ""


def format_qty(value):
    if value is None:
        return ""
    value = float(value)
    if value.is_integer():
        return str(int(value))
    return str(value)


def output_diff_against_target(output_qty, target_qty):
    if output_qty is None:
        return 0.0
    return float(output_qty) - float(target_qty or 0)


def normalize_block_status_inputs(data, default_planning="PLANNED", default_execution="NOT_STARTED"):
    planning_status = compact_text(data.get("planning_status")) or default_planning
    execution_status = compact_text(data.get("execution_status")) or default_execution
    legacy_status = compact_text(data.get("status"))
    if legacy_status:
        if legacy_status.upper() in {"NOT_STARTED", "IN_PROGRESS", "DONE"}:
            execution_status = legacy_status
        elif legacy_status.upper() in {"UNPLANNED", "PARTIALLY_PLANNED", "PLANNED"}:
            planning_status = legacy_status
    return planning_status, execution_status


def combine_date_time(d: date, hhmm: str):
    hour, minute = [int(part) for part in hhmm.split(":", 1)]
    return datetime(d.year, d.month, d.day, hour, minute, 0)


def trial_process_sheet_completed(row):
    return compact_text(row.get("status")).upper() == "COMPLETED" or compact_text(row.get("planner_status")).upper() == "COMPLETED"
